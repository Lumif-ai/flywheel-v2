"""Broker module API endpoints.

Gated by require_module("broker") — returns 403 for non-broker tenants.
"""

from __future__ import annotations

import asyncio
import io
import re
from datetime import datetime, timezone
from typing import Any
from uuid import UUID, uuid4

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import exists, func, select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from flywheel.api.deps import get_tenant_db, require_module
from flywheel.auth.jwt import TokenPayload
from flywheel.db.models import (
    BrokerActivity,
    BrokerProject,
    CarrierConfig,
    CarrierQuote,
    Document,
    Email,
    Integration,
    ProjectCoverage,
    SubmissionDocument,
    UploadedFile,
)
from flywheel.engines.gap_detector import detect_gaps, summarize_gaps
from flywheel.engines.solicitation_drafter import draft_solicitation_email
from flywheel.engines.submission_builder import build_submission_package
from flywheel.services.email_dispatch import send_email_as_user

import logging

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/broker", tags=["broker"])

# Max file size for document upload (25 MB)
_MAX_DOC_SIZE = 25 * 1024 * 1024
_ALLOWED_DOC_TYPES = {"application/pdf", "image/png", "image/jpeg"}


# ---------------------------------------------------------------------------
# Pydantic request bodies
# ---------------------------------------------------------------------------


class CreateProjectBody(BaseModel):
    name: str
    project_type: str | None = None
    description: str | None = None
    contract_value: float | None = None
    location: str | None = None


class CreateFromEmailBody(BaseModel):
    message_id: str
    integration_id: UUID
    attachment_index: int = 0


class UpdateCoverageBody(BaseModel):
    coverage_type: str | None = None
    display_name: str | None = None
    required_limit: float | None = None
    required_deductible: float | None = None
    required_terms: str | None = None
    contract_clause: str | None = None
    gap_notes: str | None = None


class CreateCarrierBody(BaseModel):
    carrier_name: str
    carrier_type: str = "insurance"
    submission_method: str = "email"
    portal_url: str | None = None
    email_address: str | None = None
    coverage_types: list[str] = []
    regions: list[str] = []
    min_project_value: float | None = None
    max_project_value: float | None = None
    avg_response_days: float | None = None
    portal_limit: float | None = None
    notes: str | None = None


class UpdateCarrierBody(BaseModel):
    carrier_name: str | None = None
    carrier_type: str | None = None
    submission_method: str | None = None
    portal_url: str | None = None
    email_address: str | None = None
    coverage_types: list[str] | None = None
    regions: list[str] | None = None
    min_project_value: float | None = None
    max_project_value: float | None = None
    avg_response_days: float | None = None
    portal_limit: float | None = None
    is_active: bool | None = None
    notes: str | None = None


class ExportComparisonBody(BaseModel):
    quote_ids: list[UUID] | None = None


class DraftSolicitationsBody(BaseModel):
    carrier_config_ids: list[UUID]


class EditDraftBody(BaseModel):
    draft_subject: str | None = None
    draft_body: str | None = None


class SolicitationDraft(BaseModel):
    quote_id: UUID
    carrier_name: str
    carrier_config_id: UUID
    submission_method: str
    draft_subject: str | None
    draft_body: str | None
    draft_status: str | None
    documents: list[dict]


class DraftSolicitationsResponse(BaseModel):
    drafts: list[SolicitationDraft]
    portal_submissions: list[dict]  # carriers needing portal submission
    skipped: list[dict]  # carriers skipped with reason


# Simple email regex for validation
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


# ---------------------------------------------------------------------------
# Status transition enforcement
# ---------------------------------------------------------------------------

ALLOWED_TRANSITIONS: dict[str, set[str]] = {
    "new_request": {"analyzing", "cancelled"},
    "analyzing": {"analysis_failed", "gaps_identified", "cancelled"},
    "analysis_failed": {"analyzing", "cancelled"},
    "gaps_identified": {"soliciting", "cancelled"},
    "soliciting": {"quotes_partial", "quotes_complete", "cancelled"},
    "quotes_partial": {"quotes_complete", "cancelled"},
    "quotes_complete": {"recommended", "cancelled"},
    "recommended": {"delivered", "cancelled"},
    "delivered": {"bound", "cancelled"},
    "bound": set(),  # terminal
    "cancelled": set(),  # terminal
}


def validate_transition(current: str, target: str) -> None:
    """Raise HTTPException(409) if status transition is not allowed."""
    allowed = ALLOWED_TRANSITIONS.get(current, set())
    if target not in allowed:
        raise HTTPException(
            status_code=409,
            detail=f"Invalid status transition: '{current}' -> '{target}'. "
            f"Allowed: {sorted(allowed) if allowed else 'none (terminal state)'}",
        )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _project_to_dict(p: BrokerProject) -> dict[str, Any]:
    """Serialize a BrokerProject to a JSON-friendly dict."""
    return {
        "id": str(p.id),
        "tenant_id": str(p.tenant_id),
        "name": p.name,
        "project_type": p.project_type,
        "description": p.description,
        "contract_value": float(p.contract_value) if p.contract_value is not None else None,
        "currency": p.currency,
        "location": p.location,
        "language": p.language,
        "status": p.status,
        "approval_status": p.approval_status,
        "analysis_status": p.analysis_status,
        "source": p.import_source,
        "source_ref": p.external_ref,
        "notes": None,
        "metadata": p.metadata_,
        "recommendation_subject": p.recommendation_subject,
        "recommendation_body": p.recommendation_body,
        "recommendation_status": p.recommendation_status,
        "recommendation_sent_at": p.recommendation_sent_at.isoformat() if p.recommendation_sent_at else None,
        "recommendation_recipient": p.recommendation_recipient,
        "created_at": p.created_at.isoformat() if p.created_at else None,
        "updated_at": p.updated_at.isoformat() if p.updated_at else None,
    }


def _coverage_to_dict(c: ProjectCoverage) -> dict[str, Any]:
    """Serialize a ProjectCoverage to a JSON-friendly dict."""
    return {
        "id": str(c.id),
        "broker_project_id": str(c.broker_project_id),
        "coverage_type": c.coverage_type,
        "category": c.category,
        "display_name": c.display_name,
        "description": c.display_name,
        "language": getattr(c, 'language', None),
        "required_limit": float(c.required_limit) if c.required_limit is not None else None,
        "required_deductible": float(c.required_deductible) if c.required_deductible is not None else None,
        "required_terms": c.required_terms,
        "contract_clause": c.contract_clause,
        "current_limit": float(c.current_limit) if c.current_limit is not None else None,
        "current_carrier": c.current_carrier,
        "gap_status": c.gap_status,
        "gap_amount": float(c.gap_amount) if c.gap_amount is not None else None,
        "gap_notes": c.gap_notes,
        "source": c.source,
        "confidence": c.confidence,
        "is_manual_override": c.is_manual_override,
        "created_at": c.created_at.isoformat() if c.created_at else None,
    }


def _activity_to_dict(a: BrokerActivity) -> dict[str, Any]:
    """Serialize a BrokerActivity to a JSON-friendly dict."""
    return {
        "id": str(a.id),
        "activity_type": a.activity_type,
        "description": a.description,
        "actor_type": a.actor_type,
        "metadata": a.metadata_,
        "occurred_at": a.occurred_at.isoformat() if a.occurred_at else None,
    }


def _carrier_to_dict(c: CarrierConfig) -> dict[str, Any]:
    """Serialize a CarrierConfig to a JSON-friendly dict."""
    return {
        "id": str(c.id),
        "tenant_id": str(c.tenant_id),
        "carrier_name": c.carrier_name,
        "carrier_type": c.carrier_type,
        "submission_method": c.submission_method,
        "portal_url": c.portal_url,
        "email_address": c.email_address,
        "coverage_types": c.coverage_types or [],
        "regions": c.regions or [],
        "min_project_value": float(c.min_project_value) if c.min_project_value is not None else None,
        "max_project_value": float(c.max_project_value) if c.max_project_value is not None else None,
        "avg_response_days": float(c.avg_response_days) if c.avg_response_days is not None else None,
        "portal_limit": float(c.portal_limit) if c.portal_limit is not None else None,
        "is_active": c.is_active,
        "notes": c.notes,
        "created_at": c.created_at.isoformat() if c.created_at else None,
        "updated_at": c.updated_at.isoformat() if c.updated_at else None,
    }


# ---------------------------------------------------------------------------
# GET /broker/health
# ---------------------------------------------------------------------------


@router.get("/health")
async def broker_health(
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Health check for broker module — verifies module access."""
    return {"status": "ok"}


# ---------------------------------------------------------------------------
# GET /broker/gate-counts — Gate strip counts
# ---------------------------------------------------------------------------


@router.get("/gate-counts")
async def get_gate_counts(
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Return counts for each gate in the broker workflow strip."""
    base = (
        select(BrokerProject.id, BrokerProject.created_at)
        .where(
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
        .order_by(BrokerProject.created_at.asc())
    )

    # Review: gaps identified, not yet approved
    review_q = base.where(
        BrokerProject.status == "gaps_identified",
        BrokerProject.approval_status == "draft",
    )
    review_rows = (await db.execute(review_q)).all()

    # Approve: approved projects with at least one pending carrier draft
    has_pending_draft = exists(
        select(CarrierQuote.id).where(
            CarrierQuote.broker_project_id == BrokerProject.id,
            CarrierQuote.draft_status == "pending",
        )
    )
    approve_q = base.where(
        BrokerProject.status == "gaps_identified",
        BrokerProject.approval_status == "approved",
        has_pending_draft,
    )
    approve_rows = (await db.execute(approve_q)).all()

    # Export: quotes complete or recommended
    export_q = base.where(
        BrokerProject.status.in_(["quotes_complete", "recommended"]),
    )
    export_rows = (await db.execute(export_q)).all()

    def _gate(rows: list) -> dict[str, Any]:
        return {
            "count": len(rows),
            "oldest_project_id": str(rows[0].id) if rows else None,
        }

    return {
        "review": _gate(review_rows),
        "approve": _gate(approve_rows),
        "export": _gate(export_rows),
    }


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/approve — Gate 1 approval
# ---------------------------------------------------------------------------


@router.post("/projects/{project_id}/approve")
async def approve_project(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Approve a broker project (Gate 1). Returns 409 if already approved."""
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if project.approval_status == "approved":
        raise HTTPException(status_code=409, detail="Project already approved")

    project.approval_status = "approved"

    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project.id,
        activity_type="project_approved",
        actor_type="user",
    )
    db.add(activity)

    await db.commit()
    await db.refresh(project)
    return _project_to_dict(project)


# ---------------------------------------------------------------------------
# GET /broker/dashboard-tasks — Urgency-ordered task list
# ---------------------------------------------------------------------------


@router.get("/dashboard-tasks")
async def get_dashboard_tasks(
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Return urgency-ordered task list for broker dashboard.

    Task priority: review (1) > approve (2) > export (3) > followup (4).
    Capped at 50 tasks total.
    """
    tasks: list[dict[str, Any]] = []

    # 1. "review" tasks (priority 1) — projects needing coverage review
    review_result = await db.execute(
        select(BrokerProject)
        .where(
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.status == "gaps_identified",
            BrokerProject.approval_status == "draft",
            BrokerProject.deleted_at.is_(None),
        )
        .order_by(BrokerProject.created_at.asc())
    )
    for p in review_result.scalars().all():
        tasks.append({
            "type": "review",
            "priority": 1,
            "project_id": str(p.id),
            "project_name": p.name,
            "created_at": p.created_at.isoformat() if p.created_at else None,
            "message": "Review extracted coverages",
        })

    # 2. "approve" tasks (priority 2) — approved projects with pending drafts
    approve_result = await db.execute(
        select(BrokerProject)
        .where(
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.status == "gaps_identified",
            BrokerProject.approval_status == "approved",
            BrokerProject.deleted_at.is_(None),
            exists(
                select(CarrierQuote.id).where(
                    CarrierQuote.broker_project_id == BrokerProject.id,
                    CarrierQuote.draft_status == "pending",
                )
            ),
        )
        .order_by(BrokerProject.created_at.asc())
    )
    for p in approve_result.scalars().all():
        tasks.append({
            "type": "approve",
            "priority": 2,
            "project_id": str(p.id),
            "project_name": p.name,
            "created_at": p.created_at.isoformat() if p.created_at else None,
            "message": "Approve & send solicitations",
        })

    # 3. "export" tasks (priority 3) — projects ready for comparison export
    export_result = await db.execute(
        select(BrokerProject)
        .where(
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.status.in_(("quotes_complete", "recommended")),
            BrokerProject.deleted_at.is_(None),
        )
        .order_by(BrokerProject.created_at.asc())
    )
    for p in export_result.scalars().all():
        tasks.append({
            "type": "export",
            "priority": 3,
            "project_id": str(p.id),
            "project_name": p.name,
            "created_at": p.created_at.isoformat() if p.created_at else None,
            "message": "Export comparison for client",
        })

    # 4. "followup" tasks (priority 4) — overdue solicitations (>7 days, no response)
    followup_result = await db.execute(
        select(
            CarrierQuote.id.label("quote_id"),
            CarrierQuote.solicited_at,
            BrokerProject.id.label("project_id"),
            BrokerProject.name.label("project_name"),
            CarrierConfig.carrier_name,
        )
        .join(BrokerProject, CarrierQuote.broker_project_id == BrokerProject.id)
        .join(CarrierConfig, CarrierQuote.carrier_config_id == CarrierConfig.id)
        .where(
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
            CarrierQuote.draft_status == "sent",
            CarrierQuote.status == "pending",
            CarrierQuote.solicited_at.isnot(None),
            func.now() - CarrierQuote.solicited_at > text("interval '7 days'"),
        )
        .order_by(CarrierQuote.solicited_at.asc())
    )
    for row in followup_result.all():
        days_overdue = (
            (datetime.now(timezone.utc) - row.solicited_at).days
            if row.solicited_at
            else 0
        )
        tasks.append({
            "type": "followup",
            "priority": 4,
            "project_id": str(row.project_id),
            "project_name": row.project_name,
            "quote_id": str(row.quote_id),
            "solicited_at": row.solicited_at.isoformat() if row.solicited_at else None,
            "carrier_name": row.carrier_name,
            "days_overdue": days_overdue,
            "message": f"Follow up with {row.carrier_name}",
        })

    # Cap at 50 tasks (already in priority order from concatenation)
    tasks = tasks[:50]

    return {"tasks": tasks, "total": len(tasks)}


# ---------------------------------------------------------------------------
# POST /broker/projects — Manual project creation
# ---------------------------------------------------------------------------


@router.post("/projects", status_code=201)
async def create_project(
    body: CreateProjectBody,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Create a new broker project manually."""
    project = BrokerProject(
        tenant_id=user.tenant_id,
        name=body.name,
        project_type=body.project_type or "construction",
        description=body.description,
        contract_value=body.contract_value,
        location=body.location,
        status="new_request",
        analysis_status="pending",
        import_source="manual",
    )
    db.add(project)
    await db.flush()  # Get project.id before creating activity

    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project.id,
        activity_type="project_created",
        actor_type="user",
        metadata_={"source": "manual"},
    )
    db.add(activity)
    await db.commit()
    await db.refresh(project)

    return _project_to_dict(project)


# ---------------------------------------------------------------------------
# GET /broker/projects — Paginated list
# ---------------------------------------------------------------------------


@router.get("/projects")
async def list_projects(
    offset: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    status: str | None = Query(None),
    search: str | None = Query(None),
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """List broker projects with pagination, filtering, and search."""
    # Base filter: exclude soft-deleted
    base_filter = BrokerProject.deleted_at.is_(None)

    # Optional status filter
    filters = [base_filter]
    if status:
        filters.append(BrokerProject.status == status)
    if search:
        like_pattern = f"%{search}%"
        filters.append(BrokerProject.name.ilike(like_pattern))

    # Count query
    count_q = select(func.count(BrokerProject.id)).where(*filters)
    total = (await db.execute(count_q)).scalar() or 0

    # Data query
    data_q = (
        select(BrokerProject)
        .where(*filters)
        .order_by(BrokerProject.updated_at.desc())
        .offset(offset)
        .limit(limit)
    )
    projects = (await db.execute(data_q)).scalars().all()

    return {
        "items": [_project_to_dict(p) for p in projects],
        "total": total,
        "offset": offset,
        "limit": limit,
        "has_more": offset + limit < total,
    }


# ---------------------------------------------------------------------------
# GET /broker/projects/{project_id} — Single project with nested data
# ---------------------------------------------------------------------------


@router.get("/projects/{project_id}")
async def get_project(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Get a single project with nested coverages and activities."""
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    # Fetch coverages
    coverages_result = await db.execute(
        select(ProjectCoverage)
        .where(ProjectCoverage.broker_project_id == project_id)
        .order_by(ProjectCoverage.created_at)
    )
    coverages = coverages_result.scalars().all()

    # Fetch recent activities (limit 50, newest first)
    activities_result = await db.execute(
        select(BrokerActivity)
        .where(BrokerActivity.broker_project_id == project_id)
        .order_by(BrokerActivity.occurred_at.desc())
        .limit(50)
    )
    activities = activities_result.scalars().all()

    project_dict = _project_to_dict(project)
    project_dict["coverages"] = [_coverage_to_dict(c) for c in coverages]
    project_dict["activities"] = [_activity_to_dict(a) for a in activities]

    return project_dict


# ---------------------------------------------------------------------------
# GET /broker/projects/{project_id}/quotes — List quotes for a project
# ---------------------------------------------------------------------------


@router.get("/projects/{project_id}/quotes")
async def list_project_quotes(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> list[dict[str, Any]]:
    """List all carrier quotes for a project, newest first."""
    # Verify project exists and belongs to tenant
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    # Fetch quotes
    quotes_result = await db.execute(
        select(CarrierQuote)
        .where(
            CarrierQuote.broker_project_id == project_id,
            CarrierQuote.tenant_id == user.tenant_id,
        )
        .order_by(CarrierQuote.created_at.desc())
    )
    quotes = quotes_result.scalars().all()

    return [_quote_to_dict(q) for q in quotes]


# ---------------------------------------------------------------------------
# DELETE /broker/projects/{project_id} — Soft delete
# ---------------------------------------------------------------------------


@router.delete("/projects/{project_id}", status_code=204)
async def delete_project(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Soft-delete a project (sets deleted_at)."""
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    project.deleted_at = func.now()

    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project_id,
        activity_type="project_deleted",
        actor_type="user",
        metadata_={},
    )
    db.add(activity)
    await db.commit()
    return None


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/cancel — Cancel project
# ---------------------------------------------------------------------------


@router.post("/projects/{project_id}/cancel")
async def cancel_project(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Cancel a project. Returns 409 if already cancelled."""
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    validate_transition(project.status, "cancelled")
    project.status = "cancelled"

    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project_id,
        activity_type="project_cancelled",
        actor_type="user",
        metadata_={},
    )
    db.add(activity)
    await db.commit()
    await db.refresh(project)

    return _project_to_dict(project)


# ---------------------------------------------------------------------------
# GET /broker/dashboard-stats — Aggregated KPIs
# ---------------------------------------------------------------------------

_NEEDS_ACTION_STATUSES = ("new_request", "analysis_failed", "gaps_identified")


@router.get("/dashboard-stats")
async def dashboard_stats(
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Aggregated dashboard KPIs for the broker module."""
    base = BrokerProject.deleted_at.is_(None)

    # Total projects
    total = (
        await db.execute(select(func.count(BrokerProject.id)).where(base))
    ).scalar() or 0

    # Projects by status
    status_rows = (
        await db.execute(
            select(BrokerProject.status, func.count(BrokerProject.id))
            .where(base)
            .group_by(BrokerProject.status)
        )
    ).all()
    projects_by_status = {row[0]: row[1] for row in status_rows}

    # Projects needing action
    needs_action = (
        await db.execute(
            select(func.count(BrokerProject.id)).where(
                base,
                BrokerProject.status.in_(_NEEDS_ACTION_STATUSES),
            )
        )
    ).scalar() or 0

    # Recent projects (top 5 by updated_at)
    recent = (
        await db.execute(
            select(BrokerProject)
            .where(base)
            .order_by(BrokerProject.updated_at.desc())
            .limit(5)
        )
    ).scalars().all()

    return {
        "total_projects": total,
        "projects_by_status": projects_by_status,
        "projects_needing_action": needs_action,
        "recent_projects": [_project_to_dict(p) for p in recent],
    }


# ---------------------------------------------------------------------------
# PATCH /broker/coverages/{coverage_id} — Inline coverage edit
# ---------------------------------------------------------------------------


@router.patch("/coverages/{coverage_id}")
async def update_coverage(
    coverage_id: UUID,
    body: UpdateCoverageBody,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Update a coverage requirement inline. Sets is_manual_override=True."""
    result = await db.execute(
        select(ProjectCoverage).where(ProjectCoverage.id == coverage_id)
    )
    coverage = result.scalar_one_or_none()
    if coverage is None:
        raise HTTPException(status_code=404, detail="Coverage not found")

    # Apply updates for any provided fields
    update_data = body.model_dump(exclude_unset=True)
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")

    for field, value in update_data.items():
        setattr(coverage, field, value)

    coverage.is_manual_override = True

    # Log activity for each updated field
    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=coverage.broker_project_id,
        activity_type="coverage_updated",
        actor_type="user",
        metadata_={"updated_fields": list(update_data.keys())},
    )
    db.add(activity)
    await db.commit()
    await db.refresh(coverage)

    return _coverage_to_dict(coverage)


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/analyze-gaps — Run gap detection
# ---------------------------------------------------------------------------


@router.post("/projects/{project_id}/analyze-gaps")
async def analyze_gaps(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Run gap analysis on a project's coverages.

    Compares required_limit vs current_limit for each coverage, persists
    gap_status and gap_amount, and logs a BrokerActivity event.
    Manual-override coverages are preserved unchanged.
    """
    # Fetch project
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    # Fetch all coverages for this project
    cov_result = await db.execute(
        select(ProjectCoverage)
        .where(ProjectCoverage.broker_project_id == project_id)
        .order_by(ProjectCoverage.created_at)
    )
    coverages = cov_result.scalars().all()

    # Convert ORM objects to dicts for the pure-Python engine
    coverage_dicts = [_coverage_to_dict(c) for c in coverages]

    # Run gap detection
    gap_results = detect_gaps(coverage_dicts)
    summary = summarize_gaps(gap_results)

    # Persist gap_status and gap_amount back to each coverage row
    result_by_id = {r["id"]: r for r in gap_results}
    for cov_orm in coverages:
        updated = result_by_id.get(str(cov_orm.id))
        if updated is None:
            continue
        cov_orm.gap_status = updated.get("gap_status")
        cov_orm.gap_amount = updated.get("gap_amount")

    # Log activity
    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project_id,
        activity_type="gap_analysis_complete",
        actor_type="system",
        description=(
            f"Gap analysis: {summary['covered']} covered, "
            f"{summary['insufficient']} insufficient, "
            f"{summary['missing']} missing"
        ),
        metadata_=summary,
    )
    db.add(activity)
    await db.commit()

    # Refresh coverages to get persisted values
    refreshed_dicts = []
    for cov_orm in coverages:
        await db.refresh(cov_orm)
        refreshed_dicts.append(_coverage_to_dict(cov_orm))

    return {"summary": summary, "coverages": refreshed_dicts}


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/documents — Multi-file upload
# ---------------------------------------------------------------------------


@router.post("/projects/{project_id}/documents", status_code=201)
async def upload_project_documents(
    project_id: UUID,
    files: list[UploadFile] = File(...),
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Upload one or more documents to a broker project."""
    from flywheel.services.document_storage import upload_file as upload_to_storage

    # Verify project exists and is not deleted
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    uploaded_docs: list[dict[str, Any]] = []
    tenant_id = str(user.tenant_id)

    for file in files:
        content = await file.read()
        size = len(content)

        # Validate size
        if size > _MAX_DOC_SIZE:
            raise HTTPException(
                status_code=413,
                detail=f"File '{file.filename}' exceeds 25 MB limit",
            )

        # Validate content type
        if file.content_type not in _ALLOWED_DOC_TYPES:
            raise HTTPException(
                status_code=415,
                detail=f"File '{file.filename}' has unsupported type '{file.content_type}'. "
                f"Allowed: PDF, PNG, JPEG",
            )

        # Upload to Supabase Storage
        file_uuid = uuid4()
        filename = file.filename or "unknown"
        try:
            storage_path = await upload_to_storage(
                tenant_id=tenant_id,
                file_id=str(file_uuid),
                filename=filename,
                content=content,
                mime_type=file.content_type or "application/octet-stream",
            )
        except Exception:
            storage_path = f"local://{tenant_id}/{file_uuid}/{filename}"

        # Create UploadedFile record
        uploaded_file = UploadedFile(
            tenant_id=user.tenant_id,
            user_id=user.sub,
            filename=filename,
            mimetype=file.content_type or "application/octet-stream",
            size_bytes=size,
            storage_path=storage_path,
        )
        db.add(uploaded_file)
        await db.flush()  # Get uploaded_file.id

        doc_ref = {
            "file_id": str(uploaded_file.id),
            "filename": filename,
            "mimetype": file.content_type,
            "size_bytes": size,
            "storage_path": storage_path,
            "uploaded_at": uploaded_file.created_at.isoformat() if uploaded_file.created_at else None,
        }
        uploaded_docs.append(doc_ref)

    # Store document references in project metadata
    existing_docs = (project.metadata_ or {}).get("documents", [])
    existing_docs.extend(uploaded_docs)
    project.metadata_ = {**(project.metadata_ or {}), "documents": existing_docs}

    # Log activity
    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project_id,
        activity_type="document_uploaded",
        actor_type="user",
        metadata_={"filenames": [f.filename for f in files]},
    )
    db.add(activity)
    await db.commit()

    return {"documents": uploaded_docs, "total": len(uploaded_docs)}


# ---------------------------------------------------------------------------
# POST /broker/projects/from-email — Create project from Gmail message
# ---------------------------------------------------------------------------


@router.post("/projects/from-email", status_code=201)
async def create_project_from_email(
    body: CreateFromEmailBody,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Create a broker project from a Gmail message with PDF attachment.

    Downloads the PDF attachment from Gmail, uploads it to Supabase Storage,
    creates an UploadedFile record, and creates a BrokerProject linked to it.
    """
    from flywheel.db.models import Integration
    from flywheel.services.document_storage import upload_file as storage_upload
    from flywheel.services.gmail_read import (
        find_pdf_attachments,
        get_attachment,
        get_valid_credentials,
    )
    from googleapiclient.discovery import build
    import asyncio

    # Load Gmail integration
    result = await db.execute(
        select(Integration).where(
            Integration.id == body.integration_id,
            Integration.tenant_id == user.tenant_id,
            Integration.provider == "gmail-read",
        )
    )
    integration = result.scalar_one_or_none()
    if integration is None:
        raise HTTPException(404, "Gmail integration not found")

    # Build Gmail service
    creds = await get_valid_credentials(integration)

    def _fetch_message():
        service = build("gmail", "v1", credentials=creds)
        return (
            service.users()
            .messages()
            .get(userId="me", id=body.message_id, format="full")
            .execute()
        )

    msg = await asyncio.to_thread(_fetch_message)

    # Find PDF attachments
    pdf_attachments = find_pdf_attachments(msg)
    if not pdf_attachments:
        raise HTTPException(422, "No PDF attachments found in message")

    if body.attachment_index >= len(pdf_attachments):
        raise HTTPException(
            422,
            f"Attachment index {body.attachment_index} out of range "
            f"(message has {len(pdf_attachments)} PDF attachments)",
        )

    pdf_info = pdf_attachments[body.attachment_index]

    # Download the PDF
    pdf_bytes = await get_attachment(creds, body.message_id, pdf_info["attachment_id"])

    # Validate size (25 MB max)
    if len(pdf_bytes) > _MAX_DOC_SIZE:
        raise HTTPException(413, "PDF exceeds 25 MB limit")

    # Upload to Supabase Storage
    file_id = str(uuid4())
    storage_path = await storage_upload(
        tenant_id=str(user.tenant_id),
        file_id=file_id,
        filename=pdf_info["filename"],
        content=pdf_bytes,
        mime_type="application/pdf",
    )

    # Create UploadedFile record
    uploaded_file = UploadedFile(
        tenant_id=user.tenant_id,
        user_id=user.sub,
        filename=pdf_info["filename"],
        mimetype="application/pdf",
        size_bytes=len(pdf_bytes),
        storage_path=storage_path,
    )
    db.add(uploaded_file)
    await db.flush()

    # Extract sender and subject from headers
    headers = {
        h["name"].lower(): h["value"]
        for h in msg.get("payload", {}).get("headers", [])
    }
    sender = headers.get("from", "Unknown Sender")
    subject = headers.get("subject", "Untitled Project")

    # Create BrokerProject
    project = BrokerProject(
        tenant_id=user.tenant_id,
        name=subject,
        project_type="construction",
        status="new_request",
        analysis_status="pending",
        import_source="email",
        external_ref=body.message_id,
        source_document_id=uploaded_file.id,
        metadata_={"sender": sender, "gmail_message_id": body.message_id},
    )
    db.add(project)
    await db.flush()

    # Log activity
    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project.id,
        activity_type="project_created_from_email",
        actor_type="user",
        metadata_={"message_id": body.message_id, "filename": pdf_info["filename"]},
    )
    db.add(activity)
    await db.commit()
    await db.refresh(project)

    return _project_to_dict(project)


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/analyze — Trigger async analysis
# ---------------------------------------------------------------------------


@router.post("/projects/{project_id}/analyze", status_code=202)
async def trigger_analysis(
    project_id: UUID,
    background_tasks: BackgroundTasks,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Trigger async contract analysis for a project.

    Returns 409 if analysis is already running. Uses SELECT FOR UPDATE to
    prevent race conditions. Background task creates its own DB session.
    """
    # Lock row to prevent concurrent triggers
    result = await db.execute(
        select(BrokerProject)
        .where(
            BrokerProject.id == project_id,
            BrokerProject.deleted_at.is_(None),
        )
        .with_for_update()
    )
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(404, "Project not found")
    if project.analysis_status == "running":
        raise HTTPException(409, "Analysis already in progress")

    # Set status immediately
    project.analysis_status = "running"
    project.status = "analyzing"
    await db.commit()

    # Add background task (passes only primitives — session is closed after response)
    background_tasks.add_task(_run_analysis, project_id, user.tenant_id)
    return {"status": "analyzing", "project_id": str(project_id)}


# ---------------------------------------------------------------------------
# Background analysis runner
# ---------------------------------------------------------------------------


async def _get_project_pdf(session: AsyncSession, project, tenant_id: UUID) -> bytes | None:
    """Retrieve PDF bytes for a project from Supabase Storage."""
    if not project.source_document_id:
        return None

    result = await session.execute(
        select(UploadedFile).where(
            UploadedFile.id == project.source_document_id,
            UploadedFile.tenant_id == tenant_id,
        )
    )
    uploaded_file = result.scalar_one_or_none()
    if not uploaded_file or not uploaded_file.storage_path:
        return None

    # Download from Supabase Storage
    import httpx
    from flywheel.config import settings as app_settings

    url = f"{app_settings.supabase_url}/storage/v1/object/{_UPLOADS_BUCKET}/{uploaded_file.storage_path}"
    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.get(
            url,
            headers={"Authorization": f"Bearer {app_settings.supabase_service_key}"},
        )
        if resp.status_code != 200:
            return None
        return resp.content


# Bucket name constant for storage downloads
_UPLOADS_BUCKET = "uploads"


async def _run_analysis(project_id: UUID, tenant_id: UUID):
    """Background: retrieve PDF from Supabase Storage, run contract_analyzer, commit results.

    CRITICAL: Creates its own DB session since the endpoint session is closed
    after the HTTP response is sent.
    """
    import logging as _logging

    from flywheel.db.session import get_session_factory
    from flywheel.engines.contract_analyzer import analyze_contract

    _logger = _logging.getLogger(__name__)
    factory = get_session_factory()

    try:
        async with factory() as session:
            await session.execute(text(f"SET LOCAL app.tenant_id = '{tenant_id}'"))

            # Load project
            project = (
                await session.execute(
                    select(BrokerProject).where(BrokerProject.id == project_id)
                )
            ).scalar_one_or_none()
            if not project:
                return

            # Retrieve PDF from Supabase Storage
            pdf_content = await _get_project_pdf(session, project, tenant_id)
            if not pdf_content:
                project.analysis_status = "failed"
                await session.commit()
                return

            await analyze_contract(session, tenant_id, project_id, pdf_content)
            await session.commit()
    except Exception as exc:
        _logger.error("Analysis failed for project %s: %s", project_id, exc)
        try:
            async with factory() as err_session:
                await err_session.execute(
                    text(f"SET LOCAL app.tenant_id = '{tenant_id}'")
                )
                proj = (
                    await err_session.execute(
                        select(BrokerProject).where(BrokerProject.id == project_id)
                    )
                ).scalar_one()
                proj.analysis_status = "failed"
                validate_transition(proj.status, "analysis_failed")
                proj.status = "analysis_failed"
                await err_session.commit()
        except Exception:
            _logger.error("Failed to update status for project %s", project_id)


# ---------------------------------------------------------------------------
# Carrier CRUD endpoints
# ---------------------------------------------------------------------------


@router.get("/carriers")
async def list_carriers(
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """List all active carrier configs for the tenant, ordered by name."""
    result = await db.execute(
        select(CarrierConfig)
        .where(
            CarrierConfig.tenant_id == user.tenant_id,
            CarrierConfig.is_active.is_(True),
        )
        .order_by(CarrierConfig.carrier_name)
    )
    carriers = result.scalars().all()
    return [_carrier_to_dict(c) for c in carriers]


@router.post("/carriers", status_code=201)
async def create_carrier(
    body: CreateCarrierBody,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Create a new carrier config."""
    carrier = CarrierConfig(
        id=uuid4(),
        tenant_id=user.tenant_id,
        carrier_name=body.carrier_name,
        carrier_type=body.carrier_type,
        submission_method=body.submission_method,
        portal_url=body.portal_url,
        email_address=body.email_address,
        coverage_types=body.coverage_types,
        regions=body.regions,
        min_project_value=body.min_project_value,
        max_project_value=body.max_project_value,
        avg_response_days=body.avg_response_days,
        portal_limit=body.portal_limit,
        notes=body.notes,
    )
    db.add(carrier)
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=409,
            detail="Carrier with this name and type already exists",
        )
    await db.refresh(carrier)
    return _carrier_to_dict(carrier)


@router.put("/carriers/{carrier_id}")
async def update_carrier(
    carrier_id: UUID,
    body: UpdateCarrierBody,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Partially update a carrier config."""
    result = await db.execute(
        select(CarrierConfig).where(
            CarrierConfig.id == carrier_id,
            CarrierConfig.tenant_id == user.tenant_id,
        )
    )
    carrier = result.scalar_one_or_none()
    if carrier is None:
        raise HTTPException(status_code=404, detail="Carrier not found")

    update_data = body.model_dump(exclude_none=True)
    for field, value in update_data.items():
        setattr(carrier, field, value)

    await db.commit()
    await db.refresh(carrier)
    return _carrier_to_dict(carrier)


@router.delete("/carriers/{carrier_id}")
async def delete_carrier(
    carrier_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, str]:
    """Soft-deactivate a carrier config (sets is_active=False)."""
    result = await db.execute(
        select(CarrierConfig).where(
            CarrierConfig.id == carrier_id,
            CarrierConfig.tenant_id == user.tenant_id,
        )
    )
    carrier = result.scalar_one_or_none()
    if carrier is None:
        raise HTTPException(status_code=404, detail="Carrier not found")

    carrier.is_active = False
    await db.commit()
    return {"status": "deactivated"}


# ---------------------------------------------------------------------------
# Carrier matching endpoint
# ---------------------------------------------------------------------------


def _compute_carrier_matches(
    carriers: list,
    project_coverage_types: list[str],
    contract_value: float | None,
) -> list[dict]:
    """Rank carriers by coverage_type intersection with project needs."""
    if not project_coverage_types:
        return []

    matches: list[dict] = []
    for carrier in carriers:
        carrier_coverages = carrier.coverage_types or []
        if not carrier_coverages:
            continue

        # Check project value bounds
        if contract_value is not None:
            if carrier.min_project_value is not None and contract_value < float(carrier.min_project_value):
                continue
            if carrier.max_project_value is not None and contract_value > float(carrier.max_project_value):
                continue

        matched = set(carrier_coverages) & set(project_coverage_types)
        if not matched:
            continue

        unmatched = set(project_coverage_types) - matched
        match_score = len(matched) / len(project_coverage_types)

        matches.append({
            "carrier_config_id": str(carrier.id),
            "carrier_name": carrier.carrier_name,
            "carrier_type": carrier.carrier_type,
            "submission_method": carrier.submission_method,
            "email_address": carrier.email_address,
            "portal_url": carrier.portal_url,
            "matched_coverages": sorted(matched),
            "unmatched_coverages": sorted(unmatched),
            "match_score": round(match_score, 2),
            "avg_response_days": float(carrier.avg_response_days) if carrier.avg_response_days is not None else None,
        })

    # Sort: highest match_score first, then lowest avg_response_days
    matches.sort(key=lambda m: (-m["match_score"], m["avg_response_days"] or 999))
    return matches


@router.get("/projects/{project_id}/carrier-matches")
async def get_carrier_matches(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Return carriers ranked by coverage_type intersection with project needs."""
    # Fetch project
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    # Fetch project coverages
    cov_result = await db.execute(
        select(ProjectCoverage).where(
            ProjectCoverage.broker_project_id == project_id,
        )
    )
    coverages = cov_result.scalars().all()
    coverage_types = list({c.coverage_type for c in coverages if c.coverage_type})

    # Fetch all active carriers for this tenant
    carrier_result = await db.execute(
        select(CarrierConfig).where(
            CarrierConfig.tenant_id == user.tenant_id,
            CarrierConfig.is_active.is_(True),
        )
    )
    carriers = carrier_result.scalars().all()

    contract_value = float(project.contract_value) if project.contract_value is not None else None
    matches = _compute_carrier_matches(carriers, coverage_types, contract_value)

    return {"matches": matches, "project_coverage_count": len(coverage_types)}


# ---------------------------------------------------------------------------
# Helper: check if all carriers solicited -> update project status
# ---------------------------------------------------------------------------


async def _check_all_solicited(db: AsyncSession, project_id: UUID) -> None:
    """If all CarrierQuotes for project are solicited, transition project status."""
    count_total = (
        await db.execute(
            select(func.count(CarrierQuote.id)).where(
                CarrierQuote.broker_project_id == project_id
            )
        )
    ).scalar() or 0

    if count_total == 0:
        return

    count_solicited = (
        await db.execute(
            select(func.count(CarrierQuote.id)).where(
                CarrierQuote.broker_project_id == project_id,
                CarrierQuote.status == "solicited",
            )
        )
    ).scalar() or 0

    if count_solicited == count_total:
        result = await db.execute(
            select(BrokerProject).where(BrokerProject.id == project_id)
        )
        project = result.scalar_one_or_none()
        if project and project.status != "soliciting":
            validate_transition(project.status, "soliciting")
            old_status = project.status
            project.status = "soliciting"
            activity = BrokerActivity(
                tenant_id=project.tenant_id,
                broker_project_id=project_id,
                activity_type="status_change",
                actor_type="system",
                metadata_={"from": old_status, "to": "soliciting"},
            )
            db.add(activity)


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/draft-solicitations
# ---------------------------------------------------------------------------


@router.post("/projects/{project_id}/draft-solicitations")
async def draft_solicitations(
    project_id: UUID,
    body: DraftSolicitationsBody,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Batch-draft solicitation emails for selected carriers.

    For email carriers: creates CarrierQuote, builds submission package, generates
    AI email draft. For portal carriers: creates CarrierQuote and submission package.
    Skips carriers without valid email.
    """
    # Load project
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    # Load project coverages for the drafter
    cov_result = await db.execute(
        select(ProjectCoverage).where(
            ProjectCoverage.broker_project_id == project_id
        )
    )
    coverages = cov_result.scalars().all()
    coverages_list = [_coverage_to_dict(c) for c in coverages]

    drafts: list[dict] = []
    portal_submissions: list[dict] = []
    skipped: list[dict] = []

    for carrier_config_id in body.carrier_config_ids:
        # Load carrier config
        carrier_result = await db.execute(
            select(CarrierConfig).where(
                CarrierConfig.id == carrier_config_id,
                CarrierConfig.tenant_id == user.tenant_id,
                CarrierConfig.is_active.is_(True),
            )
        )
        carrier = carrier_result.scalar_one_or_none()
        if carrier is None:
            skipped.append({
                "carrier_config_id": str(carrier_config_id),
                "reason": "Carrier not found or inactive",
            })
            continue

        method = carrier.submission_method or "email"

        # Handle email track
        if method in ("email", "both"):
            # Validate email
            if not carrier.email_address:
                skipped.append({
                    "carrier_config_id": str(carrier_config_id),
                    "carrier_name": carrier.carrier_name,
                    "carrier": carrier.carrier_name,
                    "reason": "No email address configured",
                })
                if method == "email":
                    continue
            elif not _EMAIL_RE.match(carrier.email_address):
                skipped.append({
                    "carrier_config_id": str(carrier_config_id),
                    "carrier_name": carrier.carrier_name,
                    "carrier": carrier.carrier_name,
                    "reason": f"Invalid email format: {carrier.email_address}",
                })
                if method == "email":
                    continue
            else:
                # Create CarrierQuote for email track
                quote = CarrierQuote(
                    tenant_id=user.tenant_id,
                    broker_project_id=project_id,
                    carrier_config_id=carrier_config_id,
                    carrier_name=carrier.carrier_name,
                    carrier_type=carrier.carrier_type or "insurance",
                    status="pending",
                    draft_status="pending",
                    source="solicitation",
                )
                db.add(quote)
                await db.flush()

                # Build submission package
                documents = await build_submission_package(db, project_id, quote.id)

                # Generate AI draft
                project_dict = _project_to_dict(project)
                carrier_dict = _carrier_to_dict(carrier)
                language = project.language if hasattr(project, "language") and project.language else "en"

                try:
                    ai_result = await draft_solicitation_email(
                        project_dict, carrier_dict, coverages_list, documents, language
                    )
                    quote.draft_subject = ai_result.get("subject", "")
                    quote.draft_body = ai_result.get("body_html", "")
                except Exception as exc:
                    logger.warning(
                        "AI draft failed for carrier %s: %s", carrier.carrier_name, exc
                    )
                    quote.draft_subject = ""
                    quote.draft_body = ""

                drafts.append({
                    "quote_id": str(quote.id),
                    "carrier_name": carrier.carrier_name,
                    "carrier_config_id": str(carrier_config_id),
                    "submission_method": method,
                    "draft_subject": quote.draft_subject,
                    "draft_body": quote.draft_body,
                    "draft_status": quote.draft_status,
                    "documents": documents,
                })

        # Handle portal track
        if method in ("portal", "both"):
            quote_portal = CarrierQuote(
                tenant_id=user.tenant_id,
                broker_project_id=project_id,
                carrier_config_id=carrier_config_id,
                carrier_name=carrier.carrier_name,
                carrier_type=carrier.carrier_type or "insurance",
                status="pending",
                draft_status=None,  # portal track has no email draft
                source="solicitation",
            )
            db.add(quote_portal)
            await db.flush()

            documents = await build_submission_package(db, project_id, quote_portal.id)

            portal_submissions.append({
                "quote_id": str(quote_portal.id),
                "carrier_name": carrier.carrier_name,
                "carrier_config_id": str(carrier_config_id),
                "submission_method": method,
                "portal_url": carrier.portal_url,
                "documents": documents,
            })

    # Commit all changes
    await db.commit()

    # Log activity
    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project_id,
        activity_type="solicitations_drafted",
        actor_type="user",
        metadata_={
            "email_count": len(drafts),
            "portal_count": len(portal_submissions),
            "skipped_count": len(skipped),
        },
    )
    db.add(activity)
    await db.commit()

    return {
        "drafts": drafts,
        "portal_submissions": portal_submissions,
        "skipped": skipped,
    }


# ---------------------------------------------------------------------------
# PUT /broker/quotes/{quote_id}/draft — Edit a solicitation draft
# ---------------------------------------------------------------------------


@router.put("/quotes/{quote_id}/draft")
async def edit_draft(
    quote_id: UUID,
    body: EditDraftBody,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Edit a solicitation email draft before sending."""
    result = await db.execute(
        select(CarrierQuote).where(
            CarrierQuote.id == quote_id,
            CarrierQuote.tenant_id == user.tenant_id,
        )
    )
    quote = result.scalar_one_or_none()
    if quote is None:
        raise HTTPException(status_code=404, detail="Quote not found")

    if quote.draft_status != "pending":
        raise HTTPException(
            status_code=409,
            detail=f"Cannot edit draft with status '{quote.draft_status}' — only 'pending' drafts are editable",
        )

    if body.draft_subject is not None:
        quote.draft_subject = body.draft_subject
    if body.draft_body is not None:
        quote.draft_body = body.draft_body

    await db.commit()
    await db.refresh(quote)

    return {
        "quote_id": str(quote.id),
        "carrier_name": quote.carrier_name,
        "draft_subject": quote.draft_subject,
        "draft_body": quote.draft_body,
        "draft_status": quote.draft_status,
    }


# ---------------------------------------------------------------------------
# POST /broker/quotes/{quote_id}/approve-send — Approve and send solicitation
# ---------------------------------------------------------------------------


@router.post("/quotes/{quote_id}/approve-send")
async def approve_and_send(
    quote_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Approve a solicitation draft and send via email_dispatch."""
    result = await db.execute(
        select(CarrierQuote).where(
            CarrierQuote.id == quote_id,
            CarrierQuote.tenant_id == user.tenant_id,
        )
    )
    quote = result.scalar_one_or_none()
    if quote is None:
        raise HTTPException(status_code=404, detail="Quote not found")

    if quote.draft_status != "pending":
        raise HTTPException(
            status_code=409,
            detail=f"Cannot send draft with status '{quote.draft_status}' — only 'pending' drafts can be sent",
        )

    # Load carrier config for email address
    carrier_result = await db.execute(
        select(CarrierConfig).where(CarrierConfig.id == quote.carrier_config_id)
    )
    carrier_config = carrier_result.scalar_one_or_none()
    if carrier_config is None or not carrier_config.email_address:
        raise HTTPException(
            status_code=422,
            detail="Carrier has no email address configured",
        )

    # Send via email dispatch
    await send_email_as_user(
        db,
        user.tenant_id,
        to=carrier_config.email_address,
        subject=quote.draft_subject or "",
        body_html=quote.draft_body or "",
    )

    # Update quote status
    quote.status = "solicited"
    quote.solicited_at = datetime.now(timezone.utc)
    quote.draft_status = "sent"
    quote.draft_body = None  # PII cleanup

    # Log activity
    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=quote.broker_project_id,
        activity_type="solicitation_sent",
        actor_type="user",
        metadata_={"method": "email", "to": carrier_config.email_address, "carrier_name": quote.carrier_name},
    )
    db.add(activity)

    # Check if all carriers solicited
    await _check_all_solicited(db, quote.broker_project_id)

    await db.commit()
    await db.refresh(quote)

    return {
        "quote_id": str(quote.id),
        "carrier_name": quote.carrier_name,
        "status": quote.status,
        "draft_status": quote.draft_status,
        "solicited_at": quote.solicited_at.isoformat() if quote.solicited_at else None,
    }


# ---------------------------------------------------------------------------
# POST /broker/quotes/{quote_id}/portal-screenshot — Upload portal screenshot
# ---------------------------------------------------------------------------


@router.post("/quotes/{quote_id}/portal-screenshot")
async def portal_screenshot(
    quote_id: UUID,
    screenshot: UploadFile = File(...),
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Upload a portal submission screenshot for review."""
    from flywheel.services.document_storage import upload_file as upload_to_storage

    result = await db.execute(
        select(CarrierQuote).where(
            CarrierQuote.id == quote_id,
            CarrierQuote.tenant_id == user.tenant_id,
        )
    )
    quote = result.scalar_one_or_none()
    if quote is None:
        raise HTTPException(status_code=404, detail="Quote not found")

    # Read and validate screenshot
    content = await screenshot.read()
    if len(content) > _MAX_DOC_SIZE:
        raise HTTPException(status_code=413, detail="Screenshot exceeds 25 MB limit")

    # Upload to Supabase Storage
    file_uuid = uuid4()
    filename = screenshot.filename or "screenshot.png"
    mime_type = screenshot.content_type or "image/png"
    tenant_id_str = str(user.tenant_id)

    try:
        storage_path = await upload_to_storage(
            tenant_id=tenant_id_str,
            file_id=str(file_uuid),
            filename=filename,
            content=content,
            mime_type=mime_type,
        )
        screenshot_url = storage_path
    except Exception:
        screenshot_url = f"local://{tenant_id_str}/{file_uuid}/{filename}"

    # Store in quote metadata
    metadata = dict(quote.metadata_ or {})
    metadata["screenshot_url"] = screenshot_url
    quote.metadata_ = metadata
    quote.draft_status = "review"

    await db.commit()

    return {"screenshot_url": screenshot_url, "status": "review"}


# ---------------------------------------------------------------------------
# POST /broker/quotes/{quote_id}/portal-confirm — Confirm portal submission
# ---------------------------------------------------------------------------


@router.post("/quotes/{quote_id}/portal-confirm")
async def portal_confirm(
    quote_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Confirm a portal submission after screenshot review."""
    result = await db.execute(
        select(CarrierQuote).where(
            CarrierQuote.id == quote_id,
            CarrierQuote.tenant_id == user.tenant_id,
        )
    )
    quote = result.scalar_one_or_none()
    if quote is None:
        raise HTTPException(status_code=404, detail="Quote not found")

    if quote.draft_status != "review":
        raise HTTPException(
            status_code=409,
            detail=f"Cannot confirm with status '{quote.draft_status}' — screenshot must be uploaded first (status='review')",
        )

    quote.status = "solicited"
    quote.solicited_at = datetime.now(timezone.utc)
    quote.draft_status = "confirmed"

    # Log activity
    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=quote.broker_project_id,
        activity_type="solicitation_sent",
        actor_type="user",
        metadata_={"method": "portal", "carrier_name": quote.carrier_name},
    )
    db.add(activity)

    # Check if all carriers solicited
    await _check_all_solicited(db, quote.broker_project_id)

    await db.commit()
    await db.refresh(quote)

    return {
        "quote_id": str(quote.id),
        "carrier_name": quote.carrier_name,
        "status": quote.status,
        "draft_status": quote.draft_status,
        "solicited_at": quote.solicited_at.isoformat() if quote.solicited_at else None,
    }


# ---------------------------------------------------------------------------
# Helper: serialize CarrierQuote
# ---------------------------------------------------------------------------


def _quote_to_dict(quote: CarrierQuote) -> dict[str, Any]:
    """Serialize a CarrierQuote to a JSON-friendly dict."""
    return {
        "id": str(quote.id),
        "broker_project_id": str(quote.broker_project_id),
        "coverage_id": str(quote.coverage_id) if quote.coverage_id else None,
        "carrier_name": quote.carrier_name,
        "carrier_config_id": str(quote.carrier_config_id) if quote.carrier_config_id else None,
        "carrier_type": quote.carrier_type,
        "premium": float(quote.premium) if quote.premium is not None else None,
        "deductible": float(quote.deductible) if quote.deductible is not None else None,
        "limit_amount": float(quote.limit_amount) if quote.limit_amount is not None else None,
        "coinsurance": float(quote.coinsurance) if quote.coinsurance is not None else None,
        "term_months": quote.term_months,
        "validity_date": quote.validity_date.isoformat() if quote.validity_date else None,
        "exclusions": quote.exclusions or [],
        "conditions": quote.conditions or [],
        "endorsements": quote.endorsements or [],
        "is_best_price": quote.is_best_price,
        "is_best_coverage": quote.is_best_coverage,
        "is_recommended": quote.is_recommended,
        "has_critical_exclusion": quote.has_critical_exclusion,
        "critical_exclusion_detail": quote.critical_exclusion_detail,
        "status": quote.status,
        "solicited_at": quote.solicited_at.isoformat() if quote.solicited_at else None,
        "received_at": quote.received_at.isoformat() if quote.received_at else None,
        "confidence": quote.confidence,
        "source": quote.source,
        "is_manual_override": quote.is_manual_override,
        "draft_subject": quote.draft_subject,
        "draft_body": quote.draft_body,
        "draft_status": quote.draft_status,
        "documents": [],
        "created_at": quote.created_at.isoformat() if quote.created_at else None,
        "updated_at": quote.updated_at.isoformat() if quote.updated_at else None,
    }


# ---------------------------------------------------------------------------
# POST /broker/quotes/{quote_id}/extract — Trigger async quote extraction
# ---------------------------------------------------------------------------


@router.post("/quotes/{quote_id}/extract", status_code=202)
async def extract_quote_endpoint(
    quote_id: UUID,
    force: bool = Query(False),
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Trigger async PDF extraction for a carrier quote.

    Returns 202 immediately. Extraction runs as a background task.
    Returns 409 if already extracted/reviewed (unless force=true).
    """
    from flywheel.engines.quote_extractor import extract_quote

    result = await db.execute(
        select(CarrierQuote).where(
            CarrierQuote.id == quote_id,
            CarrierQuote.tenant_id == user.tenant_id,
        )
    )
    quote = result.scalar_one_or_none()
    if quote is None:
        raise HTTPException(status_code=404, detail="Quote not found")

    if quote.status in ("extracted", "reviewed") and not force:
        raise HTTPException(
            status_code=409,
            detail=f"Quote already {quote.status}. Use ?force=true to re-extract.",
        )

    # Load PDF bytes from source document or source email
    pdf_content = None
    if quote.source_document_id:
        pdf_content = await _get_quote_pdf_from_document(db, quote, user.tenant_id)
    elif quote.source_email_id:
        pdf_content = await _get_quote_pdf_from_email(db, quote, user.tenant_id)

    if pdf_content is None:
        raise HTTPException(
            status_code=422,
            detail="No PDF source found for this quote (no source_document_id or source_email_id)",
        )

    # Load project coverages for cross-reference
    cov_result = await db.execute(
        select(ProjectCoverage).where(
            ProjectCoverage.broker_project_id == quote.broker_project_id
        )
    )
    coverages = cov_result.scalars().all()
    coverages_dicts = [
        {
            "id": str(c.id),
            "coverage_type": c.coverage_type,
            "category": c.category,
            "required_limit": float(c.required_limit) if c.required_limit else None,
        }
        for c in coverages
    ]

    # Dispatch extraction as background task
    tenant_id = user.tenant_id

    async def _run_extraction():
        from flywheel.db.session import get_session_factory
        factory = get_session_factory()
        try:
            async with factory() as session:
                await session.execute(text(f"SET LOCAL app.tenant_id = '{tenant_id}'"))
                await extract_quote(
                    session, tenant_id, quote_id, pdf_content, coverages_dicts, force=force
                )
                await session.commit()
        except Exception as exc:
            logger.error("Quote extraction failed for %s: %s", quote_id, exc)

    asyncio.create_task(_run_extraction())

    return {"status": "extracting", "quote_id": str(quote_id)}


async def _get_quote_pdf_from_document(
    db: AsyncSession, quote: CarrierQuote, tenant_id: UUID
) -> bytes | None:
    """Retrieve PDF bytes from UploadedFile via source_document_id."""
    result = await db.execute(
        select(UploadedFile).where(
            UploadedFile.id == quote.source_document_id,
            UploadedFile.tenant_id == tenant_id,
        )
    )
    uploaded_file = result.scalar_one_or_none()
    if not uploaded_file or not uploaded_file.storage_path:
        return None

    import httpx
    from flywheel.config import settings as app_settings

    url = f"{app_settings.supabase_url}/storage/v1/object/{_UPLOADS_BUCKET}/{uploaded_file.storage_path}"
    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.get(
            url,
            headers={"Authorization": f"Bearer {app_settings.supabase_service_key}"},
        )
        if resp.status_code != 200:
            return None
        return resp.content


async def _get_quote_pdf_from_email(
    db: AsyncSession, quote: CarrierQuote, tenant_id: UUID
) -> bytes | None:
    """Retrieve PDF bytes from the source email's attachment."""
    from flywheel.services.gmail_read import (
        find_pdf_attachments,
        get_attachment,
        get_valid_credentials,
    )

    # Load the email row to get gmail_message_id
    email_result = await db.execute(
        select(Email).where(
            Email.id == quote.source_email_id,
            Email.tenant_id == tenant_id,
        )
    )
    email_row = email_result.scalar_one_or_none()
    if not email_row:
        return None

    # We need Gmail credentials -- find integration for this tenant
    intg_result = await db.execute(
        select(Integration).where(
            Integration.tenant_id == tenant_id,
            Integration.provider == "gmail-read",
            Integration.status == "connected",
        )
    )
    integration = intg_result.scalar_one_or_none()
    if not integration:
        return None

    creds = await get_valid_credentials(integration)

    def _fetch():
        from googleapiclient.discovery import build as _build
        service = _build("gmail", "v1", credentials=creds)
        return (
            service.users()
            .messages()
            .get(userId="me", id=email_row.gmail_message_id, format="full")
            .execute()
        )

    msg = await asyncio.to_thread(_fetch)
    pdfs = find_pdf_attachments(msg)
    if not pdfs:
        return None

    # Get first PDF attachment
    return await get_attachment(creds, email_row.gmail_message_id, pdfs[0]["attachment_id"])


# ---------------------------------------------------------------------------
# PUT /broker/quotes/{quote_id} — Manual quote entry
# ---------------------------------------------------------------------------


class ManualQuoteBody(BaseModel):
    premium: float | None = None
    deductible: float | None = None
    limit_amount: float | None = None
    coinsurance: float | None = None
    term_months: int | None = None
    validity_date: str | None = None
    exclusions: list[str] | None = None
    conditions: list[str] | None = None
    endorsements: list[str] | None = None
    coverage_id: UUID | None = None
    confidence: str | None = None


@router.put("/quotes/{quote_id}")
async def update_quote_manual(
    quote_id: UUID,
    body: ManualQuoteBody,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Manual quote entry -- update CarrierQuote fields directly.

    Sets source='manual', is_manual_override=True, status='extracted'.
    """
    result = await db.execute(
        select(CarrierQuote).where(
            CarrierQuote.id == quote_id,
            CarrierQuote.tenant_id == user.tenant_id,
        )
    )
    quote = result.scalar_one_or_none()
    if quote is None:
        raise HTTPException(status_code=404, detail="Quote not found")

    # Apply provided fields
    update_data = body.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(quote, field, value)

    # Manual entry markers
    quote.source = "manual"
    quote.is_manual_override = True
    quote.status = "extracted"

    # Log activity
    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=quote.broker_project_id,
        activity_type="quote_manual_entry",
        actor_type="user",
        metadata_={"fields_set": list(update_data.keys())},
    )
    db.add(activity)
    await db.commit()
    await db.refresh(quote)

    return _quote_to_dict(quote)


# ---------------------------------------------------------------------------
# POST /broker/quotes/{quote_id}/mark-received — Manually mark quote received
# ---------------------------------------------------------------------------


@router.post("/quotes/{quote_id}/mark-received")
async def mark_quote_received(
    quote_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Manually mark a quote as received. Updates project status accordingly."""
    result = await db.execute(
        select(CarrierQuote).where(
            CarrierQuote.id == quote_id,
            CarrierQuote.tenant_id == user.tenant_id,
        )
    )
    quote = result.scalar_one_or_none()
    if quote is None:
        raise HTTPException(status_code=404, detail="Quote not found")

    quote.status = "received"
    quote.received_at = datetime.now(timezone.utc)

    # Update project status (quotes_partial / quotes_complete logic)
    project_result = await db.execute(
        select(BrokerProject).where(BrokerProject.id == quote.broker_project_id)
    )
    project = project_result.scalar_one()

    all_quotes_result = await db.execute(
        select(CarrierQuote).where(
            CarrierQuote.broker_project_id == project.id
        )
    )
    all_quotes = all_quotes_result.scalars().all()

    all_received = all(
        q.status in ("received", "extracted", "reviewed", "selected")
        for q in all_quotes
    )
    if all_received:
        target_status = "quotes_complete"
    else:
        target_status = "quotes_partial"
    validate_transition(project.status, target_status)
    project.status = target_status

    # Log activity
    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project.id,
        activity_type="quote_received",
        actor_type="user",
        description=f"Quote from {quote.carrier_name} manually marked as received",
        metadata_={"carrier_name": quote.carrier_name, "manual": True},
    )
    db.add(activity)
    await db.commit()
    await db.refresh(quote)

    return _quote_to_dict(quote)


# ---------------------------------------------------------------------------
# GET /broker/projects/{project_id}/comparison — Quote comparison matrix
# ---------------------------------------------------------------------------


@router.get("/projects/{project_id}/comparison")
async def get_comparison(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Return fresh quote comparison matrix for a project.

    Stateless computation -- compares all extracted/reviewed/selected quotes
    against project coverages on each call.
    """
    from flywheel.engines.quote_comparator import compare_quotes

    # Verify project
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    # Load coverages as dicts
    cov_result = await db.execute(
        select(ProjectCoverage).where(
            ProjectCoverage.broker_project_id == project_id
        )
    )
    coverages = cov_result.scalars().all()
    coverage_dicts = [_coverage_to_dict(c) for c in coverages]

    # Load quotes in comparable statuses
    quote_result = await db.execute(
        select(CarrierQuote).where(
            CarrierQuote.broker_project_id == project_id,
            CarrierQuote.status.in_(("extracted", "reviewed", "selected")),
        )
    )
    quotes = quote_result.scalars().all()
    quote_dicts = [_quote_to_dict(q) for q in quotes]

    # Compute comparison
    comparison = compare_quotes(coverage_dicts, quote_dicts)

    # Check for partial data
    carrier_ids = {q.carrier_config_id for q in quotes if q.carrier_config_id}
    if len(carrier_ids) < 2:
        comparison["partial"] = True

    return comparison


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/export-comparison — Excel export
# ---------------------------------------------------------------------------


def _build_comparison_xlsx(comparison: dict, project_name: str) -> bytes:
    """Build an Excel workbook from comparison data. Pure function, no async."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Group coverages by category — coverages is a LIST not a dict
    categories: dict[str, list] = {}
    for cov_data in comparison.get("coverages", []):
        cat = cov_data.get("category", "insurance")
        categories.setdefault(cat, []).append(cov_data)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    exclusion_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for cat_name, cov_list in categories.items():
        ws = wb.create_sheet(title=cat_name.capitalize())

        # Collect unique carriers across all coverages in this category
        carrier_map: dict[str, str] = {}  # carrier_id -> name
        carrier_names: list[str] = []
        for cov in cov_list:
            for quote in cov.get("quotes", []):
                cid = quote.get("carrier_config_id") or quote.get("carrier_id")
                if cid and cid not in carrier_map:
                    carrier_map[cid] = quote.get("carrier_name", "Unknown")
                    carrier_names.append(quote.get("carrier_name", "Unknown"))

        # Header row
        headers = ["Coverage", "Required Limit"] + carrier_names
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        # Coverage rows
        row = 2
        carrier_id_list = list(carrier_map.keys())
        for cov in cov_list:
            ws.cell(row=row, column=1, value=cov.get("display_name", cov.get("coverage_type", "")))
            req_limit = cov.get("required_limit")
            ws.cell(row=row, column=2, value=float(req_limit) if req_limit is not None else None)

            for q in cov.get("quotes", []):
                cid = q.get("carrier_config_id") or q.get("carrier_id")
                if cid in carrier_map:
                    col_idx = carrier_id_list.index(cid) + 3
                    premium = q.get("premium") or q.get("quoted_premium")
                    ws.cell(row=row, column=col_idx, value=float(premium) if premium is not None else None)

                    if q.get("has_critical_exclusion"):
                        ws.cell(row=row, column=col_idx).fill = exclusion_fill
            row += 1

        # Auto-width columns
        for col_cells in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)

    # If no categories found, create a single empty sheet
    if not categories:
        wb.create_sheet(title="No Data")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.post("/projects/{project_id}/export-comparison")
async def export_comparison(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
    body: ExportComparisonBody | None = None,
) -> StreamingResponse:
    """Export quote comparison as .xlsx file.

    Reuses compare_quotes engine, then builds Excel in a thread to avoid
    blocking the async event loop. Optional quote_ids filter.
    """
    from flywheel.engines.quote_comparator import compare_quotes

    # Verify project
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    # Load coverages
    cov_result = await db.execute(
        select(ProjectCoverage).where(
            ProjectCoverage.broker_project_id == project_id
        )
    )
    coverages = cov_result.scalars().all()
    coverage_dicts = [_coverage_to_dict(c) for c in coverages]

    # Load quotes in comparable statuses
    quote_query = select(CarrierQuote).where(
        CarrierQuote.broker_project_id == project_id,
        CarrierQuote.status.in_(("extracted", "reviewed", "selected")),
    )
    # Optional filter by quote_ids
    if body and body.quote_ids:
        quote_query = quote_query.where(CarrierQuote.id.in_(body.quote_ids))

    quote_result = await db.execute(quote_query)
    quotes = quote_result.scalars().all()
    quote_dicts = [_quote_to_dict(q) for q in quotes]

    # Compute comparison
    comparison = compare_quotes(coverage_dicts, quote_dicts)

    # Build Excel in thread (openpyxl is synchronous)
    content = await asyncio.to_thread(_build_comparison_xlsx, comparison, project.name or "comparison")

    safe_name = re.sub(r"[^\w\s\-]", "", project.name or "comparison").strip().replace(" ", "_")[:60]
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_comparison.xlsx"'},
    )


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/draft-followups — Generate follow-ups
# ---------------------------------------------------------------------------


@router.post("/projects/{project_id}/draft-followups")
async def draft_followups_endpoint(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Generate follow-up email drafts for carriers that haven't responded.

    Loads all solicited (not yet received) quotes for the project, checks timing
    thresholds, and drafts follow-ups for those past due.
    """
    from flywheel.engines.followup_drafter import draft_followup

    # Verify project
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    # Load all solicited quotes (not yet received)
    quote_result = await db.execute(
        select(CarrierQuote).where(
            CarrierQuote.broker_project_id == project_id,
            CarrierQuote.status == "solicited",
        )
    )
    solicited_quotes = quote_result.scalars().all()

    followups: list[dict] = []
    not_due: list[dict] = []
    skipped: list[dict] = []

    for quote in solicited_quotes:
        try:
            # Load carrier config
            if not quote.carrier_config_id:
                skipped.append({
                    "quote_id": str(quote.id),
                    "carrier_name": quote.carrier_name,
                    "reason": "No carrier_config_id",
                })
                continue

            carrier_result = await db.execute(
                select(CarrierConfig).where(CarrierConfig.id == quote.carrier_config_id)
            )
            carrier = carrier_result.scalar_one_or_none()
            if carrier is None:
                skipped.append({
                    "quote_id": str(quote.id),
                    "carrier_name": quote.carrier_name,
                    "reason": "Carrier config not found",
                })
                continue

            # Call draft_followup engine
            draft_result = await draft_followup(
                db, user.tenant_id, quote, project, carrier
            )

            if draft_result["status"] == "not_due":
                not_due.append({
                    "quote_id": str(quote.id),
                    "carrier_name": quote.carrier_name,
                    "days_remaining": draft_result["days_remaining"],
                })
            else:
                # Update quote with draft
                quote.draft_subject = draft_result["subject"]
                quote.draft_body = draft_result["body"]
                quote.draft_status = "pending"

                followups.append({
                    "quote_id": str(quote.id),
                    "carrier_name": quote.carrier_name,
                    "subject": draft_result["subject"],
                    "body": draft_result["body"],
                })

        except Exception as exc:
            logger.warning(
                "Follow-up draft failed for quote %s: %s", quote.id, exc
            )
            skipped.append({
                "quote_id": str(quote.id),
                "carrier_name": quote.carrier_name,
                "reason": str(exc),
            })

    await db.commit()

    return {"followups": followups, "not_due": not_due, "skipped": skipped}


# ---------------------------------------------------------------------------
# Recommendation: Pydantic bodies
# ---------------------------------------------------------------------------


class DraftRecommendationBody(BaseModel):
    recipient_email: str | None = None


class EditRecommendationBody(BaseModel):
    subject: str | None = None
    body: str | None = None
    recipient_email: str | None = None


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/draft-recommendation
# ---------------------------------------------------------------------------


@router.post("/projects/{project_id}/draft-recommendation")
async def draft_recommendation(
    project_id: UUID,
    body: DraftRecommendationBody | None = None,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Generate an AI recommendation email from comparison data.

    Requires project status 'quotes_complete' and at least one recommended quote.
    """
    from flywheel.engines.quote_comparator import compare_quotes, summarize_comparison
    from flywheel.engines.recommendation_drafter import draft_recommendation_email

    # Verify project exists
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    # Guard: must be in quotes_complete status
    validate_transition(project.status, "recommended")

    # Guard: at least one quote must be recommended
    rec_check = await db.execute(
        select(func.count()).select_from(CarrierQuote).where(
            CarrierQuote.broker_project_id == project_id,
            CarrierQuote.is_recommended.is_(True),
        )
    )
    if rec_check.scalar() == 0:
        raise HTTPException(
            status_code=422,
            detail="No recommended quotes found. Run comparison first.",
        )

    # Handle recipient email
    recipient = None
    if body and body.recipient_email:
        recipient = body.recipient_email
        project.recommendation_recipient = recipient
    elif project.recommendation_recipient:
        recipient = project.recommendation_recipient

    # Load coverages and quotes for comparison
    cov_result = await db.execute(
        select(ProjectCoverage).where(
            ProjectCoverage.broker_project_id == project_id
        )
    )
    coverages = cov_result.scalars().all()
    coverage_dicts = [_coverage_to_dict(c) for c in coverages]

    quote_result = await db.execute(
        select(CarrierQuote).where(
            CarrierQuote.broker_project_id == project_id,
            CarrierQuote.status.in_(("extracted", "reviewed", "selected")),
        )
    )
    quotes = quote_result.scalars().all()
    quote_dicts = [_quote_to_dict(q) for q in quotes]

    # Run comparison and summary
    comparison = compare_quotes(coverage_dicts, quote_dicts)
    summary = summarize_comparison(comparison)

    # Build project dict for the drafter engine
    project_dict = _project_to_dict(project)

    # Generate AI recommendation
    language = project.language if hasattr(project, "language") and project.language else "en"
    draft_result = await draft_recommendation_email(
        project_dict, comparison, summary, language
    )

    # Store result on project
    project.recommendation_subject = draft_result["subject"]
    project.recommendation_body = draft_result["body_html"]
    project.recommendation_status = "pending"

    # Transition status
    validate_transition(project.status, "recommended")
    project.status = "recommended"

    # Log activity
    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project.id,
        activity_type="recommendation_drafted",
        actor_type="system",
        description=f"AI recommendation drafted for {project.name}",
        metadata_={"recipient": recipient},
    )
    db.add(activity)
    await db.commit()

    return {
        "subject": draft_result["subject"],
        "body_html": draft_result["body_html"],
        "recipient": recipient,
    }


# ---------------------------------------------------------------------------
# PUT /broker/projects/{project_id}/recommendation-draft
# ---------------------------------------------------------------------------


@router.put("/projects/{project_id}/recommendation-draft")
async def edit_recommendation_draft(
    project_id: UUID,
    body: EditRecommendationBody,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Edit the recommendation email draft before sending."""
    # Verify project exists
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    # Update fields
    if body.subject is not None:
        project.recommendation_subject = body.subject
    if body.body is not None:
        project.recommendation_body = body.body
    if body.recipient_email is not None:
        project.recommendation_recipient = body.recipient_email

    await db.commit()
    await db.refresh(project)

    return {
        "subject": project.recommendation_subject,
        "body_html": project.recommendation_body,
        "recipient": project.recommendation_recipient,
    }


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/approve-send-recommendation
# ---------------------------------------------------------------------------


@router.post("/projects/{project_id}/approve-send-recommendation")
async def approve_send_recommendation(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Approve and send the recommendation email, save to document library.

    Transitions project to 'delivered' status.
    """
    # Verify project exists
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    # Guard: recommendation must be pending
    if project.recommendation_status != "pending" or project.recommendation_body is None:
        raise HTTPException(
            status_code=422,
            detail="No pending recommendation draft. Generate one first.",
        )

    # Guard: recipient must be set
    if not project.recommendation_recipient:
        raise HTTPException(
            status_code=422,
            detail="No recipient email set. Edit the draft to add a recipient.",
        )

    # Send email
    email_result = await send_email_as_user(
        db=db,
        tenant_id=user.tenant_id,
        to=project.recommendation_recipient,
        subject=project.recommendation_subject or "Insurance Recommendation",
        body_html=project.recommendation_body,
    )

    # Update project status
    now = datetime.now(timezone.utc)
    project.recommendation_status = "sent"
    project.recommendation_sent_at = now

    # Transition to delivered
    validate_transition(project.status, "delivered")
    project.status = "delivered"

    # Save recommendation to document library
    doc = Document(
        tenant_id=user.tenant_id,
        user_id=user.sub,
        title=f"Recommendation - {project.name}",
        document_type="broker-recommendation",
        module="broker",
        tags=["broker", "recommendation", project.project_type or "insurance"],
        metadata_={
            "broker_project_id": str(project.id),
            "sent_to": project.recommendation_recipient,
            "sent_at": now.isoformat(),
        },
    )
    db.add(doc)
    await db.flush()

    # Log activity
    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project.id,
        activity_type="recommendation_sent",
        actor_type="user",
        description=f"Recommendation sent to {project.recommendation_recipient}",
        metadata_={
            "recipient": project.recommendation_recipient,
            "document_id": str(doc.id),
            "email_provider": email_result.get("provider"),
        },
    )
    db.add(activity)
    await db.commit()

    return {
        "status": "sent",
        "sent_at": now.isoformat(),
        "document_id": str(doc.id),
    }
