"""Broker project CRUD and workflow endpoints.

Endpoints:
  POST /projects                               -- create project
  GET  /projects                               -- list projects (paginated)
  GET  /projects/from-email                    -- (see below)
  POST /projects/from-email                    -- create from email attachment
  GET  /projects/{id}                          -- get project with coverages + activities
  DELETE /projects/{id}                        -- soft delete
  POST /projects/{id}/cancel                   -- cancel project
  POST /projects/{id}/approve                  -- Gate 1 approval
  POST /projects/{id}/analyze-gaps             -- run gap analysis
  POST /projects/{id}/documents                -- upload documents
  POST /projects/{id}/analyze                  -- trigger async contract analysis
  GET  /projects/{id}/carrier-matches          -- ranked carrier matches
  GET  /projects/{id}/comparison               -- quote comparison matrix
  POST /projects/{id}/export-comparison        -- Excel export
  POST /projects/{id}/draft-followups          -- generate follow-up drafts
  GET  /health                                 -- broker module health check
  GET  /gate-counts                            -- gate strip counts
  GET  /dashboard-stats                        -- aggregated KPIs
  GET  /dashboard-tasks                        -- urgency-ordered task list
  PATCH /coverages/{id}                        -- inline coverage edit
"""
from __future__ import annotations

import asyncio
import io
import re
from datetime import datetime, timezone
from typing import Any
from uuid import UUID, uuid4

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import exists, func, select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from flywheel.api.broker._shared import validate_transition
from flywheel.api.deps import get_tenant_db, require_module
from flywheel.auth.jwt import TokenPayload
from flywheel.db.models import (
    BrokerActivity,
    BrokerProject,
    CarrierConfig,
    CarrierQuote,
    CoverageType,
    Document,
    Email,
    Integration,
    ProjectCoverage,
    SolicitationDraft,
    SubmissionDocument,
    UploadedFile,
)
from flywheel.engines.gap_detector import detect_gaps, summarize_gaps

import logging

logger = logging.getLogger(__name__)

projects_router = APIRouter(tags=["broker"])

# Max file size for document upload (25 MB)
_MAX_DOC_SIZE = 25 * 1024 * 1024
_ALLOWED_DOC_TYPES = {
    "application/pdf",
    "image/png",
    "image/jpeg",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "text/plain",
    "text/csv",
}
_UPLOADS_BUCKET = "uploads"
_NEEDS_ACTION_STATUSES = ("new_request", "analysis_failed", "gaps_identified")


# ---------------------------------------------------------------------------
# Pydantic request bodies
# ---------------------------------------------------------------------------


class CreateProjectBody(BaseModel):
    name: str
    project_type: str | None = None
    description: str | None = None
    contract_value: float | None = None
    location: str | None = None
    client_id: UUID | None = None
    country_code: str | None = None
    line_of_business: str | None = None


class CreateFromEmailBody(BaseModel):
    message_id: str
    integration_id: UUID
    attachment_index: int = 0


class UpdateCoverageBody(BaseModel):
    coverage_type: str | None = None
    display_name: str | None = None
    required_limit: float | None = None
    required_deductible: float | None = None
    required_terms: str | None = None
    contract_clause: str | None = None
    current_limit: float | None = None
    current_carrier: str | None = None
    current_policy_number: str | None = None
    gap_notes: str | None = None


class UpdateProjectBody(BaseModel):
    name: str | None = None
    project_type: str | None = None
    description: str | None = None
    contract_value: float | None = None
    location: str | None = None
    country_code: str | None = None
    line_of_business: str | None = None


class ExportComparisonBody(BaseModel):
    quote_ids: list[UUID] | None = None


# ---------------------------------------------------------------------------
# Serializers
# ---------------------------------------------------------------------------


def _project_to_dict(p: BrokerProject) -> dict[str, Any]:
    """Serialize a BrokerProject to a JSON-friendly dict.

    NOTE: Removed dropped columns (recommendation_subject, recommendation_body,
    recommendation_status, recommendation_sent_at, recommendation_recipient).
    Added client_id and context_entity_id (Phase 129/130).
    """
    return {
        "id": str(p.id),
        "tenant_id": str(p.tenant_id),
        "name": p.name,
        "project_type": p.project_type,
        "description": p.description,
        "contract_value": float(p.contract_value) if p.contract_value is not None else None,
        "currency": p.currency,
        "location": p.location,
        "language": p.language,
        "status": p.status,
        "approval_status": p.approval_status,
        "analysis_status": p.analysis_status,
        "source": p.import_source,
        "source_ref": p.external_ref,
        "notes": None,
        "metadata": p.metadata_,
        "client_id": str(p.client_id) if p.client_id else None,
        "country_code": p.country_code,
        "line_of_business": p.line_of_business,
        "context_entity_id": str(p.context_entity_id) if p.context_entity_id else None,
        "created_at": p.created_at.isoformat() if p.created_at else None,
        "updated_at": p.updated_at.isoformat() if p.updated_at else None,
        "start_date": p.start_date.isoformat() if p.start_date else None,
    }


def _coverage_to_dict(
    c: ProjectCoverage,
    critical_findings: list | None = None,
) -> dict[str, Any]:
    """Serialize a ProjectCoverage to a JSON-friendly dict."""
    return {
        "id": str(c.id),
        "broker_project_id": str(c.broker_project_id),
        "coverage_type": c.coverage_type,
        "coverage_type_key": c.coverage_type_key,
        "category": c.category,
        "display_name": c.display_name,
        "description": c.required_terms or c.display_name,
        "language": getattr(c, "language", None),
        "required_limit": float(c.required_limit) if c.required_limit is not None else None,
        "limit_currency": (c.metadata_ or {}).get("limit_currency"),
        "required_deductible": float(c.required_deductible) if c.required_deductible is not None else None,
        "required_terms": c.required_terms,
        "contract_clause": c.contract_clause,
        "current_limit": float(c.current_limit) if c.current_limit is not None else None,
        "current_carrier": c.current_carrier,
        "current_policy_number": c.current_policy_number,
        "current_expiry": c.current_expiry.isoformat() if c.current_expiry else None,
        "gap_status": c.gap_status,
        "gap_amount": float(c.gap_amount) if c.gap_amount is not None else None,
        "gap_notes": c.gap_notes,
        "source": c.source,
        "source_excerpt": c.source_excerpt,
        "source_page": c.source_page,
        "source_section": c.source_section,
        "confidence": c.confidence,
        "is_manual_override": c.is_manual_override,
        "ai_critical_finding": bool(
            critical_findings
            and any(
                cf.get("coverage_type") == c.coverage_type
                for cf in critical_findings
            )
        ),
        "created_at": c.created_at.isoformat() if c.created_at else None,
    }


def _activity_to_dict(a: BrokerActivity) -> dict[str, Any]:
    """Serialize a BrokerActivity to a JSON-friendly dict."""
    return {
        "id": str(a.id),
        "activity_type": a.activity_type,
        "description": a.description,
        "actor_type": a.actor_type,
        "metadata": a.metadata_,
        "occurred_at": a.occurred_at.isoformat() if a.occurred_at else None,
    }


def _quote_to_dict_minimal(q: CarrierQuote) -> dict[str, Any]:
    """Minimal quote dict for comparison/export (no dropped columns)."""
    return {
        "id": str(q.id),
        "broker_project_id": str(q.broker_project_id),
        "coverage_id": str(q.coverage_id) if q.coverage_id else None,
        "carrier_name": q.carrier_name,
        "carrier_config_id": str(q.carrier_config_id) if q.carrier_config_id else None,
        "carrier_type": q.carrier_type,
        "premium": float(q.premium) if q.premium is not None else None,
        "deductible": float(q.deductible) if q.deductible is not None else None,
        "limit_amount": float(q.limit_amount) if q.limit_amount is not None else None,
        "coinsurance": float(q.coinsurance) if q.coinsurance is not None else None,
        "term_months": q.term_months,
        "validity_date": q.validity_date.isoformat() if q.validity_date else None,
        "exclusions": q.exclusions or [],
        "conditions": q.conditions or [],
        "endorsements": q.endorsements or [],
        "has_critical_exclusion": q.has_critical_exclusion,
        "critical_exclusion_detail": q.critical_exclusion_detail,
        "status": q.status,
        "solicited_at": q.solicited_at.isoformat() if q.solicited_at else None,
        "received_at": q.received_at.isoformat() if q.received_at else None,
        "confidence": q.confidence,
        "source": q.source,
        "is_manual_override": q.is_manual_override,
        "documents": [],
        "created_at": q.created_at.isoformat() if q.created_at else None,
        "updated_at": q.updated_at.isoformat() if q.updated_at else None,
    }


# ---------------------------------------------------------------------------
# GET /broker/health
# ---------------------------------------------------------------------------


@projects_router.get("/health")
async def broker_health(
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Health check for broker module."""
    return {"status": "ok"}


# ---------------------------------------------------------------------------
# GET /broker/coverage-types
# ---------------------------------------------------------------------------


@projects_router.get("/coverage-types")
async def get_coverage_types(
    country: str | None = None,
    lob: str | None = None,
    db: AsyncSession = Depends(get_tenant_db),
    _user: TokenPayload = Depends(require_module("broker")),
) -> dict[str, Any]:
    """Return filtered taxonomy of canonical coverage types."""
    from sqlalchemy import or_

    filters = [CoverageType.is_active.is_(True)]
    if country:
        filters.append(
            or_(
                CoverageType.countries == func.cast(text("'{}'"), CoverageType.countries.type),
                func.array_position(CoverageType.countries, country).isnot(None),
            )
        )
    if lob:
        filters.append(
            or_(
                CoverageType.lines_of_business == func.cast(text("'{}'"), CoverageType.lines_of_business.type),
                func.array_position(CoverageType.lines_of_business, lob).isnot(None),
            )
        )

    result = await db.execute(
        select(CoverageType)
        .where(*filters)
        .order_by(CoverageType.sort_order, CoverageType.key)
    )
    rows = result.scalars().all()

    coverage_types = []
    for ct in rows:
        display_names = ct.display_names or {}
        aliases_map = ct.aliases or {}
        # Flatten aliases from all locales into a single list
        all_aliases: list[str] = []
        for locale_aliases in aliases_map.values():
            if isinstance(locale_aliases, list):
                all_aliases.extend(locale_aliases)

        coverage_types.append({
            "key": ct.key,
            "category": ct.category,
            "display_name": display_names.get("en", ct.key),
            "aliases": all_aliases,
            "countries": ct.countries or [],
            "lines_of_business": ct.lines_of_business or [],
            "is_verified": ct.is_verified,
        })

    return {"coverage_types": coverage_types}


# ---------------------------------------------------------------------------
# GET /broker/gate-counts
# ---------------------------------------------------------------------------


@projects_router.get("/gate-counts")
async def get_gate_counts(
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Return counts for each gate in the broker workflow strip."""
    base = (
        select(BrokerProject.id, BrokerProject.created_at)
        .where(
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
        .order_by(BrokerProject.created_at.asc())
    )

    # Review: gaps identified, not yet approved
    review_q = base.where(
        BrokerProject.status == "gaps_identified",
        BrokerProject.approval_status == "draft",
    )
    review_rows = (await db.execute(review_q)).all()

    # Approve: approved projects with at least one pending SolicitationDraft
    has_pending_draft = exists(
        select(SolicitationDraft.id).where(
            SolicitationDraft.broker_project_id == BrokerProject.id,
            SolicitationDraft.status == "pending",
        )
    )
    approve_q = base.where(
        BrokerProject.status == "gaps_identified",
        BrokerProject.approval_status == "approved",
        has_pending_draft,
    )
    approve_rows = (await db.execute(approve_q)).all()

    # Export: quotes complete or recommended
    export_q = base.where(
        BrokerProject.status.in_(["quotes_complete", "recommended"]),
    )
    export_rows = (await db.execute(export_q)).all()

    def _gate(rows: list) -> dict[str, Any]:
        return {
            "count": len(rows),
            "oldest_project_id": str(rows[0].id) if rows else None,
        }

    return {
        "review": _gate(review_rows),
        "approve": _gate(approve_rows),
        "export": _gate(export_rows),
    }


# ---------------------------------------------------------------------------
# GET /broker/dashboard-stats
# ---------------------------------------------------------------------------


@projects_router.get("/dashboard-stats")
async def dashboard_stats(
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Aggregated dashboard KPIs for the broker module."""
    base = BrokerProject.deleted_at.is_(None)

    total = (
        await db.execute(select(func.count(BrokerProject.id)).where(base))
    ).scalar() or 0

    status_rows = (
        await db.execute(
            select(BrokerProject.status, func.count(BrokerProject.id))
            .where(base)
            .group_by(BrokerProject.status)
        )
    ).all()
    projects_by_status = {row[0]: row[1] for row in status_rows}

    needs_action = (
        await db.execute(
            select(func.count(BrokerProject.id)).where(
                base,
                BrokerProject.status.in_(_NEEDS_ACTION_STATUSES),
            )
        )
    ).scalar() or 0

    recent = (
        await db.execute(
            select(BrokerProject)
            .where(base)
            .order_by(BrokerProject.updated_at.desc())
            .limit(5)
        )
    ).scalars().all()

    # Total premium: sum of best (lowest) quote premium per active project
    min_premium_sq = (
        select(
            CarrierQuote.broker_project_id,
            func.min(CarrierQuote.premium).label("best_premium"),
        )
        .where(
            CarrierQuote.premium.isnot(None),
            CarrierQuote.status.in_(["extracted", "selected"]),
        )
        .group_by(CarrierQuote.broker_project_id)
        .subquery()
    )
    total_premium_result = (
        await db.execute(select(func.sum(min_premium_sq.c.best_premium)))
    ).scalar() or 0

    return {
        "total_projects": total,
        "projects_by_status": projects_by_status,
        "projects_needing_action": needs_action,
        "total_premium": float(total_premium_result),
        "recent_projects": [_project_to_dict(p) for p in recent],
    }


# ---------------------------------------------------------------------------
# GET /broker/dashboard-tasks
# ---------------------------------------------------------------------------


@projects_router.get("/dashboard-tasks")
async def get_dashboard_tasks(
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Return urgency-ordered task list for broker dashboard."""
    tasks: list[dict[str, Any]] = []

    # 1. "review" tasks
    review_result = await db.execute(
        select(BrokerProject)
        .where(
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.status == "gaps_identified",
            BrokerProject.approval_status == "draft",
            BrokerProject.deleted_at.is_(None),
        )
        .order_by(BrokerProject.created_at.asc())
    )
    for p in review_result.scalars().all():
        tasks.append({
            "type": "review",
            "priority": 1,
            "project_id": str(p.id),
            "project_name": p.name,
            "created_at": p.created_at.isoformat() if p.created_at else None,
            "message": "Review extracted coverages",
        })

    # 2. "approve" tasks — use SolicitationDraft.status (not CarrierQuote.draft_status)
    approve_result = await db.execute(
        select(BrokerProject)
        .where(
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.status == "gaps_identified",
            BrokerProject.approval_status == "approved",
            BrokerProject.deleted_at.is_(None),
            exists(
                select(SolicitationDraft.id).where(
                    SolicitationDraft.broker_project_id == BrokerProject.id,
                    SolicitationDraft.status == "pending",
                )
            ),
        )
        .order_by(BrokerProject.created_at.asc())
    )
    for p in approve_result.scalars().all():
        tasks.append({
            "type": "approve",
            "priority": 2,
            "project_id": str(p.id),
            "project_name": p.name,
            "created_at": p.created_at.isoformat() if p.created_at else None,
            "message": "Approve & send solicitations",
        })

    # 3. "export" tasks
    export_result = await db.execute(
        select(BrokerProject)
        .where(
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.status.in_(("quotes_complete", "recommended")),
            BrokerProject.deleted_at.is_(None),
        )
        .order_by(BrokerProject.created_at.asc())
    )
    for p in export_result.scalars().all():
        tasks.append({
            "type": "export",
            "priority": 3,
            "project_id": str(p.id),
            "project_name": p.name,
            "created_at": p.created_at.isoformat() if p.created_at else None,
            "message": "Export comparison for client",
        })

    # 4. "followup" tasks — overdue solicitations
    followup_result = await db.execute(
        select(
            CarrierQuote.id.label("quote_id"),
            CarrierQuote.solicited_at,
            BrokerProject.id.label("project_id"),
            BrokerProject.name.label("project_name"),
            CarrierConfig.carrier_name,
        )
        .join(BrokerProject, CarrierQuote.broker_project_id == BrokerProject.id)
        .join(CarrierConfig, CarrierQuote.carrier_config_id == CarrierConfig.id)
        .where(
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
            CarrierQuote.status == "solicited",
            CarrierQuote.solicited_at.isnot(None),
            func.now() - CarrierQuote.solicited_at > text("interval '7 days'"),
        )
        .order_by(CarrierQuote.solicited_at.asc())
    )
    for row in followup_result.all():
        days_overdue = (
            (datetime.now(timezone.utc) - row.solicited_at).days
            if row.solicited_at
            else 0
        )
        tasks.append({
            "type": "followup",
            "priority": 4,
            "project_id": str(row.project_id),
            "project_name": row.project_name,
            "quote_id": str(row.quote_id),
            "solicited_at": row.solicited_at.isoformat() if row.solicited_at else None,
            "carrier_name": row.carrier_name,
            "days_overdue": days_overdue,
            "message": f"Follow up with {row.carrier_name}",
        })

    tasks = tasks[:50]
    return {"tasks": tasks, "total": len(tasks)}


# ---------------------------------------------------------------------------
# POST /broker/projects
# ---------------------------------------------------------------------------


@projects_router.post("/projects", status_code=201)
async def create_project(
    body: CreateProjectBody,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Create a new broker project manually."""
    from flywheel.engines.context_store_writer import create_context_entity

    project = BrokerProject(
        tenant_id=user.tenant_id,
        name=body.name,
        project_type=body.project_type or "construction",
        description=body.description,
        contract_value=body.contract_value,
        location=body.location,
        status="new_request",
        analysis_status="pending",
        import_source="manual",
        client_id=body.client_id,
        created_by_user_id=user.sub,
        **({"country_code": body.country_code} if body.country_code else {}),
        **({"line_of_business": body.line_of_business} if body.line_of_business else {}),
    )
    db.add(project)
    await db.flush()

    # Create context entity for the project
    project.context_entity_id = await create_context_entity(
        db=db,
        tenant_id=user.tenant_id,
        name=body.name,
        entity_type="broker_project",
    )

    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project.id,
        activity_type="project_created",
        actor_type="user",
        metadata_={"source": "manual"},
    )
    db.add(activity)
    await db.commit()
    await db.refresh(project)

    return _project_to_dict(project)


# ---------------------------------------------------------------------------
# GET /broker/projects
# ---------------------------------------------------------------------------


@projects_router.get("/projects")
async def list_projects(
    offset: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    status: str | None = Query(None),
    search: str | None = Query(None),
    client_id: UUID | None = Query(None),
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """List broker projects with pagination, filtering, and search.

    client_id: optional filter to return only projects for a specific client (API-06).
    """
    base_filter = BrokerProject.deleted_at.is_(None)

    filters = [base_filter, BrokerProject.tenant_id == user.tenant_id]
    if status:
        filters.append(BrokerProject.status == status)
    if search:
        filters.append(BrokerProject.name.ilike(f"%{search}%"))
    if client_id is not None:
        filters.append(BrokerProject.client_id == client_id)

    count_q = select(func.count(BrokerProject.id)).where(*filters)
    total = (await db.execute(count_q)).scalar() or 0

    data_q = (
        select(BrokerProject)
        .where(*filters)
        .order_by(BrokerProject.updated_at.desc())
        .offset(offset)
        .limit(limit)
    )
    projects = (await db.execute(data_q)).scalars().all()

    return {
        "items": [_project_to_dict(p) for p in projects],
        "total": total,
        "offset": offset,
        "limit": limit,
        "has_more": offset + limit < total,
    }


# ---------------------------------------------------------------------------
# GET /broker/projects/{project_id}
# ---------------------------------------------------------------------------


@projects_router.get("/projects/{project_id}")
async def get_project(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Get a single project with nested coverages and activities."""
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    coverages_result = await db.execute(
        select(ProjectCoverage)
        .where(ProjectCoverage.broker_project_id == project_id)
        .order_by(ProjectCoverage.created_at)
    )
    coverages = coverages_result.scalars().all()

    activities_result = await db.execute(
        select(BrokerActivity)
        .where(BrokerActivity.broker_project_id == project_id)
        .order_by(BrokerActivity.occurred_at.desc())
        .limit(50)
    )
    activities = activities_result.scalars().all()

    project_dict = _project_to_dict(project)
    cf = getattr(project, "critical_findings", None)
    project_dict["coverages"] = [
        _coverage_to_dict(c, critical_findings=cf) for c in coverages
    ]
    project_dict["activities"] = [_activity_to_dict(a) for a in activities]

    return project_dict


# ---------------------------------------------------------------------------
# DELETE /broker/projects/{project_id}
# ---------------------------------------------------------------------------


@projects_router.delete("/projects/{project_id}", status_code=204)
async def delete_project(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Soft-delete a project (sets deleted_at)."""
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    project.deleted_at = func.now()

    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project_id,
        activity_type="project_deleted",
        actor_type="user",
        metadata_={},
    )
    db.add(activity)
    await db.commit()
    return None


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/cancel
# ---------------------------------------------------------------------------


@projects_router.post("/projects/{project_id}/cancel")
async def cancel_project(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Cancel a project."""
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    validate_transition(project.status, "cancelled", client_id=project.client_id)
    project.status = "cancelled"

    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project_id,
        activity_type="project_cancelled",
        actor_type="user",
        metadata_={},
    )
    db.add(activity)
    await db.commit()
    await db.refresh(project)

    return _project_to_dict(project)


# ---------------------------------------------------------------------------
# PATCH /broker/projects/{project_id}
# ---------------------------------------------------------------------------


@projects_router.patch("/projects/{project_id}")
async def update_project(
    project_id: UUID,
    body: UpdateProjectBody,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Update project fields (name, type, description, market context, etc.)."""
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    update_data = body.model_dump(exclude_unset=True)
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")

    for field, value in update_data.items():
        setattr(project, field, value)

    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project_id,
        activity_type="project_updated",
        actor_type="user",
        metadata_={"updated_fields": list(update_data.keys())},
    )
    db.add(activity)
    await db.commit()
    await db.refresh(project)

    return _project_to_dict(project)


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/approve
# ---------------------------------------------------------------------------


@projects_router.post("/projects/{project_id}/approve")
async def approve_project(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Approve a broker project (Gate 1). Returns 409 if already approved."""
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if project.approval_status == "approved":
        raise HTTPException(status_code=409, detail="Project already approved")

    project.approval_status = "approved"

    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project.id,
        activity_type="project_approved",
        actor_type="user",
    )
    db.add(activity)

    await db.commit()
    await db.refresh(project)
    return _project_to_dict(project)


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/analyze-gaps
# ---------------------------------------------------------------------------


@projects_router.post("/projects/{project_id}/analyze-gaps")
async def analyze_gaps(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Run gap analysis on a project's coverages."""
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    cov_result = await db.execute(
        select(ProjectCoverage)
        .where(ProjectCoverage.broker_project_id == project_id)
        .order_by(ProjectCoverage.created_at)
    )
    coverages = cov_result.scalars().all()
    coverage_dicts = [_coverage_to_dict(c) for c in coverages]

    gap_results = detect_gaps(coverage_dicts)
    summary = summarize_gaps(gap_results)

    result_by_id = {r["id"]: r for r in gap_results}
    for cov_orm in coverages:
        updated = result_by_id.get(str(cov_orm.id))
        if updated is None:
            continue
        cov_orm.gap_status = updated.get("gap_status")
        cov_orm.gap_amount = updated.get("gap_amount")

    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project_id,
        activity_type="gap_analysis_complete",
        actor_type="system",
        description=(
            f"Gap analysis: {summary['covered']} covered, "
            f"{summary['insufficient']} insufficient, "
            f"{summary['missing']} missing"
        ),
        metadata_=summary,
    )
    db.add(activity)
    await db.commit()

    refreshed_dicts = []
    for cov_orm in coverages:
        await db.refresh(cov_orm)
        refreshed_dicts.append(_coverage_to_dict(cov_orm))

    return {"summary": summary, "coverages": refreshed_dicts}


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/documents
# ---------------------------------------------------------------------------


@projects_router.post("/projects/{project_id}/documents", status_code=201)
async def upload_project_documents(
    project_id: UUID,
    files: list[UploadFile] = File(...),
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Upload one or more documents to a broker project."""
    from flywheel.services.document_storage import upload_file as upload_to_storage

    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    uploaded_docs: list[dict[str, Any]] = []
    tenant_id = str(user.tenant_id)

    for file in files:
        content = await file.read()
        size = len(content)

        if size > _MAX_DOC_SIZE:
            raise HTTPException(
                status_code=413,
                detail=f"File '{file.filename}' exceeds 25 MB limit",
            )

        if file.content_type not in _ALLOWED_DOC_TYPES:
            raise HTTPException(
                status_code=415,
                detail=f"File '{file.filename}' has unsupported type '{file.content_type}'. "
                f"Allowed: PDF, PNG, JPEG",
            )

        file_uuid = uuid4()
        filename = file.filename or "unknown"
        storage_path = await upload_to_storage(
            tenant_id=tenant_id,
            file_id=str(file_uuid),
            filename=filename,
            content=content,
            mime_type=file.content_type or "application/octet-stream",
        )

        uploaded_file = UploadedFile(
            tenant_id=user.tenant_id,
            user_id=user.sub,
            filename=filename,
            mimetype=file.content_type or "application/octet-stream",
            size_bytes=size,
            storage_path=storage_path,
        )
        db.add(uploaded_file)
        await db.flush()

        doc_ref = {
            "file_id": str(uploaded_file.id),
            "name": filename,
            "mimetype": file.content_type,
            "size": size,
            "storage_path": storage_path,
            "uploaded_at": uploaded_file.created_at.isoformat() if uploaded_file.created_at else None,
        }
        uploaded_docs.append(doc_ref)

    # Set source_document_id to the first uploaded PDF if not already set
    if not project.source_document_id:
        first_pdf = next(
            (d for d in uploaded_docs if d.get("mimetype") == "application/pdf"),
            None,
        )
        if first_pdf:
            project.source_document_id = UUID(first_pdf["file_id"])

    existing_docs = (project.metadata_ or {}).get("documents", [])
    existing_docs.extend(uploaded_docs)
    project.metadata_ = {**(project.metadata_ or {}), "documents": existing_docs}

    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project_id,
        activity_type="document_uploaded",
        actor_type="user",
        metadata_={"filenames": [f.filename for f in files]},
    )
    db.add(activity)
    await db.commit()

    return {"documents": uploaded_docs, "total": len(uploaded_docs)}


# ---------------------------------------------------------------------------
# POST /broker/projects/from-email
# ---------------------------------------------------------------------------


@projects_router.post("/projects/from-email", status_code=201)
async def create_project_from_email(
    body: CreateFromEmailBody,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Create a broker project from a Gmail message with PDF attachment."""
    from flywheel.services.document_storage import upload_file as storage_upload
    from flywheel.services.gmail_read import (
        find_pdf_attachments,
        get_attachment,
        get_valid_credentials,
    )
    from googleapiclient.discovery import build

    result = await db.execute(
        select(Integration).where(
            Integration.id == body.integration_id,
            Integration.tenant_id == user.tenant_id,
            Integration.provider == "gmail-read",
        )
    )
    integration = result.scalar_one_or_none()
    if integration is None:
        raise HTTPException(404, "Gmail integration not found")

    creds = await get_valid_credentials(integration)

    def _fetch_message():
        service = build("gmail", "v1", credentials=creds)
        return (
            service.users()
            .messages()
            .get(userId="me", id=body.message_id, format="full")
            .execute()
        )

    msg = await asyncio.to_thread(_fetch_message)

    pdf_attachments = find_pdf_attachments(msg)
    if not pdf_attachments:
        raise HTTPException(422, "No PDF attachments found in message")

    if body.attachment_index >= len(pdf_attachments):
        raise HTTPException(
            422,
            f"Attachment index {body.attachment_index} out of range "
            f"(message has {len(pdf_attachments)} PDF attachments)",
        )

    pdf_info = pdf_attachments[body.attachment_index]
    pdf_bytes = await get_attachment(creds, body.message_id, pdf_info["attachment_id"])

    if len(pdf_bytes) > _MAX_DOC_SIZE:
        raise HTTPException(413, "PDF exceeds 25 MB limit")

    file_id = str(uuid4())
    storage_path = await storage_upload(
        tenant_id=str(user.tenant_id),
        file_id=file_id,
        filename=pdf_info["filename"],
        content=pdf_bytes,
        mime_type="application/pdf",
    )

    uploaded_file = UploadedFile(
        tenant_id=user.tenant_id,
        user_id=user.sub,
        filename=pdf_info["filename"],
        mimetype="application/pdf",
        size_bytes=len(pdf_bytes),
        storage_path=storage_path,
    )
    db.add(uploaded_file)
    await db.flush()

    headers = {
        h["name"].lower(): h["value"]
        for h in msg.get("payload", {}).get("headers", [])
    }
    sender = headers.get("from", "Unknown Sender")
    subject = headers.get("subject", "Untitled Project")

    project = BrokerProject(
        tenant_id=user.tenant_id,
        name=subject,
        project_type="construction",
        status="new_request",
        analysis_status="pending",
        import_source="email",
        external_ref=body.message_id,
        source_document_id=uploaded_file.id,
        metadata_={"sender": sender, "gmail_message_id": body.message_id},
    )
    db.add(project)
    await db.flush()

    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project.id,
        activity_type="project_created_from_email",
        actor_type="user",
        metadata_={"message_id": body.message_id, "filename": pdf_info["filename"]},
    )
    db.add(activity)
    await db.commit()
    await db.refresh(project)

    return _project_to_dict(project)


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/analyze
# ---------------------------------------------------------------------------


@projects_router.post("/projects/{project_id}/analyze", status_code=202)
async def trigger_analysis(
    project_id: UUID,
    background_tasks: BackgroundTasks,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
):
    """Trigger async contract analysis for a project."""
    result = await db.execute(
        select(BrokerProject)
        .where(
            BrokerProject.id == project_id,
            BrokerProject.deleted_at.is_(None),
        )
        .with_for_update()
    )
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(404, "Project not found")
    if project.analysis_status == "running":
        raise HTTPException(409, "Analysis already in progress")

    project.analysis_status = "running"
    project.status = "analyzing"
    await db.commit()

    background_tasks.add_task(_run_analysis, project_id, user.tenant_id)
    return {"status": "analyzing", "project_id": str(project_id)}


async def _get_project_pdf(session, project, tenant_id: UUID) -> bytes | None:
    """Retrieve PDF bytes for a project from storage (Supabase or local fallback)."""
    from flywheel.services.document_storage import download_file

    doc_id = project.source_document_id

    # Fallback: if source_document_id is not set, use the first PDF from metadata
    if not doc_id:
        docs = (project.metadata_ or {}).get("documents", [])
        first_pdf = next(
            (d for d in docs if d.get("mimetype") == "application/pdf"),
            None,
        )
        if not first_pdf:
            return None
        from uuid import UUID as _UUID
        doc_id = _UUID(first_pdf["file_id"]) if isinstance(first_pdf["file_id"], str) else first_pdf["file_id"]
        # Backfill the FK so future calls don't need the fallback
        project.source_document_id = doc_id

    result = await session.execute(
        select(UploadedFile).where(
            UploadedFile.id == doc_id,
            UploadedFile.tenant_id == tenant_id,
        )
    )
    uploaded_file = result.scalar_one_or_none()
    if not uploaded_file or not uploaded_file.storage_path:
        return None

    return await download_file(uploaded_file.storage_path)


async def _run_analysis(project_id: UUID, tenant_id: UUID):
    """Background: retrieve PDF from Supabase Storage, run contract_analyzer."""
    import logging as _logging

    from flywheel.db.session import get_session_factory
    from flywheel.engines.contract_analyzer import analyze_contract

    _logger = _logging.getLogger(__name__)
    factory = get_session_factory()

    try:
        async with factory() as session:
            await session.execute(text(f"SET LOCAL app.tenant_id = '{tenant_id}'"))

            project = (
                await session.execute(
                    select(BrokerProject).where(BrokerProject.id == project_id)
                )
            ).scalar_one_or_none()
            if not project:
                return

            pdf_content = await _get_project_pdf(session, project, tenant_id)
            if not pdf_content:
                project.analysis_status = "failed"
                await session.commit()
                return

            await analyze_contract(session, tenant_id, project_id, pdf_content)
            await session.commit()
    except Exception as exc:
        import traceback as _tb
        _logger.error("Analysis failed for project %s: %s\n%s", project_id, exc, _tb.format_exc())
        try:
            async with factory() as err_session:
                await err_session.execute(
                    text(f"SET LOCAL app.tenant_id = '{tenant_id}'")
                )
                proj = (
                    await err_session.execute(
                        select(BrokerProject).where(BrokerProject.id == project_id)
                    )
                ).scalar_one()
                proj.analysis_status = "failed"
                validate_transition(proj.status, "analysis_failed")
                proj.status = "analysis_failed"
                await err_session.commit()
        except Exception:
            _logger.error("Failed to update status for project %s", project_id)


# ---------------------------------------------------------------------------
# GET /broker/projects/{project_id}/carrier-matches
# ---------------------------------------------------------------------------


def _compute_carrier_matches(
    carriers: list,
    project_coverage_keys: list[str],
    contract_value: float | None,
) -> list[dict]:
    """Rank carriers by exact coverage_type_key intersection with project needs.

    Uses canonical coverage_type_key values for exact set intersection —
    no fuzzy normalization needed since keys are already canonical.
    """
    if not project_coverage_keys:
        return []

    project_keys = {ct for ct in project_coverage_keys if ct}

    matches: list[dict] = []
    for carrier in carriers:
        carrier_coverages = carrier.coverage_types or []
        if not carrier_coverages:
            continue

        if contract_value is not None:
            if carrier.min_project_value is not None and contract_value < float(carrier.min_project_value):
                continue
            if carrier.max_project_value is not None and contract_value > float(carrier.max_project_value):
                continue

        carrier_keys = set(carrier_coverages)  # already canonical keys from carrier config
        matched = project_keys & carrier_keys
        if not matched:
            continue

        unmatched = project_keys - matched
        match_score = len(matched) / len(project_keys)

        matches.append({
            "carrier_config_id": str(carrier.id),
            "carrier_name": carrier.carrier_name,
            "carrier_type": carrier.carrier_type,
            "submission_method": carrier.submission_method,
            "portal_url": carrier.portal_url,
            "matched_coverages": sorted(matched),
            "unmatched_coverages": sorted(unmatched),
            "match_score": round(match_score, 2),
            "avg_response_days": float(carrier.avg_response_days) if carrier.avg_response_days is not None else None,
        })

    matches.sort(key=lambda m: (-m["match_score"], m["avg_response_days"] or 999))
    return matches


@projects_router.get("/projects/{project_id}/carrier-matches")
async def get_carrier_matches(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Return carriers ranked by coverage_type intersection with project needs."""
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    cov_result = await db.execute(
        select(ProjectCoverage).where(
            ProjectCoverage.broker_project_id == project_id,
        )
    )
    coverages = cov_result.scalars().all()
    coverage_keys = list({c.coverage_type_key for c in coverages if c.coverage_type_key})

    carrier_result = await db.execute(
        select(CarrierConfig).where(
            CarrierConfig.tenant_id == user.tenant_id,
            CarrierConfig.is_active.is_(True),
        )
    )
    carriers = carrier_result.scalars().all()

    contract_value = float(project.contract_value) if project.contract_value is not None else None
    matches = _compute_carrier_matches(carriers, coverage_keys, contract_value)

    return {"matches": matches, "project_coverage_count": len(coverage_keys)}


# ---------------------------------------------------------------------------
# GET /broker/projects/{project_id}/comparison
# ---------------------------------------------------------------------------


@projects_router.get("/projects/{project_id}/comparison")
async def get_comparison(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Return fresh quote comparison matrix for a project."""
    from flywheel.engines.quote_comparator import compare_quotes

    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    cov_result = await db.execute(
        select(ProjectCoverage).where(
            ProjectCoverage.broker_project_id == project_id
        )
    )
    coverages = cov_result.scalars().all()
    coverage_dicts = [_coverage_to_dict(c) for c in coverages]

    quote_result = await db.execute(
        select(CarrierQuote).where(
            CarrierQuote.broker_project_id == project_id,
            CarrierQuote.status.in_(("extracted", "reviewed", "selected")),
        )
    )
    quotes = quote_result.scalars().all()
    quote_dicts = [_quote_to_dict_minimal(q) for q in quotes]

    comparison = compare_quotes(coverage_dicts, quote_dicts)

    carrier_ids = {q.carrier_config_id for q in quotes if q.carrier_config_id}
    if len(carrier_ids) < 2:
        comparison["partial"] = True

    return comparison


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/export-comparison
# ---------------------------------------------------------------------------


def _build_comparison_xlsx(comparison: dict, project_name: str) -> bytes:
    """Build an Excel workbook from comparison data."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    categories: dict[str, list] = {}
    for cov_data in comparison.get("coverages", []):
        cat = cov_data.get("category", "insurance")
        categories.setdefault(cat, []).append(cov_data)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    exclusion_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for cat_name, cov_list in categories.items():
        ws = wb.create_sheet(title=cat_name.capitalize())

        carrier_map: dict[str, str] = {}
        carrier_names: list[str] = []
        for cov in cov_list:
            for quote in cov.get("quotes", []):
                cid = quote.get("carrier_config_id") or quote.get("carrier_id")
                if cid and cid not in carrier_map:
                    carrier_map[cid] = quote.get("carrier_name", "Unknown")
                    carrier_names.append(quote.get("carrier_name", "Unknown"))

        headers = ["Coverage", "Required Limit"] + carrier_names
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        row = 2
        carrier_id_list = list(carrier_map.keys())
        for cov in cov_list:
            ws.cell(row=row, column=1, value=cov.get("display_name", cov.get("coverage_type", "")))
            req_limit = cov.get("required_limit")
            ws.cell(row=row, column=2, value=float(req_limit) if req_limit is not None else None)

            for q in cov.get("quotes", []):
                cid = q.get("carrier_config_id") or q.get("carrier_id")
                if cid in carrier_map:
                    col_idx = carrier_id_list.index(cid) + 3
                    premium = q.get("premium") or q.get("quoted_premium")
                    ws.cell(row=row, column=col_idx, value=float(premium) if premium is not None else None)

                    if q.get("has_critical_exclusion"):
                        ws.cell(row=row, column=col_idx).fill = exclusion_fill
            row += 1

        for col_cells in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)

    if not categories:
        wb.create_sheet(title="No Data")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@projects_router.post("/projects/{project_id}/export-comparison")
async def export_comparison(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
    body: ExportComparisonBody | None = None,
) -> StreamingResponse:
    """Export quote comparison as .xlsx file."""
    from flywheel.engines.quote_comparator import compare_quotes

    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    cov_result = await db.execute(
        select(ProjectCoverage).where(
            ProjectCoverage.broker_project_id == project_id
        )
    )
    coverages = cov_result.scalars().all()
    coverage_dicts = [_coverage_to_dict(c) for c in coverages]

    quote_query = select(CarrierQuote).where(
        CarrierQuote.broker_project_id == project_id,
        CarrierQuote.status.in_(("extracted", "reviewed", "selected")),
    )
    if body and body.quote_ids:
        quote_query = quote_query.where(CarrierQuote.id.in_(body.quote_ids))

    quote_result = await db.execute(quote_query)
    quotes = quote_result.scalars().all()
    quote_dicts = [_quote_to_dict_minimal(q) for q in quotes]

    comparison = compare_quotes(coverage_dicts, quote_dicts)

    content = await asyncio.to_thread(_build_comparison_xlsx, comparison, project.name or "comparison")

    safe_name = re.sub(r"[^\w\s\-]", "", project.name or "comparison").strip().replace(" ", "_")[:60]
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_comparison.xlsx"'},
    )


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/draft-followups
# ---------------------------------------------------------------------------


@projects_router.post("/projects/{project_id}/draft-followups")
async def draft_followups_endpoint(
    project_id: UUID,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Generate follow-up email drafts for carriers that haven't responded."""
    from flywheel.engines.followup_drafter import draft_followup

    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    quote_result = await db.execute(
        select(CarrierQuote).where(
            CarrierQuote.broker_project_id == project_id,
            CarrierQuote.status == "solicited",
        )
    )
    solicited_quotes = quote_result.scalars().all()

    followups: list[dict] = []
    not_due: list[dict] = []
    skipped: list[dict] = []

    for quote in solicited_quotes:
        try:
            if not quote.carrier_config_id:
                skipped.append({
                    "quote_id": str(quote.id),
                    "carrier_name": quote.carrier_name,
                    "reason": "No carrier_config_id",
                })
                continue

            carrier_result = await db.execute(
                select(CarrierConfig).where(CarrierConfig.id == quote.carrier_config_id)
            )
            carrier = carrier_result.scalar_one_or_none()
            if carrier is None:
                skipped.append({
                    "quote_id": str(quote.id),
                    "carrier_name": quote.carrier_name,
                    "reason": "Carrier config not found",
                })
                continue

            draft_result = await draft_followup(
                db, user.tenant_id, quote, project, carrier
            )

            if draft_result["status"] == "not_due":
                not_due.append({
                    "quote_id": str(quote.id),
                    "carrier_name": quote.carrier_name,
                    "days_remaining": draft_result["days_remaining"],
                })
            else:
                # Store follow-up as a SolicitationDraft if carrier contact email exists
                followups.append({
                    "quote_id": str(quote.id),
                    "carrier_name": quote.carrier_name,
                    "subject": draft_result["subject"],
                    "body": draft_result["body"],
                })

        except Exception as exc:
            logger.warning("Follow-up draft failed for quote %s: %s", quote.id, exc)
            skipped.append({
                "quote_id": str(quote.id),
                "carrier_name": quote.carrier_name,
                "reason": str(exc),
            })

    await db.commit()

    return {"followups": followups, "not_due": not_due, "skipped": skipped}


# ---------------------------------------------------------------------------
# PATCH /broker/coverages/{coverage_id}
# ---------------------------------------------------------------------------


@projects_router.patch("/coverages/{coverage_id}")
async def update_coverage(
    coverage_id: UUID,
    body: UpdateCoverageBody,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Update a coverage requirement inline. Sets is_manual_override=True."""
    result = await db.execute(
        select(ProjectCoverage).where(ProjectCoverage.id == coverage_id)
    )
    coverage = result.scalar_one_or_none()
    if coverage is None:
        raise HTTPException(status_code=404, detail="Coverage not found")

    update_data = body.model_dump(exclude_unset=True)
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")

    for field, value in update_data.items():
        setattr(coverage, field, value)

    coverage.is_manual_override = True

    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=coverage.broker_project_id,
        activity_type="coverage_updated",
        actor_type="user",
        metadata_={"updated_fields": list(update_data.keys())},
    )
    db.add(activity)
    await db.commit()
    await db.refresh(coverage)

    return _coverage_to_dict(coverage)


# ---------------------------------------------------------------------------
# POST /broker/projects/{project_id}/coverages  (batch create)
# ---------------------------------------------------------------------------


@projects_router.post("/projects/{project_id}/coverages")
async def batch_create_coverages(
    project_id: UUID,
    body: dict,
    user: TokenPayload = Depends(require_module("broker")),
    db: AsyncSession = Depends(get_tenant_db),
) -> dict[str, Any]:
    """Create multiple coverages for a project in a single request.

    Expects body: {"coverages": [{...}, {...}, ...]}
    """
    # Validate project exists and belongs to tenant
    result = await db.execute(
        select(BrokerProject).where(
            BrokerProject.id == project_id,
            BrokerProject.tenant_id == user.tenant_id,
            BrokerProject.deleted_at.is_(None),
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    coverage_items = body.get("coverages")
    if not coverage_items or not isinstance(coverage_items, list):
        raise HTTPException(
            status_code=400,
            detail="Request body must contain a 'coverages' array",
        )

    created: list[ProjectCoverage] = []
    for item in coverage_items:
        cov = ProjectCoverage(
            id=uuid4(),
            tenant_id=user.tenant_id,
            broker_project_id=project_id,
            coverage_type=item.get("coverage_type", "unknown"),
            category=item.get("category"),
            display_name=item.get("display_name"),
            required_limit=item.get("required_limit"),
            required_deductible=item.get("required_deductible"),
            required_terms=item.get("required_terms"),
            contract_clause=item.get("contract_clause"),
            current_limit=item.get("current_limit"),
            current_carrier=item.get("current_carrier"),
            current_policy_number=item.get("current_policy_number"),
            current_expiry=item.get("current_expiry"),
            gap_status=item.get("gap_status"),
            gap_amount=item.get("gap_amount"),
            gap_notes=item.get("gap_notes"),
            source=item.get("source", "manual"),
            source_excerpt=item.get("source_excerpt"),
            source_page=item.get("source_page"),
            source_section=item.get("source_section"),
            confidence=item.get("confidence"),
        )
        db.add(cov)
        created.append(cov)

    await db.flush()

    activity = BrokerActivity(
        tenant_id=user.tenant_id,
        broker_project_id=project_id,
        activity_type="coverages_created",
        actor_type="system",
        description=f"Batch created {len(created)} coverages",
    )
    db.add(activity)
    await db.commit()

    for cov in created:
        await db.refresh(cov)

    return {
        "items": [_coverage_to_dict(c) for c in created],
        "created": len(created),
    }
