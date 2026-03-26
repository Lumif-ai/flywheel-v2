"""Seed CRM tables from GTM stack files.

Reads xlsx, csv, and json files from the GTM stack directory and populates
the Account, AccountContact, and OutreachActivity tables for a given tenant.
Designed to be idempotent: running the command multiple times produces no
duplicate rows.

Usage:
    python -m flywheel.seed_crm --tenant-id <uuid> [--gtm-dir ~/.claude/gtm-stack/] [--dry-run] [--verbose]
    python -m flywheel.seed_crm --tenant-id <uuid> --force   # delete existing and re-seed
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import logging
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from flywheel.db.models import Account, AccountContact, OutreachActivity
from flywheel.utils.normalize import normalize_company_name

logger = logging.getLogger("flywheel.seed_crm")


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class AccountData:
    """Accumulated data for a single company/account."""

    name: str                        # Most complete raw name encountered
    normalized_name: str
    domain: str | None = None
    fit_score: float | None = None
    fit_tier: str | None = None
    industry: str | None = None
    description: str | None = None
    fit_reasoning: str | None = None
    sources: list[str] = field(default_factory=list)


@dataclass
class ContactData:
    """Data for a single person/contact."""

    name: str
    normalized_company: str          # key into account map
    email: str | None = None
    title: str | None = None
    linkedin_url: str | None = None
    notes: str | None = None
    source: str = "unknown"


@dataclass
class ActivityData:
    """Data for a single outreach activity."""

    normalized_company: str          # key into account map
    contact_email: str | None        # key into contact map (may be None)
    contact_name: str | None         # fallback key when email is None
    channel: str
    direction: str
    status: str
    sent_at: datetime | None


@dataclass
class SeedSummary:
    """Summary of a seeding run."""

    accounts_inserted: int = 0
    accounts_updated: int = 0
    accounts_skipped: int = 0
    contacts_inserted: int = 0
    contacts_skipped: int = 0
    activities_inserted: int = 0
    activities_skipped: int = 0
    parse_errors: list[str] = field(default_factory=list)
    files_parsed: list[str] = field(default_factory=list)
    files_skipped: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _parse_float(value: Any) -> float | None:
    """Parse a float from int, float, or string; return None on failure."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return None


def _parse_date(value: str | None) -> datetime | None:
    """Parse a date string (YYYY-MM-DD) into a timezone-aware datetime."""
    if not value or not str(value).strip():
        return None
    val = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            dt = datetime.strptime(val[:10], fmt[:8] if len(fmt) > 8 else fmt)
            return dt.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    try:
        dt = datetime.strptime(val[:10], "%Y-%m-%d")
        return dt.replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def _merge_account(existing: AccountData, new_name: str, new_source: str,
                   fit_score: float | None = None,
                   fit_tier: str | None = None,
                   domain: str | None = None,
                   industry: str | None = None,
                   description: str | None = None,
                   fit_reasoning: str | None = None) -> None:
    """Merge new data into an existing AccountData in-place."""
    # Keep the longer (more complete) name
    if len(new_name) > len(existing.name):
        existing.name = new_name

    # Add source if not already present
    if new_source and new_source not in existing.sources:
        existing.sources.append(new_source)

    # Keep higher fit_score
    if fit_score is not None:
        if existing.fit_score is None or fit_score > existing.fit_score:
            existing.fit_score = fit_score

    # Keep fit_tier from highest-scoring source
    if fit_tier and existing.fit_tier is None:
        existing.fit_tier = fit_tier

    # Keep first non-null domain
    if domain and not existing.domain:
        existing.domain = domain

    # Merge intel fields — first non-null wins
    if industry and not existing.industry:
        existing.industry = industry
    if description and not existing.description:
        existing.description = description
    if fit_reasoning and not existing.fit_reasoning:
        existing.fit_reasoning = fit_reasoning


# ---------------------------------------------------------------------------
# File parsers
# ---------------------------------------------------------------------------

def parse_gtm_xlsx(
    xlsx_path: Path,
    accounts: dict[str, AccountData],
    contacts: list[ContactData],
    summary: SeedSummary,
    verbose: bool = False,
) -> None:
    """Parse gtm-leads-master.xlsx.

    Reads:
    - Sheet 'Company Summary': Company, Industry, Best Score → Account rows
    - Sheet 'All Leads Scored': Name, Title, Company, Email, Score, Tier → Contact rows
    """
    try:
        import openpyxl
    except ImportError:
        summary.parse_errors.append("openpyxl not installed — cannot parse xlsx files")
        return

    if not xlsx_path.exists():
        summary.files_skipped.append(str(xlsx_path))
        if verbose:
            logger.info("xlsx not found, skipping: %s", xlsx_path)
        return

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        summary.parse_errors.append(f"Cannot open {xlsx_path}: {e}")
        return

    summary.files_parsed.append(str(xlsx_path))

    # --- Sheet: Company Summary ---
    if "Company Summary" in wb.sheetnames:
        ws = wb["Company Summary"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            pass
        else:
            # Detect header row
            header = [str(c).strip() if c else "" for c in rows[0]]
            col = {h: i for i, h in enumerate(header)}
            company_col = col.get("Company", 0)
            industry_col = col.get("Industry")
            score_col = col.get("Best Score")

            for row in rows[1:]:
                if not row or not row[company_col]:
                    continue
                raw_name = str(row[company_col]).strip()
                norm = normalize_company_name(raw_name)
                if not norm:
                    continue
                industry = str(row[industry_col]).strip() if industry_col is not None and row[industry_col] else None
                fit_score = _parse_float(row[score_col]) if score_col is not None else None

                if norm in accounts:
                    _merge_account(accounts[norm], raw_name, "gtm-leads-master",
                                   fit_score=fit_score, industry=industry)
                else:
                    accounts[norm] = AccountData(
                        name=raw_name,
                        normalized_name=norm,
                        fit_score=fit_score,
                        industry=industry,
                        sources=["gtm-leads-master"],
                    )
    else:
        logger.warning("Sheet 'Company Summary' not found in %s", xlsx_path)

    # --- Sheet: All Leads Scored ---
    if "All Leads Scored" in wb.sheetnames:
        ws = wb["All Leads Scored"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(c).strip() if c else "" for c in rows[0]]
            col = {h: i for i, h in enumerate(header)}

            name_col = col.get("Name", 0)
            title_col = col.get("Title")
            company_col = col.get("Company")
            email_col = col.get("Email")
            score_col = col.get("Score")
            tier_col = col.get("Tier")
            industry_col = col.get("Industry Tag")

            for row in rows[1:]:
                if not row or not row[name_col]:
                    continue
                raw_contact_name = str(row[name_col]).strip()
                raw_company = str(row[company_col]).strip() if company_col is not None and row[company_col] else None
                if not raw_company:
                    continue
                norm = normalize_company_name(raw_company)
                if not norm:
                    continue

                email = str(row[email_col]).strip() if email_col is not None and row[email_col] else None
                title = str(row[title_col]).strip() if title_col is not None and row[title_col] else None
                fit_score = _parse_float(row[score_col]) if score_col is not None else None
                fit_tier = str(row[tier_col]).strip() if tier_col is not None and row[tier_col] else None
                industry = str(row[industry_col]).strip() if industry_col is not None and row[industry_col] else None

                # Ensure account exists
                if norm in accounts:
                    _merge_account(accounts[norm], raw_company, "gtm-leads-master",
                                   fit_score=fit_score, fit_tier=fit_tier, industry=industry)
                else:
                    accounts[norm] = AccountData(
                        name=raw_company,
                        normalized_name=norm,
                        fit_score=fit_score,
                        fit_tier=fit_tier,
                        industry=industry,
                        sources=["gtm-leads-master"],
                    )

                # Create contact
                contacts.append(ContactData(
                    name=raw_contact_name,
                    normalized_company=norm,
                    email=email if email and email.lower() not in ("", "n/a", "none") else None,
                    title=title,
                    source="gtm-leads-master",
                ))
    else:
        logger.warning("Sheet 'All Leads Scored' not found in %s", xlsx_path)

    wb.close()


def parse_outreach_tracker(
    csv_path: Path,
    accounts: dict[str, AccountData],
    contacts: list[ContactData],
    activities: list[ActivityData],
    summary: SeedSummary,
    verbose: bool = False,
) -> None:
    """Parse outreach-tracker.csv.

    Columns: Name, First_Name, Email, Company, Title, Score, Tier,
             Industry_Tag, Email_Status, Email_Sent, Email_Sent_Date,
             LinkedIn_Status, LinkedIn_Date, Notes
    """
    if not csv_path.exists():
        summary.files_skipped.append(str(csv_path))
        if verbose:
            logger.info("outreach tracker not found, skipping: %s", csv_path)
        return

    summary.files_parsed.append(str(csv_path))
    try:
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row_num, row in enumerate(reader, 1):
                raw_company = (row.get("Company") or "").strip()
                if not raw_company:
                    continue
                norm = normalize_company_name(raw_company)
                if not norm:
                    continue

                raw_name = (row.get("Name") or "").strip()
                email = (row.get("Email") or "").strip() or None
                title = (row.get("Title") or "").strip() or None
                notes = (row.get("Notes") or "").strip() or None
                tier = (row.get("Tier") or "").strip() or None
                industry = (row.get("Industry_Tag") or "").strip() or None
                fit_score = _parse_float(row.get("Score"))

                # Ensure account exists
                if norm in accounts:
                    _merge_account(accounts[norm], raw_company, "outreach-tracker",
                                   fit_score=fit_score, fit_tier=tier, industry=industry)
                else:
                    accounts[norm] = AccountData(
                        name=raw_company,
                        normalized_name=norm,
                        fit_score=fit_score,
                        fit_tier=tier,
                        industry=industry,
                        sources=["outreach-tracker"],
                    )

                # Contact
                if raw_name:
                    contacts.append(ContactData(
                        name=raw_name,
                        normalized_company=norm,
                        email=email,
                        title=title,
                        notes=notes,
                        source="outreach-tracker",
                    ))

                # Email activity
                email_sent = (row.get("Email_Sent") or "").strip().lower()
                email_sent_date = (row.get("Email_Sent_Date") or "").strip()
                if email_sent == "yes" and email_sent_date:
                    sent_at = _parse_date(email_sent_date)
                    if sent_at:
                        activities.append(ActivityData(
                            normalized_company=norm,
                            contact_email=email,
                            contact_name=raw_name,
                            channel="email",
                            direction="outbound",
                            status="sent",
                            sent_at=sent_at,
                        ))

                # LinkedIn activity
                li_status = (row.get("LinkedIn_Status") or "").strip()
                li_date = (row.get("LinkedIn_Date") or "").strip()
                if li_status and li_status.lower() not in ("", "n/a", "not_found", "not found"):
                    sent_at = _parse_date(li_date) if li_date else None
                    if sent_at:  # Only create activity if we have a date
                        activities.append(ActivityData(
                            normalized_company=norm,
                            contact_email=email,
                            contact_name=raw_name,
                            channel="linkedin",
                            direction="outbound",
                            status=li_status.lower(),
                            sent_at=sent_at,
                        ))
    except Exception as e:
        summary.parse_errors.append(f"Error parsing {csv_path}: {e}")


def parse_scored_csvs(
    pipeline_runs_path: Path,
    downloads_dir: Path,
    accounts: dict[str, AccountData],
    contacts: list[ContactData],
    summary: SeedSummary,
    verbose: bool = False,
) -> None:
    """Parse scored CSVs referenced from pipeline-runs.json.

    Columns include: Name, Title, Company, Email, Fit_Score, Fit_Tier,
                     Fit_Reasoning, Company_Description, Company_Website, LinkedIn_URL
    """
    if not pipeline_runs_path.exists():
        summary.files_skipped.append(str(pipeline_runs_path))
        if verbose:
            logger.info("pipeline-runs.json not found, skipping: %s", pipeline_runs_path)
        return

    summary.files_parsed.append(str(pipeline_runs_path))
    try:
        with open(pipeline_runs_path, encoding="utf-8") as f:
            runs = json.load(f)
    except Exception as e:
        summary.parse_errors.append(f"Error parsing {pipeline_runs_path}: {e}")
        return

    if not isinstance(runs, list):
        summary.parse_errors.append(f"pipeline-runs.json is not a list")
        return

    for run in runs:
        if not isinstance(run, dict):
            continue

        status = run.get("status", "")
        # Handle runs with 'status' field (older format)
        if status and status.lower() != "complete":
            continue

        # Some runs don't have explicit status — treat those as complete
        run_id = run.get("id") or run.get("run_id") or "unknown"
        source_label = f"pipeline-{run_id}"

        # Find csv_path — supports both 'csv_path' and 'scored_csv' keys
        csv_path_raw = run.get("csv_path") or run.get("scored_csv") or ""
        if not csv_path_raw:
            if verbose:
                logger.info("Run %s has no csv_path, skipping", run_id)
            continue

        # Resolve ~ and relative paths
        csv_path = Path(csv_path_raw).expanduser()
        if not csv_path.is_absolute():
            csv_path = downloads_dir / csv_path

        if not csv_path.exists():
            summary.files_skipped.append(str(csv_path))
            logger.warning("Scored CSV not found: %s (run: %s)", csv_path, run_id)
            continue

        summary.files_parsed.append(str(csv_path))
        try:
            _parse_single_scored_csv(csv_path, source_label, accounts, contacts, verbose)
        except Exception as e:
            summary.parse_errors.append(f"Error parsing {csv_path}: {e}")
            if verbose:
                logger.exception("Error parsing scored CSV: %s", csv_path)


def _parse_single_scored_csv(
    csv_path: Path,
    source_label: str,
    accounts: dict[str, AccountData],
    contacts: list[ContactData],
    verbose: bool = False,
) -> None:
    """Parse a single scored CSV file and merge data into accounts/contacts."""
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []

        # Build case-insensitive column map
        col_map = {h.strip().lower(): h.strip() for h in fieldnames}

        def get(row: dict, *keys: str) -> str | None:
            """Get first non-empty value from row using case-insensitive key lookup."""
            for k in keys:
                actual = col_map.get(k.lower())
                if actual and row.get(actual, "").strip():
                    return row[actual].strip()
            return None

        for row in reader:
            raw_company = get(row, "Company", "company") or ""
            if not raw_company:
                continue
            norm = normalize_company_name(raw_company)
            if not norm:
                continue

            fit_score = _parse_float(get(row, "Fit_Score", "Score", "score"))
            fit_tier = get(row, "Fit_Tier", "Tier", "tier")
            fit_reasoning = get(row, "Fit_Reasoning", "Rationale")
            description = get(row, "Company_Description", "Description")
            domain = get(row, "Company_Website", "Website", "domain")
            # Normalize domain — keep just the URL
            if domain and domain.startswith("http"):
                # Extract just the host
                from urllib.parse import urlparse
                parsed = urlparse(domain)
                domain = parsed.netloc or domain

            if norm in accounts:
                _merge_account(
                    accounts[norm], raw_company, source_label,
                    fit_score=fit_score, fit_tier=fit_tier, domain=domain,
                    description=description, fit_reasoning=fit_reasoning,
                )
            else:
                accounts[norm] = AccountData(
                    name=raw_company,
                    normalized_name=norm,
                    fit_score=fit_score,
                    fit_tier=fit_tier,
                    domain=domain,
                    description=description,
                    fit_reasoning=fit_reasoning,
                    sources=[source_label],
                )

            # Contact
            raw_name = get(row, "Name", "name") or ""
            email = get(row, "Email", "email")
            title = get(row, "Title", "title")
            linkedin_url = get(row, "LinkedIn_URL", "LinkedIn", "linkedin_url")

            if raw_name:
                contacts.append(ContactData(
                    name=raw_name,
                    normalized_company=norm,
                    email=email,
                    title=title,
                    linkedin_url=linkedin_url if linkedin_url and linkedin_url.startswith("http") else None,
                    source=source_label,
                ))


# ---------------------------------------------------------------------------
# Database upsert logic
# ---------------------------------------------------------------------------

async def upsert_accounts(
    session: AsyncSession,
    accounts: dict[str, AccountData],
    tenant_id: str,
    dry_run: bool,
    verbose: bool,
) -> dict[str, Any]:
    """Upsert all accounts. Return mapping of normalized_name -> account_id."""
    from sqlalchemy.dialects.postgresql import insert as pg_insert

    account_id_map: dict[str, Any] = {}
    inserted = updated = skipped = 0

    for norm, data in accounts.items():
        intel: dict[str, Any] = {}
        if data.industry:
            intel["industry"] = data.industry
        if data.description:
            intel["description"] = data.description
        if data.fit_reasoning:
            intel["fit_reasoning"] = data.fit_reasoning

        source_str = ",".join(sorted(set(data.sources))) if data.sources else "unknown"

        if dry_run:
            # In dry-run, generate a fake UUID for downstream counting
            import uuid
            account_id_map[norm] = str(uuid.uuid4())
            continue

        stmt = pg_insert(Account).values(
            tenant_id=tenant_id,
            name=data.name,
            normalized_name=norm,
            domain=data.domain,
            fit_score=data.fit_score,
            fit_tier=data.fit_tier,
            intel=intel,
            source=source_str,
            status="prospect",
        )

        # ON CONFLICT: merge — keep higher fit_score, merge intel, keep first domain/name
        stmt = stmt.on_conflict_do_update(
            constraint="uq_account_tenant_normalized",
            set_={
                # Keep the longer/better name (use case: first insert wins unless new is longer)
                "name": text(
                    "CASE WHEN length(EXCLUDED.name) > length(accounts.name) "
                    "THEN EXCLUDED.name ELSE accounts.name END"
                ),
                # Keep higher fit_score
                "fit_score": text(
                    "CASE WHEN EXCLUDED.fit_score IS NOT NULL AND "
                    "(accounts.fit_score IS NULL OR EXCLUDED.fit_score > accounts.fit_score) "
                    "THEN EXCLUDED.fit_score ELSE accounts.fit_score END"
                ),
                # Keep fit_tier if not set
                "fit_tier": text(
                    "CASE WHEN accounts.fit_tier IS NULL THEN EXCLUDED.fit_tier "
                    "ELSE accounts.fit_tier END"
                ),
                # Keep first non-null domain
                "domain": text(
                    "CASE WHEN accounts.domain IS NULL THEN EXCLUDED.domain "
                    "ELSE accounts.domain END"
                ),
                # Merge intel JSONB
                "intel": text("accounts.intel || EXCLUDED.intel"),
                # Append new source
                "source": text(
                    "CASE WHEN accounts.source NOT LIKE '%' || EXCLUDED.source || '%' "
                    "THEN accounts.source || ',' || EXCLUDED.source "
                    "ELSE accounts.source END"
                ),
                "updated_at": text("now()"),
            },
        )
        stmt = stmt.returning(Account.id)

        result = await session.execute(stmt)
        account_id = result.scalar_one()
        account_id_map[norm] = account_id

        if verbose:
            logger.debug("Account upsert: %s → %s", norm, account_id)

    if not dry_run:
        await session.flush()

    return account_id_map


async def upsert_contacts(
    session: AsyncSession,
    contacts: list[ContactData],
    account_id_map: dict[str, Any],
    tenant_id: str,
    dry_run: bool,
    verbose: bool,
) -> tuple[dict[tuple, Any], int, int]:
    """Insert contacts. Return (contact_key_map, inserted, skipped).

    contact_key_map maps (account_id, email_or_name) → contact_id.
    """
    contact_id_map: dict[tuple, Any] = {}
    inserted = skipped = 0

    # Deduplicate input contacts before hitting the DB
    # Key: (normalized_company, email_lower_or_name_lower)
    seen_keys: set[tuple[str, str]] = set()
    deduped: list[ContactData] = []
    for c in contacts:
        dedup_key = (
            c.normalized_company,
            (c.email or "").lower() if c.email else c.name.lower()
        )
        if dedup_key not in seen_keys:
            seen_keys.add(dedup_key)
            deduped.append(c)

    for c in deduped:
        account_id = account_id_map.get(c.normalized_company)
        if account_id is None:
            if verbose:
                logger.debug("No account found for contact: %s (company: %s)", c.name, c.normalized_company)
            skipped += 1
            continue

        if dry_run:
            import uuid
            contact_id = str(uuid.uuid4())
            if c.email:
                contact_id_map[(account_id, c.email.lower())] = contact_id
            else:
                contact_id_map[(account_id, c.name.lower())] = contact_id
            inserted += 1
            continue

        # SELECT first
        if c.email:
            stmt = select(AccountContact).where(
                AccountContact.tenant_id == tenant_id,
                AccountContact.account_id == account_id,
                AccountContact.email == c.email,
            )
        else:
            stmt = select(AccountContact).where(
                AccountContact.tenant_id == tenant_id,
                AccountContact.account_id == account_id,
                AccountContact.name == c.name,
            )
        result = await session.execute(stmt)
        existing = result.scalar_one_or_none()

        if existing:
            # Update notes/linkedin if we have new data
            if c.notes and not existing.notes:
                existing.notes = c.notes
            if c.linkedin_url and not existing.linkedin_url:
                existing.linkedin_url = c.linkedin_url
            if c.title and not existing.title:
                existing.title = c.title
            contact_id = existing.id
            skipped += 1
        else:
            new_contact = AccountContact(
                tenant_id=tenant_id,
                account_id=account_id,
                name=c.name,
                email=c.email,
                title=c.title,
                linkedin_url=c.linkedin_url,
                notes=c.notes,
                source=c.source,
            )
            session.add(new_contact)
            await session.flush()
            contact_id = new_contact.id
            inserted += 1

        if c.email:
            contact_id_map[(account_id, c.email.lower())] = contact_id
        else:
            contact_id_map[(account_id, c.name.lower())] = contact_id

        if verbose:
            logger.debug(
                "Contact %s: %s (%s)",
                "INSERT" if not existing else "SKIP",
                c.name,
                c.email or "no email",
            )

    return contact_id_map, inserted, skipped


async def insert_activities(
    session: AsyncSession,
    activities: list[ActivityData],
    account_id_map: dict[str, Any],
    contact_id_map: dict[tuple, Any],
    tenant_id: str,
    dry_run: bool,
    verbose: bool,
) -> tuple[int, int]:
    """Insert outreach activities. Return (inserted, skipped)."""
    inserted = skipped = 0

    # Deduplicate input activities before hitting the DB
    seen_keys: set[tuple] = set()
    deduped: list[ActivityData] = []
    for a in activities:
        if a.sent_at is None:
            skipped += 1
            continue
        dedup_key = (a.normalized_company, a.channel, a.sent_at.date() if a.sent_at else None,
                     (a.contact_email or "").lower())
        if dedup_key not in seen_keys:
            seen_keys.add(dedup_key)
            deduped.append(a)

    for a in deduped:
        account_id = account_id_map.get(a.normalized_company)
        if account_id is None:
            skipped += 1
            continue

        if a.sent_at is None:
            skipped += 1
            continue

        # Resolve contact_id
        contact_id = None
        if a.contact_email:
            contact_id = contact_id_map.get((account_id, a.contact_email.lower()))
        if contact_id is None and a.contact_name:
            contact_id = contact_id_map.get((account_id, a.contact_name.lower()))

        if dry_run:
            inserted += 1
            continue

        # SELECT first — dedup by (account_id, channel, sent_at date)
        stmt = select(OutreachActivity).where(
            OutreachActivity.tenant_id == tenant_id,
            OutreachActivity.account_id == account_id,
            OutreachActivity.channel == a.channel,
            # Cast sent_at to date for comparison (same day = same activity)
            text("DATE(outreach_activities.sent_at AT TIME ZONE 'UTC') = :sent_date").bindparams(
                sent_date=a.sent_at.date()
            ),
        )
        result = await session.execute(stmt)
        existing = result.scalar_one_or_none()

        if existing:
            skipped += 1
        else:
            new_activity = OutreachActivity(
                tenant_id=tenant_id,
                account_id=account_id,
                contact_id=contact_id,
                channel=a.channel,
                direction=a.direction,
                status=a.status,
                sent_at=a.sent_at,
                metadata_={},
            )
            session.add(new_activity)
            inserted += 1

        if verbose:
            logger.debug(
                "Activity %s: %s / %s / %s",
                "INSERT" if not existing else "SKIP",
                a.normalized_company,
                a.channel,
                a.sent_at.date() if a.sent_at else "no date",
            )

    if not dry_run:
        await session.flush()

    return inserted, skipped


# ---------------------------------------------------------------------------
# Core seed function
# ---------------------------------------------------------------------------

async def seed_crm(
    tenant_id: str,
    gtm_dir: Path,
    downloads_dir: Path,
    database_url: str,
    dry_run: bool = False,
    force: bool = False,
    verbose: bool = False,
) -> SeedSummary:
    """Seed CRM tables from GTM stack files.

    Args:
        tenant_id: Target tenant UUID.
        gtm_dir: Path to GTM stack directory.
        downloads_dir: Path to scored CSV downloads directory.
        database_url: Async Postgres connection URL.
        dry_run: Parse and report counts without inserting.
        force: Delete existing seeded data for tenant before re-inserting.
        verbose: Detailed per-record output.

    Returns:
        SeedSummary with counts and errors.
    """
    summary = SeedSummary()

    # Step 1: Parse all files into intermediate data structures
    accounts: dict[str, AccountData] = {}  # normalized_name -> AccountData
    contacts: list[ContactData] = []
    activities: list[ActivityData] = []

    # 1a. Parse gtm-leads-master.xlsx
    xlsx_path = gtm_dir / "gtm-leads-master.xlsx"
    parse_gtm_xlsx(xlsx_path, accounts, contacts, summary, verbose=verbose)

    # 1b. Parse outreach-tracker.csv
    outreach_csv = gtm_dir / "outreach-tracker.csv"
    parse_outreach_tracker(outreach_csv, accounts, contacts, activities, summary, verbose=verbose)

    # 1c. Parse scored CSVs from pipeline-runs.json
    pipeline_runs = gtm_dir / "pipeline-runs.json"
    parse_scored_csvs(pipeline_runs, downloads_dir, accounts, contacts, summary, verbose=verbose)

    # Report parse counts
    print(f"\nParsed from files:")
    print(f"  Accounts (unique companies): {len(accounts)}")
    print(f"  Contacts (raw):              {len(contacts)}")
    print(f"  Activities (raw):            {len(activities)}")
    print(f"  Files parsed:                {len(summary.files_parsed)}")
    print(f"  Files skipped:               {len(summary.files_skipped)}")

    if summary.parse_errors:
        print(f"\nParse errors ({len(summary.parse_errors)}):")
        for err in summary.parse_errors:
            print(f"  - {err}")

    if dry_run:
        print("\n[DRY RUN] No database changes made.")
        return summary

    # Step 2: Connect to database
    engine = create_async_engine(database_url, echo=False)
    session_factory = async_sessionmaker(
        bind=engine, class_=AsyncSession, expire_on_commit=False
    )

    try:
        async with session_factory() as session:
            # Step 3: Force mode — delete existing data
            if force:
                from sqlalchemy import delete
                print(f"\n[FORCE] Deleting existing data for tenant {tenant_id}...")
                await session.execute(
                    delete(OutreachActivity).where(
                        OutreachActivity.tenant_id == tenant_id
                    )
                )
                await session.execute(
                    delete(AccountContact).where(
                        AccountContact.tenant_id == tenant_id
                    )
                )
                await session.execute(
                    delete(Account).where(
                        Account.tenant_id == tenant_id
                    )
                )
                await session.flush()
                print("[FORCE] Existing data deleted.")

            # Step 4: Upsert accounts
            account_id_map = await upsert_accounts(
                session, accounts, tenant_id, dry_run=False, verbose=verbose
            )
            summary.accounts_inserted = len(account_id_map)

            # Step 5: Upsert contacts
            contact_id_map, c_inserted, c_skipped = await upsert_contacts(
                session, contacts, account_id_map, tenant_id, dry_run=False, verbose=verbose
            )
            summary.contacts_inserted = c_inserted
            summary.contacts_skipped = c_skipped

            # Step 6: Insert outreach activities
            a_inserted, a_skipped = await insert_activities(
                session, activities, account_id_map, contact_id_map,
                tenant_id, dry_run=False, verbose=verbose
            )
            summary.activities_inserted = a_inserted
            summary.activities_skipped = a_skipped

            await session.commit()
            print("\nDatabase commit complete.")

    finally:
        await engine.dispose()

    return summary


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _print_summary(summary: SeedSummary, dry_run: bool) -> None:
    """Print human-readable seed summary."""
    mode = "Dry-run" if dry_run else "Seeding"
    print(f"\n{mode} complete:")
    print(f"  Accounts:   {summary.accounts_inserted} upserted")
    print(f"  Contacts:   {summary.contacts_inserted} inserted, {summary.contacts_skipped} skipped")
    print(f"  Activities: {summary.activities_inserted} inserted, {summary.activities_skipped} skipped")

    if summary.files_skipped:
        print(f"\nFiles not found ({len(summary.files_skipped)}):")
        for f in summary.files_skipped:
            print(f"  - {f}")

    if summary.parse_errors:
        print(f"\nParse errors ({len(summary.parse_errors)}):")
        for err in summary.parse_errors:
            print(f"  - {err}")


def build_parser() -> argparse.ArgumentParser:
    """Build the argument parser."""
    parser = argparse.ArgumentParser(
        description="Seed CRM tables from GTM stack files (xlsx, csv, json).",
        prog="python -m flywheel.seed_crm",
    )
    parser.add_argument(
        "--tenant-id",
        required=True,
        help="Target tenant UUID",
    )
    parser.add_argument(
        "--gtm-dir",
        type=Path,
        default=Path.home() / ".claude" / "gtm-stack",
        help="Directory containing GTM stack files (default: ~/.claude/gtm-stack/)",
    )
    parser.add_argument(
        "--downloads-dir",
        type=Path,
        default=Path.home() / "Downloads",
        help="Directory containing scored CSV files (default: ~/Downloads/)",
    )
    parser.add_argument(
        "--database-url",
        default=None,
        help="Async Postgres URL (default: from FLYWHEEL settings)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse files and print counts without touching the database",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Delete existing seeded data for this tenant before re-inserting",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print per-record details during processing",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    """CLI entry point."""
    parser = build_parser()
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s: %(message)s",
    )

    gtm_dir = args.gtm_dir.expanduser()
    if not gtm_dir.is_dir():
        print(f"Error: GTM directory does not exist: {gtm_dir}", file=sys.stderr)
        return 1

    db_url = args.database_url
    if db_url is None:
        if not args.dry_run:
            from flywheel.config import settings
            db_url = settings.database_url
        else:
            # In dry-run mode, we don't need a real DB URL
            db_url = "postgresql+asyncpg://localhost/flywheel"

    downloads_dir = args.downloads_dir.expanduser()

    summary = asyncio.run(
        seed_crm(
            tenant_id=args.tenant_id,
            gtm_dir=gtm_dir,
            downloads_dir=downloads_dir,
            database_url=db_url,
            dry_run=args.dry_run,
            force=args.force,
            verbose=args.verbose,
        )
    )

    _print_summary(summary, args.dry_run)

    return 1 if summary.parse_errors else 0


if __name__ == "__main__":
    sys.exit(main())
