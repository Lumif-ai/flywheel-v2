#!/usr/bin/env python3
"""
Generate the GTM Command Center dashboard from the master workbook.

Primary source: ~/.claude/gtm-stack/gtm-leads-master.xlsx
Fallback:       pipeline-runs.json + scored CSVs + outreach-tracker.csv

Outputs:        ~/.claude/gtm-stack/gtm-dashboard.html

Usage:
    python generate_dashboard.py [--output PATH]
"""

import csv
import json
import os
import sys
import argparse
from datetime import datetime
from collections import defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install",
                           "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import load_workbook

# Import shared utilities
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
STACK_DIR = os.path.dirname(os.path.dirname(SCRIPTS_DIR))
sys.path.insert(0, os.path.join(STACK_DIR, "gtm-shared"))
try:
    from gtm_utils import backup_file, sanitize_for_script_embed, atomic_write_json
except ImportError:
    # Fallback
    import shutil
    def backup_file(fp, max_backups=5):
        if not os.path.exists(fp): return None
        bd = os.path.join(os.path.dirname(fp), ".backups"); os.makedirs(bd, exist_ok=True)
        n, e = os.path.splitext(os.path.basename(fp))
        bp = os.path.join(bd, f"{n}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{e}")
        shutil.copy2(fp, bp); return bp
    def sanitize_for_script_embed(s):
        return s.replace("</", "<\\/").replace("<!--", "<\\!--")

GTM_DIR = os.path.expanduser("~/.claude/gtm-stack")
MASTER_PATH = os.path.join(GTM_DIR, "gtm-leads-master.xlsx")
RUNS_PATH = os.path.join(GTM_DIR, "pipeline-runs.json")
TRACKER_PATH = os.path.join(GTM_DIR, "outreach-tracker.csv")
DNC_PATH = os.path.join(GTM_DIR, "do-not-contact.csv")
DEFAULT_OUTPUT = os.path.join(GTM_DIR, "gtm-dashboard.html")

TIER_RENAME = {
    "Hot": "Strong Fit", "Warm": "Moderate Fit",
    "Cool": "Low Fit", "Pass": "No Fit",
    "Pass (Quick)": "No Fit (Quick)",
}


def normalize_tier(t):
    if not t:
        return "Unscored"
    return TIER_RENAME.get(t.strip(), t.strip())


def load_json(path, default=None):
    if default is None:
        default = []
    if not os.path.exists(path):
        return default
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return default


def load_csv_rows(path):
    if not os.path.exists(path):
        return []
    try:
        with open(path, encoding="utf-8") as f:
            return list(csv.DictReader(f))
    except (IOError, csv.Error):
        return []


def load_from_master():
    """Load all data from master XLSX. Returns (companies, outreach, runs) or None."""
    if not os.path.exists(MASTER_PATH):
        return None

    try:
        wb = load_workbook(MASTER_PATH, read_only=True, data_only=True)
    except Exception:
        return None

    companies = []
    if "All Companies" in wb.sheetnames:
        ws = wb["All Companies"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
            for row in rows[1:]:
                record = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        record[headers[i]] = str(val) if val is not None else ""
                companies.append(record)

    outreach = []
    if "Outreach Log" in wb.sheetnames:
        ws = wb["Outreach Log"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
            for row in rows[1:]:
                record = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        record[headers[i]] = str(val) if val is not None else ""
                outreach.append(record)

    wb.close()

    # Runs come from JSON (more reliable than XLSX for nested data)
    runs = load_json(RUNS_PATH)

    return companies, outreach, runs


def load_from_raw_files():
    """Fallback: load from individual CSVs + JSON."""
    runs = load_json(RUNS_PATH)
    outreach = load_csv_rows(TRACKER_PATH)

    companies = []
    for run in runs:
        csv_path = os.path.expanduser(run.get("csv_path", ""))
        if not csv_path or not os.path.exists(csv_path):
            continue
        rows = load_csv_rows(csv_path)
        for row in rows:
            score = 0
            try:
                score = int(row.get("Fit_Score", 0))
            except (ValueError, TypeError):
                pass
            companies.append({
                "Company": row.get("Company", row.get("Name", "")),
                "Fit_Score": str(score),
                "Fit_Tier": normalize_tier(row.get("Fit_Tier", "")),
                "Location": row.get("Location", ""),
                "Est_Employees": row.get("Est_Employees", ""),
                "DM_Name": row.get("DM_Name", ""),
                "DM_Title": row.get("DM_Title", ""),
                "DM_LinkedIn": row.get("DM_LinkedIn", ""),
                "Website": row.get("Website_Found", ""),
                "Fit_Reasoning": row.get("Fit_Reasoning", ""),
                "Top_Fit_Signal": row.get("Top_Fit_Signal", ""),
                "Source": run.get("source", ""),
                "Contacted": "No",
                "Replied": "No",
                "Meeting": "No",
                "Outcome": "",
            })

    return companies, outreach, runs


def sanitize_for_json(s):
    """Escape for safe embedding in JS."""
    if not s:
        return ""
    return (s.replace("\\", "\\\\")
             .replace('"', '\\"')
             .replace("\n", "\\n")
             .replace("\r", "")
             .replace("\t", " "))


def build_json_data(companies, outreach, runs):
    """Build clean JSON strings for embedding in HTML."""
    # Normalize tier names in runs
    for run in runs:
        for old, new in [("hot", "strong_fit"), ("warm", "moderate_fit"),
                         ("cool", "low_fit"), ("pass", "no_fit")]:
            if old in run and new not in run:
                run[new] = run.pop(old)

    # Normalize tier names in outreach
    for o in outreach:
        if "Fit_Tier" in o:
            o["Fit_Tier"] = normalize_tier(o["Fit_Tier"])

    # Normalize tier names in companies
    for c in companies:
        c["Fit_Tier"] = normalize_tier(c.get("Fit_Tier", ""))

    return (
        sanitize_for_script_embed(json.dumps(runs, ensure_ascii=False)),
        sanitize_for_script_embed(json.dumps(companies, ensure_ascii=False)),
        sanitize_for_script_embed(json.dumps(outreach, ensure_ascii=False)),
    )


def build_react_json(companies, outreach, runs):
    """
    Build a JSON data file with field names matching the React dashboard's expectations.

    Field mapping:
        Python/CSV (from xlsx)     →  React (JS convention)
        ─────────────────────────────────────────────────────
        Company                    →  company
        Fit_Score                  →  score (as int)
        Fit_Tier                   →  tier
        Est_Employees              →  employees
        DM_Name                    →  dm_name
        DM_Title                   →  dm_title
        DM_LinkedIn                →  dm_linkedin
        Website                    →  website
        Fit_Reasoning              →  reasoning
        Top_Fit_Signal             →  top_signal
        Location                   →  location
        Contacted                  →  contacted (as bool)
        Replied                    →  replied (as bool)
        Meeting                    →  meeting (as bool)
        Source                     →  source (mapped to run_id)

        Contact_Name               →  contact
        Channel                    →  channel
        Status                     →  status
        Email_Subject              →  email_subject
        Email_Body                 →  email_body
        LinkedIn_DM                →  linkedin_dm
        Follow_Up_Date             →  follow_up_date
        Follow_Up_Status           →  follow_up_status
        Notes                      →  notes
    """

    def to_bool(val):
        return str(val).strip().lower() in ("yes", "true", "1", "done", "sent")

    def to_int(val, default=0):
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return default

    # Map companies to React format
    react_companies = []
    for c in companies:
        react_companies.append({
            "company": c.get("Company", ""),
            "score": to_int(c.get("Fit_Score", 0)),
            "tier": normalize_tier(c.get("Fit_Tier", "")),
            "employees": c.get("Est_Employees", ""),
            "location": c.get("Location", ""),
            "dm_name": c.get("DM_Name", ""),
            "dm_title": c.get("DM_Title", ""),
            "dm_linkedin": c.get("DM_LinkedIn", ""),
            "website": c.get("Website", ""),
            "reasoning": c.get("Fit_Reasoning", ""),
            "top_signal": c.get("Top_Fit_Signal", ""),
            "contacted": to_bool(c.get("Contacted", "No")),
            "replied": to_bool(c.get("Replied", "No")),
            "meeting": to_bool(c.get("Meeting", "No")),
            "run_id": c.get("Source", ""),
        })

    # Map outreach to React format
    react_outreach = []
    for o in outreach:
        react_outreach.append({
            "date": o.get("Date", ""),
            "company": o.get("Company", ""),
            "contact": o.get("Contact_Name", ""),
            "title": o.get("Title", ""),
            "email": o.get("Email", ""),
            "linkedin": o.get("LinkedIn_URL", ""),
            "channel": o.get("Channel", ""),
            "status": o.get("Status", ""),
            "follow_up_date": o.get("Follow_Up_Date", ""),
            "follow_up_status": o.get("Follow_Up_Status", ""),
            "notes": o.get("Notes", ""),
            "email_subject": o.get("Email_Subject", ""),
            "email_body": o.get("Email_Body", ""),
            "linkedin_dm": o.get("LinkedIn_DM", ""),
            "score": to_int(o.get("Fit_Score", 0)),
            "tier": normalize_tier(o.get("Fit_Tier", "")),
        })

    # Runs are already in a compatible format from pipeline-runs.json
    # Just ensure tier field names are normalized
    react_runs = []
    for run in runs:
        r = dict(run)
        for old, new in [("hot", "strong_fit"), ("warm", "moderate_fit"),
                         ("cool", "low_fit"), ("pass", "no_fit")]:
            if old in r and new not in r:
                r[new] = r.pop(old)
        react_runs.append(r)

    return {
        "pipeline_runs": react_runs,
        "scored_companies": react_companies,
        "outreach_log": react_outreach,
        "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "generated_by": "generate_dashboard.py",
    }


def generate_html(runs_json, companies_json, outreach_json, today, output_path):
    """Write the standalone HTML dashboard with Lumif.ai brand design."""

    # Load logo
    import base64 as _b64
    logo_path = os.path.expanduser("~/.claude/brand-assets/lumifai-logo-sm.b64")
    logo_b64 = ""
    if os.path.exists(logo_path):
        with open(logo_path) as _f:
            logo_b64 = _f.read().strip()

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>GTM Command Center — {today}</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500;700&display=swap" rel="stylesheet">
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Inter',ui-sans-serif,system-ui,-apple-system,sans-serif;background:#FAFAFA;color:#121212;min-height:100vh}}
.m{{font-family:'JetBrains Mono',ui-monospace,monospace}}
button{{font-family:inherit;cursor:pointer;border:none;background:none}}
button:hover{{opacity:0.85}}

.topbar{{background:#FFF;border-bottom:1px solid #E5E7EB;padding:16px 28px;box-shadow:0 1px 2px rgba(0,0,0,0.04)}}
.topbar-inner{{max-width:1120px;margin:0 auto;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px}}

.container{{max-width:1120px;margin:0 auto;padding:0 28px 60px}}

.tabs{{display:flex;gap:0;border-bottom:1px solid #E5E7EB;margin:0 0 28px;overflow-x:auto}}
.tab{{padding:14px 18px;font-size:13px;border-bottom:2px solid transparent;color:#6B7280;display:flex;align-items:center;gap:6px;margin-bottom:-1px;white-space:nowrap;transition:all 0.15s ease}}
.tab:hover{{color:#121212}}
.tab.active{{font-weight:700;color:#121212;border-bottom-color:#E94D35}}
.tab-count{{background:#E94D35;color:white;font-size:10px;font-weight:700;padding:1px 6px;border-radius:8px}}

.slabel{{font-size:11px;font-weight:600;color:#6B7280;letter-spacing:0.1em;text-transform:uppercase;margin-bottom:16px}}

.stats{{display:grid;grid-template-columns:repeat(auto-fit,minmax(155px,1fr));gap:10px;margin-bottom:32px}}
.stat{{background:#FFF;border:1px solid #E5E7EB;border-radius:12px;padding:16px 18px;position:relative;transition:all 0.2s ease}}
.stat:hover{{border-color:rgba(233,77,53,0.2);box-shadow:0 4px 12px rgba(0,0,0,0.04)}}

.funnel-row{{display:flex;align-items:center;gap:12px;margin-bottom:6px}}

.card{{background:#FFF;border:1px solid #E5E7EB;border-radius:12px;overflow:hidden;margin-bottom:8px;cursor:pointer;transition:all 0.2s ease}}
.card:hover{{border-color:rgba(233,77,53,0.2);box-shadow:0 4px 16px rgba(0,0,0,0.06)}}

.badge{{font-size:9px;font-weight:700;padding:2px 7px;border-radius:9999px;display:inline-block}}

.msg-block{{background:#F9FAFB;border:1px solid #E5E7EB;border-radius:8px;padding:12px 16px;margin-top:8px}}

.filters{{display:flex;gap:6px;margin-bottom:16px;flex-wrap:wrap}}
.fbtn{{padding:5px 14px;font-size:11px;border:1px solid #E5E7EB;border-radius:9999px;color:#6B7280;transition:all 0.15s ease}}
.fbtn:hover{{border-color:#D1D5DB;color:#121212}}
.fbtn.active{{background:rgba(233,77,53,0.08);color:#E94D35;border-color:#E94D35;font-weight:700}}

.action-card{{border-radius:12px;padding:14px 18px;display:flex;gap:14px;align-items:flex-start;margin-bottom:8px}}

.tbl-header{{display:grid;padding:10px 16px;background:#F9FAFB;border:1px solid #E5E7EB;border-radius:12px 12px 0 0;align-items:center}}
.tbl-row{{display:grid;padding:11px 16px;border:1px solid #E5E7EB;border-top:none;cursor:pointer;align-items:center;transition:background 0.1s ease}}
.tbl-row:hover{{background:#F9FAFB}}
.tbl-expand{{background:#F9FAFB;border:1px solid #D1D5DB;border-top:none;padding:16px 20px}}

.empty{{text-align:center;padding:60px 20px}}
.footer{{text-align:center;padding:40px 0 0;font-size:11px;color:#9CA3AF}}

@media print{{.tabs,.filters,.action-badge{{display:none}}.tbl-expand{{display:block!important}}}}
</style>
</head>
<body>

<div class="topbar">
<div class="topbar-inner">
  <div style="display:flex;align-items:center;gap:14px">
    {"<img src='data:image/png;base64," + logo_b64 + "' alt='Lumif.ai' style='height:28px'>" if logo_b64 else "<span style='font-size:16px;font-weight:700;color:#E94D35'>Lumif.ai</span>"}
    <div style="width:1px;height:24px;background:#E5E7EB;margin:0 2px"></div>
    <div>
      <div class="m" style="font-size:10px;font-weight:600;color:#E94D35;letter-spacing:0.12em;text-transform:uppercase">GTM Stack</div>
      <div style="font-size:15px;font-weight:700;color:#121212;letter-spacing:-0.02em">Command Center</div>
    </div>
  </div>
  <div style="display:flex;align-items:center;gap:16px">
    <button class="action-badge" id="actionBadge" style="display:none;background:#E94D35;color:white;padding:6px 16px;border-radius:9999px;font-size:12px;font-weight:700;box-shadow:0 2px 8px rgba(233,77,53,0.2)" onclick="switchTab('actions')"></button>
    <span class="m" style="font-size:12px;color:#6B7280">{today}</span>
    <span class="m" style="font-size:9px;color:#9CA3AF;margin-left:4px">updated {datetime.now().strftime('%H:%M')}</span>
  </div>
</div>
</div>

<div class="container">
  <div class="tabs" id="tabBar"></div>
  <div id="content"></div>
  <div class="footer m" id="footerText"></div>
</div>

<script>
const RUNS={runs_json};
const COMPANIES={companies_json};
const OUTREACH={outreach_json};
const TODAY="{today}";

const TIER={{
  "Strong Fit":{{bar:"#E94D35",bg:"rgba(233,77,53,0.06)",bd:"rgba(233,77,53,0.15)"}},
  "Moderate Fit":{{bar:"#F59E0B",bg:"rgba(245,158,11,0.06)",bd:"rgba(245,158,11,0.15)"}},
  "Low Fit":{{bar:"#3B82F6",bg:"rgba(59,130,246,0.06)",bd:"rgba(59,130,246,0.15)"}},
  "No Fit":{{bar:"#9CA3AF",bg:"#F3F4F6",bd:"#E5E7EB"}},
  "No Fit (Quick)":{{bar:"#9CA3AF",bg:"#F3F4F6",bd:"#E5E7EB"}},
  "Unscored":{{bar:"#9CA3AF",bg:"#F3F4F6",bd:"#E5E7EB"}},
}};

const pct=(n,d)=>d>0?Math.round(n/d*100):0;
const parseD=d=>new Date(d+"T00:00:00");
const dAgo=d=>Math.floor((parseD(TODAY)-parseD(d))/86400000);
const overdue=d=>d&&d<=TODAY;
const esc=s=>s?s.replace(/</g,"&lt;").replace(/>/g,"&gt;"):"";
const int=v=>parseInt(v)||0;

// Computed stats
const totalPeople=RUNS.reduce((s,r)=>s+(r.people_scraped||0),0);
const strong=COMPANIES.filter(c=>c.Fit_Tier==="Strong Fit");
const moderate=COMPANIES.filter(c=>c.Fit_Tier==="Moderate Fit");
const low=COMPANIES.filter(c=>c.Fit_Tier==="Low Fit");
const nofit=COMPANIES.filter(c=>(c.Fit_Tier||"").startsWith("No Fit"));
const contacted=COMPANIES.filter(c=>c.Contacted==="Yes");
const replied=COMPANIES.filter(c=>c.Replied==="Yes");
const meetings=COMPANIES.filter(c=>c.Meeting==="Yes");
const sentMsgs=OUTREACH.filter(o=>o.Status==="Sent");
const failedMsgs=OUTREACH.filter(o=>o.Status==="Failed");

// Actions
const fuMap=new Map();
OUTREACH.filter(o=>(o.Follow_Up_Status||"").trim()==="Pending"&&overdue(o.Follow_Up_Date))
  .forEach(o=>{{const k=(o.Contact_Name||"")+(o.Company||"");if(!fuMap.has(k))fuMap.set(k,o)}});
const actions=[];
COMPANIES.filter(c=>c.Fit_Tier==="Strong Fit"&&c.Contacted!=="Yes").forEach(c=>actions.push({{t:"sf",p:0,d:c}}));
[...fuMap.values()].forEach(o=>actions.push({{t:"fu",p:1,d:o}}));
failedMsgs.forEach(o=>actions.push({{t:"fail",p:2,d:o}}));
COMPANIES.filter(c=>c.Fit_Tier==="Moderate Fit"&&c.Contacted!=="Yes").forEach(c=>actions.push({{t:"mf",p:3,d:c}}));
actions.sort((a,b)=>a.p-b.p);

// People (grouped by contact+company)
const peopleMap=new Map();
OUTREACH.forEach(o=>{{
  const k=(o.Contact_Name||"")+"||"+(o.Company||"");
  if(!peopleMap.has(k)){{
    const co=COMPANIES.find(c=>(c.Company||"").toLowerCase()===(o.Company||"").toLowerCase());
    peopleMap.set(k,{{contact:o.Contact_Name||"",title:o.Contact_Title||"",company:o.Company||"",
      email:o.Contact_Email||"",linkedin:o.Contact_LinkedIn||"",
      score:int(o.Fit_Score||(co&&co.Fit_Score)||0),
      tier:o.Fit_Tier||(co&&co.Fit_Tier)||"",location:co&&co.Location||"",
      employees:co&&co.Est_Employees||"",reasoning:co&&co.Fit_Reasoning||"",
      firstDate:o.Date||"",messages:[],hasReply:false,hasMeeting:false,
      followUpDue:false,status:"Sent",outcome:""}});
  }}
  const p=peopleMap.get(k);
  p.messages.push(o);
  if(o.Date&&o.Date<p.firstDate)p.firstDate=o.Date;
  const out=(o.Outcome||"").toLowerCase();const notes=(o.Notes||"").toLowerCase();
  if(out.includes("interested")||out.includes("replied")||notes.match(/replied|booked|demo|meeting|interested|pilot/))p.hasReply=true;
  if(out.includes("meeting")||out.includes("booked")||notes.match(/booked|demo|meeting|pilot/))p.hasMeeting=true;
  if(o.Status==="Failed")p.status="Failed";
  if((o.Follow_Up_Status||"").trim()==="Pending"&&overdue(o.Follow_Up_Date))p.followUpDue=true;
  if(o.Outcome&&!p.outcome)p.outcome=o.Outcome;
  if(!p.outcome&&o.Notes&&notes.match(/replied|booked/))p.outcome=o.Notes;
}});
const people=[...peopleMap.values()];

// Badge
if(actions.length){{const b=document.getElementById("actionBadge");b.textContent=actions.length+" action"+(actions.length>1?"s":"");b.style.display="inline-block";}}
document.getElementById("footerText").textContent=`Lumif.ai GTM Stack · ${{RUNS.length}} runs · ${{COMPANIES.length}} companies · ${{OUTREACH.length}} messages · ${{people.length}} contacts`;

// Tab system
let activeTab="overview",runFilter=null,tierFilter="all",expanded=null,pSort="date",pDir="desc";

function switchTab(t){{activeTab=t;if(t!=="companies")runFilter=null;render();}}

function render(){{
  document.getElementById("tabBar").innerHTML=[
    {{id:"overview",label:"Overview"}},{{id:"actions",label:"Actions",count:actions.length}},
    {{id:"people",label:"People"}},{{id:"pipeline",label:"Runs"}},
    {{id:"companies",label:"Companies"}}
  ].map(t=>`<button class="tab ${{t.id===activeTab?"active":""}}" onclick="switchTab('${{t.id}}')">${{t.label}}${{t.count?` <span class="tab-count m">${{t.count}}</span>`:""}}
  </button>`).join("");

  const el=document.getElementById("content");
  if(activeTab==="overview")renderOverview(el);
  else if(activeTab==="actions")renderActions(el);
  else if(activeTab==="people")renderPeople(el);
  else if(activeTab==="pipeline")renderPipeline(el);
  else if(activeTab==="companies")renderCompanies(el);
}}

function tierBadge(tier){{
  const t=TIER[tier]||TIER["Unscored"];
  return `<span class="badge" style="background:${{t.bg}};color:${{t.bar}};border:1px solid ${{t.bd}}">${{esc(tier)}}</span>`;
}}

function statusBadge(c){{
  if(c.Meeting==="Yes"||c.hasMeeting)return `<span class="badge" style="background:rgba(34,197,94,0.1);color:#16A34A">Meeting</span>`;
  if(c.Replied==="Yes"||c.hasReply)return `<span class="badge" style="background:rgba(168,85,247,0.1);color:#7C3AED">Replied</span>`;
  if(c.Contacted==="Yes"||(c.status&&c.status!=="Failed"))return `<span class="badge" style="background:rgba(59,130,246,0.1);color:#2563EB">Sent</span>`;
  if(c.status==="Failed")return `<span class="badge" style="background:rgba(239,68,68,0.1);color:#DC2626">Failed</span>`;
  return `<span class="badge" style="background:#F3F4F6;color:#9CA3AF">Not sent</span>`;
}}

function scoreDot(score,tier){{
  const t=TIER[tier]||TIER["Unscored"];
  return `<div style="width:34px;height:34px;border-radius:50%;background:#FFF;border:2px solid ${{t.bar}};display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:700;color:${{t.bar}};flex-shrink:0;box-shadow:0 1px 3px rgba(0,0,0,0.06)" class="m">${{score}}</div>`;
}}

function msgBlock(label,icon,content,subject){{
  if(!content)return "";
  return `<div class="msg-block">
    <div class="m" style="font-size:10px;font-weight:600;color:#6B7280;letter-spacing:0.08em;text-transform:uppercase;margin-bottom:8px">${{icon}} ${{label}}</div>
    ${{subject?`<div style="font-size:12px;font-weight:600;color:#374151;margin-bottom:6px">Subject: ${{esc(subject)}}</div>`:""}}
    <div style="font-size:12px;color:#6B7280;line-height:1.6;white-space:pre-wrap">${{esc(content)}}</div>
  </div>`;
}}

// ═══ OVERVIEW ═══
function renderOverview(el){{
  const funnel=[
    {{l:"People Scraped",v:totalPeople,c:"#6B7280"}},
    {{l:"Companies",v:COMPANIES.length,c:"#374151"}},
    {{l:"Strong + Moderate",v:strong.length+moderate.length,c:"#F59E0B"}},
    {{l:"Contacted",v:contacted.length,c:"#3B82F6"}},
    {{l:"People Reached",v:people.length,c:"#60A5FA"}},
    {{l:"Replied",v:replied.length,c:"#7C3AED"}},
    {{l:"Meetings",v:meetings.length,c:"#16A34A"}},
  ];
  const mx=funnel[0].v||1;

  let h=`<div style="margin-bottom:32px"><div class="slabel m">Pipeline Funnel</div>`;
  funnel.forEach((s,i)=>{{
    h+=`<div class="funnel-row">
      <div style="width:130px;font-size:12px;color:#6B7280;text-align:right;flex-shrink:0">${{s.l}}</div>
      <div style="flex:1;height:28px;background:#F3F4F6;border-radius:6px;overflow:hidden">
        <div style="width:${{Math.max(pct(s.v,mx),2)}}%;height:100%;background:${{s.c}};border-radius:6px;opacity:0.85"></div>
      </div>
      <div class="m" style="width:50px;font-size:14px;font-weight:700;color:${{s.c}};text-align:right">${{s.v}}</div>
      <div class="m" style="width:44px;font-size:10px;color:#9CA3AF;text-align:right">${{i>0?pct(s.v,funnel[i-1].v)+"%":""}}</div>
    </div>`;
  }});
  h+=`</div>`;

  const stats=[
    {{l:"Pipeline Runs",v:RUNS.length,s:totalPeople+" people scraped",c:"#6B7280"}},
    {{l:"Companies Scored",v:COMPANIES.length,s:strong.length+" strong · "+moderate.length+" moderate",c:"#F59E0B"}},
    {{l:"Contacted",v:contacted.length,s:people.length+" people · "+sentMsgs.length+" messages",c:"#3B82F6"}},
    {{l:"Reply Rate",v:pct(replied.length,contacted.length)+"%",s:replied.length+" of "+contacted.length,c:"#7C3AED"}},
    {{l:"Meetings",v:meetings.length,s:pct(meetings.length,contacted.length)+"% conversion",c:"#16A34A"}},
    {{l:"Needs Action",v:actions.length,s:actions.length?"tap Actions tab":"all clear",c:actions.length?"#E94D35":"#D1D5DB"}},
  ];
  h+=`<div class="stats">`;
  stats.forEach(s=>{{h+=`<div class="stat"><div style="position:absolute;top:0;left:0;width:3px;height:100%;background:${{s.c}};border-radius:12px 0 0 12px"></div><div class="m" style="font-size:26px;font-weight:700;color:${{s.c}};line-height:1">${{s.v}}</div><div style="font-size:12px;font-weight:600;color:#121212;margin-top:6px">${{s.l}}</div><div class="m" style="font-size:11px;color:#9CA3AF;margin-top:2px">${{s.s}}</div></div>`}});
  h+=`</div>`;

  // Distribution
  h+=`<div style="margin-bottom:32px"><div class="slabel m">Fit Distribution</div><div style="display:flex;gap:8px">`;
  [{{t:"Strong Fit",n:strong.length}},{{t:"Moderate Fit",n:moderate.length}},{{t:"Low Fit",n:low.length}},{{t:"No Fit",n:nofit.length}}].forEach(d=>{{
    const ts=TIER[d.t];
    h+=`<div style="flex:${{Math.max(pct(d.n,COMPANIES.length),8)}};background:#FFF;border:1px solid #E5E7EB;border-radius:12px;padding:14px 16px;text-align:center;position:relative;overflow:hidden;min-width:60px;cursor:pointer;transition:all 0.2s ease" onclick="tierFilter='${{d.t}}';switchTab('companies')" onmouseover="this.style.borderColor='${{ts.bar}}'" onmouseout="this.style.borderColor='#E5E7EB'">
      <div style="position:absolute;bottom:0;left:0;right:0;height:3px;background:${{ts.bar}}"></div>
      <div class="m" style="font-size:22px;font-weight:700;color:${{ts.bar}}">${{d.n}}</div>
      <div style="font-size:11px;color:#6B7280;margin-top:2px">${{d.t}}</div>
    </div>`;
  }});
  h+=`</div></div>`;

  // Recent runs
  h+=`<div><div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:14px"><div class="slabel m" style="margin:0">Recent Runs</div><button style="font-size:12px;color:#E94D35;font-weight:500" onclick="switchTab('pipeline')">View all →</button></div>`;
  RUNS.slice(0,3).forEach(r=>{{
    const sf=r.strong_fit||r.hot||0;
    h+=`<div class="card" onclick="runFilter='${{r.id}}';switchTab('companies')" style="padding:14px 18px;display:flex;align-items:center;gap:16px">
      <div class="m" style="width:40px;height:40px;border-radius:50%;background:#F0FDF4;border:1px solid #BBF7D0;display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:700;color:#16A34A;flex-shrink:0">${{sf}}</div>
      <div style="flex:1;min-width:0"><div style="font-size:13px;font-weight:600;color:#121212">${{esc(r.source||"")}}</div>
      <div class="m" style="font-size:11px;color:#9CA3AF;margin-top:2px">${{r.date||""}} · ${{r.people_scraped||0}} people → ${{r.unique_companies||0}} co</div></div>
      <div style="display:flex;gap:12px;flex-shrink:0">
        <div style="text-align:center"><div class="m" style="font-size:14px;font-weight:700;color:#3B82F6">${{r.contacted||0}}</div><div style="font-size:9px;color:#9CA3AF">sent</div></div>
        <div style="text-align:center"><div class="m" style="font-size:14px;font-weight:700;color:#7C3AED">${{r.replied||0}}</div><div style="font-size:9px;color:#9CA3AF">replied</div></div>
        <div style="text-align:center"><div class="m" style="font-size:14px;font-weight:700;color:#16A34A">${{r.meetings||0}}</div><div style="font-size:9px;color:#9CA3AF">mtg</div></div>
      </div>
    </div>`;
  }});
  h+=`</div>`;
  el.innerHTML=h;
}}

// ═══ ACTIONS ═══
function renderActions(el){{
  if(!actions.length){{el.innerHTML='<div class="empty"><div style="font-size:40px;opacity:0.3;margin-bottom:12px">✓</div><div style="font-size:16px;font-weight:600;color:#6B7280">All clear</div></div>';return;}}
  const cfg={{sf:{{icon:"🔥",bg:"rgba(233,77,53,0.05)",bd:"rgba(233,77,53,0.15)",txt:"Strong Fit — not yet contacted"}},fu:{{icon:"⏰",bg:"rgba(245,158,11,0.05)",bd:"rgba(245,158,11,0.15)",txt:"Follow-up overdue"}},fail:{{icon:"⚠",bg:"rgba(239,68,68,0.05)",bd:"rgba(239,68,68,0.15)",txt:"Send failed"}},mf:{{icon:"☀",bg:"rgba(245,158,11,0.04)",bd:"rgba(245,158,11,0.12)",txt:"Moderate Fit — not yet contacted"}}}};
  el.innerHTML=actions.map(a=>{{
    const c=cfg[a.t];const d=a.d;const isCo=a.t==="sf"||a.t==="mf";
    let detail=c.txt;
    if(a.t==="fu")detail=`Follow-up overdue ${{dAgo(d.Follow_Up_Date)}}d`;
    if(a.t==="fail")detail=`Failed: ${{esc(d.Failure_Reason||d.Notes||"unknown")}}`;
    return `<div class="action-card" style="background:${{c.bg}};border:1px solid ${{c.bd}}">
      <div style="font-size:18px;margin-top:2px;flex-shrink:0">${{c.icon}}</div>
      <div style="flex:1;min-width:0">
        <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap">
          <span style="font-size:14px;font-weight:700;color:#121212">${{esc(d.Company||"")}}</span>
          ${{isCo?`<span class="m" style="font-size:12px;font-weight:700;color:${{(TIER[d.Fit_Tier]||TIER["Unscored"]).bar}}">${{d.Fit_Score||""}}</span>`:""}}
        </div>
        <div style="font-size:12px;color:#6B7280;margin-top:3px">${{isCo?esc((d.DM_Name||"")+(d.DM_Title?", "+d.DM_Title:"")+(d.Location?" · "+d.Location:"")):esc((d.Contact_Name||"")+" · "+(d.Channel||""))}}</div>
        <div style="font-size:12px;color:#9CA3AF;margin-top:4px">${{esc(detail)}}</div>
      </div>
    </div>`;
  }}).join("");
}}

// ═══ PEOPLE ═══
function renderPeople(el){{
  const cols="38px 1.2fr 1fr 52px 82px 76px 1.3fr";
  let sorted=[...people];
  const dir=pDir==="desc"?-1:1;
  sorted.sort((a,b)=>{{
    if(pSort==="date")return dir*(a.firstDate>b.firstDate?1:-1);
    if(pSort==="score")return dir*(a.score-b.score);
    if(pSort==="company")return dir*a.company.localeCompare(b.company);
    if(pSort==="status"){{const sa=a.hasMeeting?0:a.hasReply?1:a.followUpDue?2:3;const sb=b.hasMeeting?0:b.hasReply?1:b.followUpDue?2:3;return dir*(sa-sb);}}
    return 0;
  }});

  let h=`<div class="m" style="font-size:11px;color:#9CA3AF;margin-bottom:16px">${{people.length}} contacts reached · click any row to see full emails + DMs</div>`;
  h+=`<div style="overflow-x:auto"><div class="tbl-header m" style="grid-template-columns:${{cols}};min-width:700px">`;
  [{{}},{{c:"contact",l:"Contact"}},{{c:"company",l:"Company"}},{{c:"score",l:"Score"}},{{c:"status",l:"Status"}},{{c:"date",l:"Sent"}},{{l:"Subject / Channel"}}].forEach(h2=>{{
    h+=`<div style="font-size:10px;font-weight:600;color:#6B7280;letter-spacing:0.06em;text-transform:uppercase;cursor:${{h2.c?"pointer":"default"}};user-select:none" ${{h2.c?`onclick="pSort=pSort==='${{h2.c}}'?pSort:('${{h2.c}}');pDir=pSort==='${{h2.c}}'&&pDir==='desc'?'asc':'desc';render()"`:""}}>${{h2.l||""}}</div>`;
  }});
  h+=`</div>`;

  sorted.forEach((p,i)=>{{
    const k=p.contact+"||"+p.company;
    const isExp=expanded===k;
    const chs=[...new Set(p.messages.map(m=>m.Channel||""))];
    const subj=(p.messages.find(m=>m.Email_Subject)||{{}}).Email_Subject||"";
    const note=p.outcome||"";
    const ts=TIER[p.tier]||TIER["Unscored"];

    let stBadge;
    if(p.hasMeeting)stBadge=`<span class="badge" style="background:rgba(34,197,94,0.1);color:#16A34A">Meeting</span>`;
    else if(p.hasReply)stBadge=`<span class="badge" style="background:rgba(168,85,247,0.1);color:#7C3AED">Replied</span>`;
    else if(p.followUpDue)stBadge=`<span class="badge" style="background:rgba(239,68,68,0.1);color:#DC2626">Follow up</span>`;
    else if(p.status==="Failed")stBadge=`<span class="badge" style="background:rgba(239,68,68,0.1);color:#DC2626">Failed</span>`;
    else stBadge=`<span class="badge" style="background:rgba(59,130,246,0.1);color:#2563EB">Sent</span>`;

    h+=`<div class="tbl-row" style="grid-template-columns:${{cols}};min-width:700px;background:${{isExp?"#F3F4F6":i%2===0?"#FFF":"#FAFAFA"}}" onclick="expanded=expanded==='${{k}}'?null:'${{k}}';render()">
      ${{scoreDot(p.score,p.tier)}}
      <div style="min-width:0"><div style="font-size:13px;font-weight:600;color:#121212;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${{esc(p.contact)}}</div><div style="font-size:11px;color:#9CA3AF;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${{esc(p.title)}}</div></div>
      <div style="min-width:0"><div style="font-size:12px;color:#374151;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${{esc(p.company)}}</div><div class="m" style="font-size:10px;color:#D1D5DB">${{esc(p.location)}}</div></div>
      <div style="text-align:center"><span class="m" style="font-size:12px;font-weight:700;color:${{ts.bar}}">${{p.score}}</span></div>
      <div>${{stBadge}}</div>
      <div class="m" style="font-size:11px;color:#9CA3AF">${{esc(p.firstDate)}}</div>
      <div style="min-width:0"><div style="font-size:11px;color:#6B7280;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${{esc(subj)}}</div>
        <div style="display:flex;gap:4px;margin-top:3px">${{chs.map(ch=>`<span class="m" style="font-size:9px;background:#F3F4F6;color:#6B7280;padding:1px 6px;border-radius:9999px">${{ch==="linkedin"?"in":"✉"}} ${{esc(ch)}}</span>`).join("")}}</div>
      </div>
    </div>`;

    if(isExp){{
      h+=`<div class="tbl-expand">`;
      h+=`<div style="display:flex;gap:20px;margin-bottom:14px;flex-wrap:wrap;font-size:11px;color:#6B7280">`;
      if(p.email)h+=`<span>✉ <span class="m" style="color:#374151">${{esc(p.email)}}</span></span>`;
      if(p.linkedin)h+=`<span>in <span class="m" style="color:#374151">${{esc(p.linkedin)}}</span></span>`;
      if(p.employees)h+=`<span>👥 ${{esc(p.employees)}} emp</span>`;
      h+=`</div>`;

      if(p.reasoning){{
        h+=`<div style="background:#FFF;border:1px solid #E5E7EB;border-radius:8px;padding:12px 16px;margin-bottom:12px">
          <div class="m" style="font-size:10px;font-weight:600;color:#6B7280;letter-spacing:0.08em;text-transform:uppercase;margin-bottom:6px">Why ${{esc(p.company)}} scored ${{p.score}}</div>
          <div style="font-size:12px;color:#6B7280;line-height:1.5">${{esc(p.reasoning)}}</div>
        </div>`;
      }}

      if(note){{
        h+=`<div style="background:rgba(34,197,94,0.04);border:1px solid rgba(34,197,94,0.15);border-radius:8px;padding:10px 16px;margin-bottom:12px">
          <div style="font-size:12px;color:#16A34A;font-weight:600">📝 ${{esc(note)}}</div>
        </div>`;
      }}

      h+=`<div class="m" style="font-size:10px;font-weight:600;color:#6B7280;letter-spacing:0.08em;text-transform:uppercase;margin-bottom:10px">Messages Sent (${{p.messages.length}})</div>`;
      p.messages.forEach(msg=>{{
        const stSym=msg.Status==="Sent"?"✓":"✕";
        const stClr=msg.Status==="Sent"?"#16A34A":"#DC2626";
        const stBg=msg.Status==="Sent"?"rgba(34,197,94,0.1)":"rgba(239,68,68,0.1)";
        h+=`<div style="margin-bottom:14px">
          <div style="display:flex;align-items:center;gap:8px;margin-bottom:4px">
            <span style="width:18px;height:18px;border-radius:50%;display:inline-flex;align-items:center;justify-content:center;font-size:10px;font-weight:700;background:${{stBg}};color:${{stClr}}">${{stSym}}</span>
            <span class="m" style="font-size:11px;color:#6B7280">${{esc(msg.Channel||"")}}</span>
            <span style="font-size:11px;color:#D1D5DB">·</span>
            <span class="m" style="font-size:11px;color:#9CA3AF">${{esc(msg.Date||"")}}</span>
            <span style="font-size:11px;color:${{stClr}};font-weight:600">${{esc(msg.Status||"")}}</span>
            ${{(msg.Follow_Up_Status||"").trim()==="Pending"&&msg.Follow_Up_Date?`<span class="m" style="font-size:10px;color:${{overdue(msg.Follow_Up_Date)?"#DC2626":"#9CA3AF"}}">${{overdue(msg.Follow_Up_Date)?"⚠ ":""}}f/u: ${{msg.Follow_Up_Date}}</span>`:""}}
          </div>
          ${{msgBlock("Email","✉",msg.Email_Body||"",msg.Email_Subject||"")}}
          ${{msgBlock("LinkedIn DM","in",msg.LinkedIn_DM||"",null)}}
          ${{msg.Outcome?`<div style="font-size:11px;color:#16A34A;margin-top:6px;font-style:italic">📋 Outcome: ${{esc(msg.Outcome)}}</div>`:""}}
          ${{msg.Notes&&msg.Notes!==msg.Outcome?`<div style="font-size:11px;color:#6B7280;margin-top:4px;font-style:italic">📝 ${{esc(msg.Notes)}}</div>`:""}}
        </div>`;
      }});
      h+=`</div>`;
    }}
  }});
  h+=`</div>`;
  if(!people.length)h+=`<div class="empty" style="font-size:13px;color:#9CA3AF">No outreach sent yet.</div>`;
  el.innerHTML=h;
}}

// ═══ PIPELINE ═══
function renderPipeline(el){{
  el.innerHTML=RUNS.map(r=>{{
    const sf=r.strong_fit||r.hot||0,mf=r.moderate_fit||r.warm||0,lf=r.low_fit||r.cool||0,nf=r.no_fit||r.pass||0;
    return `<div class="card" onclick="runFilter='${{r.id}}';switchTab('companies')">
      <div style="padding:18px 20px;display:flex;gap:16px;align-items:flex-start">
        <div style="flex:1"><div style="font-size:15px;font-weight:700;color:#121212">${{esc(r.source||"")}}</div>
        <div class="m" style="font-size:11px;color:#9CA3AF;margin-top:4px">${{r.date||""}} · ${{esc(r.filters||"")}}</div>
        <div style="font-size:12px;color:#6B7280;margin-top:8px">${{esc(r.csv_path||"")}}</div></div>
        <div class="m" style="font-size:11px;color:#D1D5DB">${{r.duration_min||"?"}}m</div>
      </div>
      <div style="background:#F9FAFB;padding:14px 20px;border-top:1px solid #E5E7EB;display:flex;gap:20px;flex-wrap:wrap">
        ${{[["Scraped",r.people_scraped,"#6B7280"],["Companies",r.unique_companies,"#374151"],["Strong",sf,"#E94D35"],["Moderate",mf,"#F59E0B"],["Low",lf,"#3B82F6"],["No Fit",nf,"#9CA3AF"],["Contacted",r.contacted,"#3B82F6"],["Replied",r.replied,"#7C3AED"],["Meetings",r.meetings,"#16A34A"]].map(s=>`<div style="text-align:center;min-width:48px"><div class="m" style="font-size:16px;font-weight:700;color:${{s[2]}}">${{s[1]||0}}</div><div style="font-size:9px;color:#9CA3AF;margin-top:1px">${{s[0]}}</div></div>`).join("")}}
      </div>
    </div>`;
  }}).join("");
}}

// ═══ COMPANIES ═══
function renderCompanies(el){{
  let filtered=COMPANIES.slice();
  if(tierFilter!=="all")filtered=filtered.filter(c=>c.Fit_Tier===tierFilter);
  if(runFilter){{
    const run=RUNS.find(r=>r.id===runFilter);
    if(run&&run.csv_path){{
      const src=run.source||"";
      filtered=filtered.filter(c=>(c.Source||"").includes(src));
    }}
  }}
  filtered.sort((a,b)=>int(b.Fit_Score)-int(a.Fit_Score));

  let h="";
  if(runFilter){{
    const rn=RUNS.find(r=>r.id===runFilter);
    h+=`<div style="display:flex;align-items:center;gap:8px;margin-bottom:16px"><span style="font-size:12px;color:#6B7280">Showing: <strong style="color:#121212">${{esc(rn?rn.source:"")}}</strong></span>
    <button class="m" style="font-size:11px;color:#E94D35;border:1px solid #E5E7EB;border-radius:9999px;padding:2px 10px" onclick="runFilter=null;render()">× clear</button></div>`;
  }}

  h+=`<div class="filters">`;
  ["all","Strong Fit","Moderate Fit","Low Fit","No Fit"].forEach(t=>{{
    const cnt=t==="all"?filtered.length:filtered.filter(c=>c.Fit_Tier===t).length;
    h+=`<button class="fbtn m ${{tierFilter===t?"active":""}}" onclick="tierFilter='${{t}}';render()">${{t==="all"?"All":t}} (${{cnt}})</button>`;
  }});
  h+=`</div>`;

  const display=tierFilter!=="all"?filtered.filter(c=>c.Fit_Tier===tierFilter):filtered;

  display.forEach(c=>{{
    const cid="co_"+(c.Company||"").replace(/[^a-zA-Z0-9]/g,"_");
    const isExp=expanded===cid;
    const ors=OUTREACH.filter(o=>(o.Company||"").toLowerCase()===(c.Company||"").toLowerCase());
    h+=`<div class="card" onclick="expanded=expanded==='${{cid}}'?null:'${{cid}}';render()">
      <div style="padding:12px 18px;display:flex;align-items:center;gap:14px">
        ${{scoreDot(int(c.Fit_Score),c.Fit_Tier)}}
        <div style="flex:1;min-width:0">
          <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap">
            <span style="font-size:13px;font-weight:700;color:#121212">${{esc(c.Company||"")}}</span>
            ${{statusBadge(c)}}
          </div>
          <div style="font-size:11px;color:#9CA3AF;margin-top:3px">${{esc(c.DM_Name?c.DM_Name+(c.DM_Title?", "+c.DM_Title:""):"No DM")}} · ${{esc(c.Est_Employees||"?")}} emp · ${{esc(c.Location||"")}}</div>
        </div>
        <span style="font-size:12px;color:#D1D5DB;transform:${{isExp?"rotate(180deg)":"none"}};transition:transform 0.2s">▾</span>
      </div>
    </div>`;

    if(isExp){{
      h+=`<div style="border:1px solid #D1D5DB;border-top:none;border-radius:0 0 12px 12px;padding:14px 18px;background:#F9FAFB;margin-top:-8px;margin-bottom:8px">`;
      if(c.Fit_Reasoning){{
        h+=`<div style="font-size:12px;color:#6B7280;line-height:1.5;margin-bottom:12px;padding:10px 14px;background:#FFF;border-radius:8px;border:1px solid #E5E7EB">
          <span class="m" style="font-size:10px;font-weight:600;color:#6B7280">FIT REASONING: </span>${{esc(c.Fit_Reasoning)}}</div>`;
      }}
      if(ors.length){{
        ors.forEach(o=>{{
          const stSym=o.Status==="Sent"?"✓":"✕";const stClr=o.Status==="Sent"?"#16A34A":"#DC2626";const stBg=o.Status==="Sent"?"rgba(34,197,94,0.1)":"rgba(239,68,68,0.1)";
          h+=`<div style="margin-bottom:12px">
            <div class="m" style="display:flex;align-items:center;gap:8px;font-size:11px;margin-bottom:4px">
              <span style="width:18px;height:18px;border-radius:50%;display:inline-flex;align-items:center;justify-content:center;font-size:9px;font-weight:700;background:${{stBg}};color:${{stClr}}">${{stSym}}</span>
              <span style="color:#6B7280">${{esc(o.Channel||"")}}</span><span style="color:#D1D5DB">·</span><span style="color:#9CA3AF">${{esc(o.Date||"")}}</span>
            </div>
            ${{msgBlock("Email","✉",o.Email_Body||"",o.Email_Subject||"")}}
            ${{msgBlock("LinkedIn DM","in",o.LinkedIn_DM||"",null)}}
            ${{o.Outcome?`<div style="font-size:11px;color:#16A34A;margin-top:4px;font-style:italic">📋 ${{esc(o.Outcome)}}</div>`:""}}
            ${{o.Notes&&o.Notes!==o.Outcome?`<div style="font-size:11px;color:#6B7280;margin-top:4px;font-style:italic">📝 ${{esc(o.Notes)}}</div>`:""}}
          </div>`;
        }});
      }} else h+=`<div style="font-size:12px;color:#9CA3AF">No outreach sent yet.</div>`;
      h+=`</div>`;
    }}
  }});
  el.innerHTML=h;
}}

// Initial render
render();
</script>
</body>
</html>"""

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    return output_path



def main():
    parser = argparse.ArgumentParser(description="Generate GTM Command Center dashboard")
    parser.add_argument("--output", default=None, help="Output HTML path")
    parser.add_argument("--data-dir", default=None,
                        help="Override GTM data directory (default: ~/.claude/gtm-stack)")
    args = parser.parse_args()

    # Allow overriding the data directory (used by tests and custom setups)
    data_dir = args.data_dir or GTM_DIR
    master_path = os.path.join(data_dir, "gtm-leads-master.xlsx")
    runs_path = os.path.join(data_dir, "pipeline-runs.json")
    tracker_path = os.path.join(data_dir, "outreach-tracker.csv")
    output = args.output or os.path.join(data_dir, "gtm-dashboard.html")

    today = datetime.now().strftime("%Y-%m-%d")

    # Try master XLSX first (using overridden paths)
    result = None
    if os.path.exists(master_path):
        try:
            wb = load_workbook(master_path, read_only=True, data_only=True)
            companies_list = []
            if "All Companies" in wb.sheetnames:
                ws = wb["All Companies"]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
                    for row in rows[1:]:
                        record = {}
                        for i, val in enumerate(row):
                            if i < len(headers):
                                record[headers[i]] = str(val) if val is not None else ""
                        companies_list.append(record)
            outreach_list = []
            if "Outreach Log" in wb.sheetnames:
                ws = wb["Outreach Log"]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) > 1:
                    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
                    for row in rows[1:]:
                        record = {}
                        for i, val in enumerate(row):
                            if i < len(headers):
                                record[headers[i]] = str(val) if val is not None else ""
                        outreach_list.append(record)
            wb.close()
            runs_list = load_json(runs_path)
            result = (companies_list, outreach_list, runs_list)
            print(f"📊 Reading from master workbook: {master_path}")
        except Exception:
            pass

    if not result:
        runs_list = load_json(runs_path)
        outreach_list = load_csv_rows(tracker_path)
        companies_list = []
        for run in runs_list:
            csv_path = os.path.expanduser(run.get("csv_path", ""))
            if not csv_path or not os.path.exists(csv_path):
                continue
            rows = load_csv_rows(csv_path)
            for row in rows:
                score = 0
                try:
                    score = int(row.get("Fit_Score", 0))
                except (ValueError, TypeError):
                    pass
                companies_list.append({
                    "Company": row.get("Company", row.get("Name", "")),
                    "Fit_Score": str(score),
                    "Fit_Tier": normalize_tier(row.get("Fit_Tier", "")),
                    "Location": row.get("Location", ""),
                    "Est_Employees": row.get("Est_Employees", ""),
                    "DM_Name": row.get("DM_Name", ""),
                    "DM_Title": row.get("DM_Title", ""),
                    "DM_LinkedIn": row.get("DM_LinkedIn", ""),
                    "Website": row.get("Website_Found", ""),
                    "Fit_Reasoning": row.get("Fit_Reasoning", ""),
                    "Top_Fit_Signal": row.get("Top_Fit_Signal", ""),
                    "Source": run.get("source", ""),
                    "Contacted": "No", "Replied": "No", "Meeting": "No", "Outcome": "",
                })
        result = (companies_list, outreach_list, runs_list)
        print(f"📊 Reading from raw files (no master workbook yet)")

    companies, outreach, runs = result

    runs_json, companies_json, outreach_json = build_json_data(companies, outreach, runs)

    # Build React-compatible JSON data file
    react_data = build_react_json(companies, outreach, runs)
    json_output = os.path.join(os.path.dirname(output), "gtm-dashboard-data.json")
    atomic_write_json(json_output, react_data)
    print(f"📦 React data file: {json_output}")

    # Backup existing dashboard before overwriting
    backup_path = backup_file(output)
    if backup_path:
        print(f"💾 Backed up existing dashboard to: {os.path.basename(backup_path)}")

    path = generate_html(runs_json, companies_json, outreach_json, today, output)

    print(f"✅ Dashboard generated: {path}")
    print(f"   {len(runs)} runs · {len(companies)} companies · {len(outreach)} messages")


if __name__ == "__main__":
    main()
