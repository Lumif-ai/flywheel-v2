#!/usr/bin/env python3
"""
Build the GTM master workbook from all pipeline data.

Reads:
    - ~/.claude/gtm-stack/pipeline-runs.json
    - Scored CSVs referenced in pipeline-runs.json
    - ~/.claude/gtm-stack/outreach-tracker.csv
    - ~/.claude/gtm-stack/do-not-contact.csv

Produces:
    - ~/.claude/gtm-stack/gtm-leads-master.xlsx
      Tab 1: All Companies (deduped, enriched with outreach status)
      Tab 2: Outreach Log
      Tab 3: Pipeline Runs
      Tab 4+: One tab per scrape source (raw data)

Usage:
    python merge_master.py [--output PATH]
"""

import csv
import json
import os
import re
import sys
import argparse
from datetime import datetime
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...", file=sys.stderr)
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install",
                           "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Import shared utilities (backup, atomic writes, company key normalization)
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
STACK_DIR = os.path.dirname(os.path.dirname(SCRIPTS_DIR))
sys.path.insert(0, os.path.join(STACK_DIR, "gtm-shared"))
try:
    from gtm_utils import (backup_file, normalize_company_key,
                           ensure_utf8_csv, atomic_write_json)
except ImportError:
    # Fallback if shared utils not found — inline minimal versions
    import shutil, re, tempfile
    def backup_file(fp, max_backups=5):
        if not os.path.exists(fp): return None
        bd = os.path.join(os.path.dirname(fp), ".backups"); os.makedirs(bd, exist_ok=True)
        n, e = os.path.splitext(os.path.basename(fp))
        bp = os.path.join(bd, f"{n}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{e}")
        shutil.copy2(fp, bp)
        from glob import glob as _glob
        for old in sorted(_glob(os.path.join(bd, f"{n}_*{e}")), reverse=True)[max_backups:]:
            try: os.remove(old)
            except OSError: pass
        return bp
    _SUFFIXES = re.compile(r'\b(inc|llc|ltd|co|corp|corporation|company|group|partners|holdings|enterprises|solutions|services|international|intl)\b\.?', re.IGNORECASE)
    def normalize_company_key(name):
        if not name: return ""
        k = name.strip().lower(); k = _SUFFIXES.sub("", k)
        k = re.sub(r'[^a-z0-9\s]', '', k); k = re.sub(r'\s+', ' ', k).strip()
        return k
    def ensure_utf8_csv(fp): return "utf-8"
    def atomic_write_json(fp, data, indent=2):
        d = os.path.dirname(fp) or "."; os.makedirs(d, exist_ok=True)
        import tempfile as _tf; fd, tmp = _tf.mkstemp(dir=d, suffix=".tmp")
        try:
            with os.fdopen(fd, "w", encoding="utf-8") as f: json.dump(data, f, indent=indent, ensure_ascii=False)
            os.replace(tmp, fp)
        except: os.remove(tmp); raise

GTM_DIR = os.path.expanduser("~/.claude/gtm-stack")
RUNS_PATH = os.path.join(GTM_DIR, "pipeline-runs.json")
TRACKER_PATH = os.path.join(GTM_DIR, "outreach-tracker.csv")
DNC_PATH = os.path.join(GTM_DIR, "do-not-contact.csv")
DEFAULT_OUTPUT = os.path.join(GTM_DIR, "gtm-leads-master.xlsx")

# Tier display config
TIER_CONFIG = {
    "Strong Fit":  {"fill": "FEE2E2", "font_color": "991B1B"},
    "Moderate Fit": {"fill": "FEF3C7", "font_color": "92400E"},
    "Low Fit":     {"fill": "DBEAFE", "font_color": "1E40AF"},
    "No Fit":      {"fill": "F3F4F6", "font_color": "6B7280"},
    "Unscored":    {"fill": "F9FAFB", "font_color": "9CA3AF"},
}

# Map old tier names to new
TIER_RENAME = {
    "Hot": "Strong Fit", "Warm": "Moderate Fit",
    "Cool": "Low Fit", "Pass": "No Fit",
    "Pass (Quick)": "No Fit (Quick)",
}

# Columns for the All Companies tab
COMPANY_COLUMNS = [
    "Company", "Industry", "Fit_Score", "Fit_Tier",
    "Contact_Name", "Contact_Title", "Contact_LinkedIn", "Contact_Email",
    "Location", "Est_Employees", "Website",
    "Top_Fit_Signal", "Top_Concern", "Fit_Reasoning",
    "Contacted", "Outreach_Date", "Outreach_Channel", "Replied", "Meeting", "Outcome",
    "Email_Subject", "Email_Body", "LinkedIn_DM",
    "Source", "Last_Updated",
]


# Industry inference from company name and row context
_UNIVERSITY_KEYWORDS = re.compile(
    r'\b(university|college|institute of technology|MIT|school of|polytechnic|'
    r'community college|state univ|liberal arts)\b', re.IGNORECASE)
_BROKERAGE_KEYWORDS = re.compile(
    r'\b(marsh|gallagher|aon|willis|lockton|hub international|brokerage|'
    r'insurance broker|ajg|brown & brown)\b', re.IGNORECASE)
_CONSTRUCTION_KEYWORDS = re.compile(
    r'\b(construction|builder|contractor|GC|general contract|turner|skanska|'
    r'hensel phelps|AECOM|jacobs|bechtel)\b', re.IGNORECASE)
_ENERGY_KEYWORDS = re.compile(
    r'\b(energy|oil|gas|pipeline|utilities|power gen|solar|wind)\b', re.IGNORECASE)


def _infer_industry(company_name, row=None):
    """Best-effort industry classification from company name and row data."""
    text = company_name
    if row:
        text += " " + " ".join(str(v) for v in row.values() if v)
    if _UNIVERSITY_KEYWORDS.search(text):
        return "Higher Education"
    if _BROKERAGE_KEYWORDS.search(text):
        return "Insurance Brokerage"
    if _CONSTRUCTION_KEYWORDS.search(text):
        return "Construction"
    if _ENERGY_KEYWORDS.search(text):
        return "Energy"
    return ""


def load_json(path, default=None):
    if default is None:
        default = []
    if not os.path.exists(path):
        return default
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return default


def load_csv_rows(path):
    if not os.path.exists(path):
        return []
    try:
        with open(path, encoding="utf-8") as f:
            return list(csv.DictReader(f))
    except (IOError, csv.Error):
        return []


def normalize_tier(tier_str):
    """Convert old tier names to new, or return as-is if already new."""
    if not tier_str:
        return "Unscored"
    tier_str = tier_str.strip()
    return TIER_RENAME.get(tier_str, tier_str)


def load_scored_companies(runs):
    """Read all scored CSVs, normalize tiers, attach source info."""
    companies = {}  # key: company name (lower) -> best record

    for run in runs:
        csv_path = os.path.expanduser(run.get("csv_path", ""))
        if not csv_path or not os.path.exists(csv_path):
            continue

        source_name = run.get("source", os.path.basename(csv_path))
        run_date = run.get("date", "")

        rows = load_csv_rows(csv_path)
        for row in rows:
            # Determine company name: try Company, Organization, Employer, then Name
            company_name = (row.get("Company") or row.get("Organization")
                           or row.get("Employer") or "").strip()
            if not company_name:
                continue

            # Determine contact name: try DM_Name, Contact_Name, then Name
            contact_name = (row.get("DM_Name") or row.get("Contact_Name")
                           or row.get("Name") or "").strip()
            contact_title = (row.get("DM_Title") or row.get("Contact_Title")
                            or row.get("Title") or "").strip()
            contact_linkedin = (row.get("DM_LinkedIn") or row.get("Contact_LinkedIn")
                               or row.get("LinkedIn_URL") or "").strip()
            contact_email = (row.get("Contact_Email") or row.get("Email") or "").strip()

            # Determine industry from row or infer from context
            industry = (row.get("Industry") or "").strip()
            if not industry:
                industry = _infer_industry(company_name, row)

            score = 0
            try:
                score = int(row.get("Fit_Score", 0))
            except (ValueError, TypeError):
                pass

            tier = normalize_tier(row.get("Fit_Tier", ""))
            if tier == "Unscored" and score > 0:
                if score >= 75: tier = "Strong Fit"
                elif score >= 50: tier = "Moderate Fit"
                elif score >= 25: tier = "Low Fit"
                else: tier = "No Fit"

            key = normalize_company_key(company_name)
            record = {
                "Company": company_name,
                "Industry": industry,
                "Fit_Score": score,
                "Fit_Tier": tier,
                "Contact_Name": contact_name,
                "Contact_Title": contact_title,
                "Contact_LinkedIn": contact_linkedin,
                "Contact_Email": contact_email,
                "Location": row.get("Location", ""),
                "Est_Employees": row.get("Est_Employees", ""),
                "Website": row.get("Website_Found", row.get("Website", "")),
                "Top_Fit_Signal": row.get("Top_Fit_Signal", ""),
                "Top_Concern": row.get("Top_Concern", ""),
                "Fit_Reasoning": row.get("Fit_Reasoning", row.get("Reasoning", "")),
                "Email_Subject": row.get("Email_Subject", ""),
                "Email_Body": row.get("Email_Body", ""),
                "LinkedIn_DM": row.get("LinkedIn_DM", ""),
                "Source": source_name,
                "Last_Updated": run_date,
                "Contacted": "No",
                "Outreach_Date": "",
                "Outreach_Channel": "",
                "Replied": "No",
                "Meeting": "No",
                "Outcome": "",
                "_run_id": run.get("id", ""),
                "_raw_row": row,
            }

            # Keep highest score if company appears in multiple runs
            if key in companies:
                existing = companies[key]
                if score > existing["Fit_Score"]:
                    # Merge source names
                    old_sources = existing["Source"]
                    if source_name not in old_sources:
                        record["Source"] = f"{old_sources}, {source_name}"
                    companies[key] = record
                else:
                    # Keep existing but add source
                    if source_name not in existing["Source"]:
                        existing["Source"] += f", {source_name}"
            else:
                companies[key] = record

    return companies


def enrich_with_outreach(companies, outreach_rows):
    """Update companies with outreach status from tracker.
    Also creates company records for outreach-only contacts not in scored CSVs."""
    reply_keywords = {"replied", "booked", "demo", "meeting", "interested", "pilot"}
    meeting_keywords = {"booked", "demo", "meeting", "pilot", "call scheduled"}

    by_company = defaultdict(list)
    for row in outreach_rows:
        company_name = row.get("Company", "").strip()
        # Strip "(retired)" suffix for matching
        clean_name = re.sub(r'\s*\(retired\)\s*$', '', company_name, flags=re.IGNORECASE)
        key = normalize_company_key(clean_name)
        if key:
            by_company[key].append(row)

    for key, records in by_company.items():
        # If company not in scored CSVs, create a stub record from tracker data
        if key not in companies:
            first = records[0]
            company_name = re.sub(r'\s*\(retired\)\s*$', '',
                                  first.get("Company", ""), flags=re.IGNORECASE).strip()
            score = 0
            try:
                score = int(first.get("Fit_Score", 0))
            except (ValueError, TypeError):
                pass
            tier = normalize_tier(first.get("Fit_Tier", ""))
            if tier == "Unscored" and score > 0:
                if score >= 75: tier = "Strong Fit"
                elif score >= 50: tier = "Moderate Fit"
                elif score >= 25: tier = "Low Fit"
                else: tier = "No Fit"

            companies[key] = {
                "Company": company_name,
                "Industry": _infer_industry(company_name, first),
                "Fit_Score": score,
                "Fit_Tier": tier if tier != "Unscored" else (first.get("Fit_Tier", "") or "Unscored"),
                "Contact_Name": first.get("Contact_Name", ""),
                "Contact_Title": first.get("Contact_Title", ""),
                "Contact_LinkedIn": first.get("Contact_LinkedIn", ""),
                "Contact_Email": first.get("Contact_Email", ""),
                "Location": "",
                "Est_Employees": "",
                "Website": "",
                "Top_Fit_Signal": "",
                "Top_Concern": "",
                "Fit_Reasoning": first.get("Notes", ""),
                "Email_Subject": "",
                "Email_Body": "",
                "LinkedIn_DM": "",
                "Source": "Outreach Tracker",
                "Last_Updated": first.get("Date", ""),
                "Contacted": "No",
                "Outreach_Date": "",
                "Outreach_Channel": "",
                "Replied": "No",
                "Meeting": "No",
                "Outcome": "",
            }

        co = companies[key]

        has_sent = any(r.get("Status") == "Sent" for r in records)
        if has_sent:
            co["Contacted"] = "Yes"

        # Populate outreach date and channel from earliest send
        sent_records = [r for r in records if r.get("Status") == "Sent"]
        if sent_records:
            dates = [r.get("Date", "") for r in sent_records if r.get("Date")]
            if dates:
                co["Outreach_Date"] = min(dates)
            channels = set()
            for r in sent_records:
                ch = r.get("Channel", "")
                if ch:
                    channels.add(ch)
            co["Outreach_Channel"] = ", ".join(sorted(channels))

        # Pull contact info from tracker if not already populated
        for r in records:
            if r.get("Contact_Name") and not co.get("Contact_Name"):
                co["Contact_Name"] = r["Contact_Name"]
            if r.get("Contact_Title") and not co.get("Contact_Title"):
                co["Contact_Title"] = r["Contact_Title"]
            if r.get("Contact_LinkedIn") and not co.get("Contact_LinkedIn"):
                co["Contact_LinkedIn"] = r["Contact_LinkedIn"]
            if r.get("Contact_Email") and not co.get("Contact_Email"):
                co["Contact_Email"] = r["Contact_Email"]
            if r.get("Email_Body") and not co.get("Email_Body"):
                co["Email_Body"] = r["Email_Body"]
            if r.get("LinkedIn_DM") and not co.get("LinkedIn_DM"):
                co["LinkedIn_DM"] = r["LinkedIn_DM"]
            if r.get("Subject") and not co.get("Email_Subject"):
                co["Email_Subject"] = r["Subject"]
            if r.get("Email_Subject") and not co.get("Email_Subject"):
                co["Email_Subject"] = r["Email_Subject"]

        # Check for structured Outcome first, then fall back to Notes
        outcomes = [r.get("Outcome", "") for r in records if r.get("Outcome")]
        notes = [r.get("Notes", "") for r in records if r.get("Notes")]

        # Structured outcome takes priority
        for outcome in outcomes:
            ol = outcome.lower()
            if any(kw in ol for kw in meeting_keywords):
                co["Meeting"] = "Yes"
                co["Replied"] = "Yes"
                co["Outcome"] = outcome
            elif "interested" in ol or "replied" in ol:
                co["Replied"] = "Yes"
                if not co["Outcome"]:
                    co["Outcome"] = outcome

        # Fall back to Notes parsing if no structured outcome
        if co["Replied"] == "No":
            for note in notes:
                nl = note.lower()
                if any(kw in nl for kw in reply_keywords):
                    co["Replied"] = "Yes"
                    if not co["Outcome"]:
                        co["Outcome"] = note
                if any(kw in nl for kw in meeting_keywords):
                    co["Meeting"] = "Yes"


def enrich_with_dnc(companies, dnc_rows):
    """Mark companies that are on the Do Not Contact list."""
    dnc_companies = {normalize_company_key(r.get("Company", "")) for r in dnc_rows}
    for key, co in companies.items():
        if key in dnc_companies:
            co["_dnc"] = True


def build_workbook(companies, outreach_rows, runs, output_path):
    """Build the master XLSX workbook."""
    wb = Workbook()

    # Styles
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell_font = Font(name="Calibri", size=10)
    thin_border = Border(
        bottom=Side(style="thin", color="E5E7EB")
    )

    def style_header(ws, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.freeze_panes = "A2"

    def auto_width(ws, max_width=50):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells[:50]:  # Sample first 50 rows
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)

    # ── Tab 1: All Companies ──
    ws = wb.active
    ws.title = "All Companies"

    ws.append(COMPANY_COLUMNS)
    style_header(ws, len(COMPANY_COLUMNS))

    sorted_companies = sorted(
        companies.values(),
        key=lambda c: c["Fit_Score"],
        reverse=True
    )

    for co in sorted_companies:
        row_data = [co.get(col, "") for col in COMPANY_COLUMNS]
        ws.append(row_data)
        row_num = ws.max_row

        # Tier color coding
        tier = co.get("Fit_Tier", "")
        tier_cfg = TIER_CONFIG.get(tier, TIER_CONFIG.get("Unscored"))
        if tier_cfg:
            tier_cell = ws.cell(row=row_num, column=COMPANY_COLUMNS.index("Fit_Tier") + 1)
            tier_cell.fill = PatternFill(
                start_color=tier_cfg["fill"], end_color=tier_cfg["fill"], fill_type="solid"
            )
            tier_cell.font = Font(
                name="Calibri", size=10, bold=True, color=tier_cfg["font_color"]
            )

        # Apply cell font and border
        for col in range(1, len(COMPANY_COLUMNS) + 1):
            cell = ws.cell(row=row_num, column=col)
            if not cell.font.bold:
                cell.font = cell_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col > 10))

    auto_width(ws)

    # ── Tab 2: Outreach Log ──
    if outreach_rows:
        ws2 = wb.create_sheet("Outreach Log")
        outreach_cols = list(outreach_rows[0].keys())
        ws2.append(outreach_cols)
        style_header(ws2, len(outreach_cols))
        for row in outreach_rows:
            ws2.append([row.get(col, "") for col in outreach_cols])
        auto_width(ws2)

    # ── Tab 3: Pipeline Runs ──
    if runs:
        ws3 = wb.create_sheet("Pipeline Runs")
        run_cols = [
            "id", "date", "source", "source_url", "filters",
            "people_scraped", "duplicates_removed", "unique_companies",
            "scored", "strong_fit", "moderate_fit", "low_fit", "no_fit",
            "csv_path", "status", "duration_min"
        ]
        ws3.append(run_cols)
        style_header(ws3, len(run_cols))
        for run in runs:
            # Normalize old field names
            normalized = dict(run)
            for old, new in [("hot", "strong_fit"), ("warm", "moderate_fit"),
                             ("cool", "low_fit"), ("pass", "no_fit")]:
                if old in normalized and new not in normalized:
                    normalized[new] = normalized.pop(old)
            ws3.append([normalized.get(col, "") for col in run_cols])
        auto_width(ws3)

    # ── Tab 4+: Source tabs ──
    source_data = defaultdict(list)
    for run in runs:
        csv_path = os.path.expanduser(run.get("csv_path", ""))
        source_name = run.get("source", "Unknown")
        if not csv_path or not os.path.exists(csv_path):
            continue
        rows = load_csv_rows(csv_path)
        if rows:
            source_data[source_name] = rows

    for source_name, rows in source_data.items():
        # Sanitize sheet name (Excel max 31 chars, no special chars)
        safe_name = source_name[:28].replace("/", "-").replace("\\", "-")
        safe_name = safe_name.replace("[", "").replace("]", "")
        safe_name = safe_name.replace("*", "").replace("?", "")
        safe_name = safe_name.replace(":", "-")

        ws_src = wb.create_sheet(safe_name)
        if rows:
            cols = list(rows[0].keys())
            ws_src.append(cols)
            style_header(ws_src, len(cols))
            for row in rows:
                ws_src.append([row.get(col, "") for col in cols])
            auto_width(ws_src)

    # Save
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Build GTM master workbook")
    parser.add_argument("--output", default=None, help="Output XLSX path")
    parser.add_argument("--data-dir", default=None,
                        help="Override GTM data directory (default: ~/.claude/gtm-stack)")
    args = parser.parse_args()

    # Allow overriding the data directory (used by tests and custom setups)
    data_dir = args.data_dir or GTM_DIR
    runs_path = os.path.join(data_dir, "pipeline-runs.json")
    tracker_path = os.path.join(data_dir, "outreach-tracker.csv")
    dnc_path = os.path.join(data_dir, "do-not-contact.csv")
    output = args.output or os.path.join(data_dir, "gtm-leads-master.xlsx")

    runs = load_json(runs_path)

    # Validate CSV encoding before reading
    for run in runs:
        csv_path = os.path.expanduser(run.get("csv_path", ""))
        if csv_path and os.path.exists(csv_path):
            enc = ensure_utf8_csv(csv_path)
            if enc not in ("utf-8", "utf-8-bom"):
                print(f"   ℹ️  Converted {os.path.basename(csv_path)} from {enc} to UTF-8")

    if os.path.exists(tracker_path):
        ensure_utf8_csv(tracker_path)

    companies = load_scored_companies(runs)
    outreach_rows = load_csv_rows(tracker_path)
    dnc_rows = load_csv_rows(dnc_path)

    enrich_with_outreach(companies, outreach_rows)
    enrich_with_dnc(companies, dnc_rows)

    # Backup existing master workbook before overwriting
    backup_path = backup_file(output)
    if backup_path:
        print(f"💾 Backed up existing master to: {os.path.basename(backup_path)}")

    path = build_workbook(companies, outreach_rows, runs, output)

    co_count = len(companies)
    strong = sum(1 for c in companies.values() if c["Fit_Tier"] == "Strong Fit")
    moderate = sum(1 for c in companies.values() if c["Fit_Tier"] == "Moderate Fit")
    contacted = sum(1 for c in companies.values() if c["Contacted"] == "Yes")

    print(f"✅ Master workbook built: {path}")
    print(f"   {co_count} companies ({strong} strong fit, {moderate} moderate fit)")
    print(f"   {contacted} contacted | {len(outreach_rows)} outreach records")
    print(f"   {len(runs)} pipeline runs | {len(dnc_rows)} on DNC list")


if __name__ == "__main__":
    main()
